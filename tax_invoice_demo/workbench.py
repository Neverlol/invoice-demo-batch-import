from __future__ import annotations

import json
import re
import shutil
import subprocess
from dataclasses import asdict, dataclass, replace
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from werkzeug.datastructures import FileStorage

from .case_events import batch_snapshot, diff_drafts, draft_snapshot, record_case_event
from .coding_library import enrich_invoice_lines, load_formal_coding_library
from .customer_profiles import LineHistoryMatch, apply_line_history_hints, resolve_buyer_from_history, resolve_invoice_record_from_history, seller_default_line_profile
from .extraction_pipeline import compose_parse_source, extract_invoice_structured_data
from .ledger import sync_draft_to_ledger
from .models import BuyerInfo, DraftAttachment, DraftBatch, DraftBatchItem, InvoiceDraft, InvoiceLine
from .ocr import run_optional_ocr
from .parsing import parse_bulk_invoice_lines
from .platform_invoice_screenshots import PlatformInvoiceRequest, extract_platform_invoice_requests
from .source_documents import extract_supported_documents, serialize_document_results
from .tax_rule_engine import write_learned_rules_from_manual_update

WORKBENCH_ROOT = Path(__file__).resolve().parent.parent / "output" / "workbench" / "tax_invoice_demo"
BATCH_LLM_MAX_ATTACHMENTS = 5


@dataclass(frozen=True)
class WorkbookInvoiceUnit:
    source_name: str
    lines: list[InvoiceLine]
    source_excerpt: str


@dataclass(frozen=True)
class HistoryInvoiceRecord:
    invoice_no: str
    seller_name: str
    buyer: BuyerInfo
    invoice_kind: str
    total_amount: str
    is_positive: bool
    status: str
    issued_at: str
    note: str
    lines: list[InvoiceLine]
    source_name: str


@dataclass(frozen=True)
class ReissueDraftUnit:
    source_name: str
    buyer: BuyerInfo
    invoice_kind: str
    note: str
    lines: list[InvoiceLine]
    target_amount: str
    source_excerpt: str


def default_workbench_form() -> dict[str, str]:
    return {
        "company_name": "",
        "raw_text": "",
        "note": "",
    }


def _merge_extracted_note(user_note: str, extracted_note: str) -> str:
    base = (user_note or "").strip()
    extra = (extracted_note or "").strip()
    if not extra:
        return base
    if not base:
        return extra
    if extra in base:
        return base
    return f"{base}\n{extra}"


def _extract_invoice_note_from_context(text: str) -> str:
    """Extract stable invoice remark lines from OCR/text without sending data to LLM.

    Real seed-customer screenshots often say "按该发票对象和备注".  Tesseract may
    read "项目地址" as "顺目地址", so keep this intentionally conservative and only
    extract project name/address style remarks.
    """

    project_name_candidates: list[str] = []
    project_address = ""
    for raw_line in str(text or "").splitlines():
        line = re.sub(r"\s+", " ", raw_line).strip(" |\t")
        if not line:
            continue
        normalized = line.replace("顺目地址", "项目地址").replace("项日地址", "项目地址")
        if "项目名称" in normalized:
            value = _value_after_label(normalized, "项目名称")
            candidate = _cleanup_invoice_note_value(value)
            if candidate:
                project_name_candidates.append(candidate)
        if "项目地址" in normalized and not project_address:
            value = _value_after_label(normalized, "项目地址")
            project_address = _cleanup_invoice_note_value(value)
    project_name = _select_invoice_note_project_name(project_name_candidates)
    parts = []
    if project_name:
        parts.append(f"项目名称:{project_name}")
    if project_address:
        parts.append(f"项目地址:{project_address}")
    return "\n".join(parts)


def _select_invoice_note_project_name(candidates: list[str]) -> str:
    if not candidates:
        return ""

    def score(value: str) -> int:
        compact = re.sub(r"\s+", "", value)
        result = len(compact)
        if re.search(r"项目经理部|项目部|冬运|沈阳市|王家", compact):
            result += 100
        if "公司" in compact:
            result += 20
        if re.search(r"规格型号|单位|数量|单价|金额|税率|编码|需求单位", compact):
            result -= 80
        return result

    return sorted(candidates, key=score, reverse=True)[0]


def _value_after_label(line: str, label: str) -> str:
    _, _, value = line.partition(label)
    value = value.lstrip(" :：|｜")
    return value.strip()


def _cleanup_invoice_note_value(value: str) -> str:
    cleaned = str(value or "").strip().replace("〈", "（").replace("〉", "）")
    cleaned = cleaned.replace("(冬运)", "（冬运）").replace("（冬运)", "（冬运）")
    cleaned = re.sub(r"\s+", " ", cleaned)
    cleaned = re.sub(r"\s+项目", "项目", cleaned)
    if re.search(r"\d+\s*选|[（(]?\d+[）)]?\s*选|URS|ARSE|SRE|RRO|请选择|没有我可以开具的项目", cleaned, re.IGNORECASE):
        return ""
    return cleaned.strip()


def create_draft_from_workbench(
    company_name: str,
    raw_text: str,
    note: str,
    uploaded_files: list[FileStorage],
    *,
    force_batch: bool = False,
) -> InvoiceDraft | DraftBatch:
    draft_id = uuid4().hex[:10]
    case_id = draft_id
    draft_dir = draft_directory(draft_id)
    draft_dir.mkdir(parents=True, exist_ok=True)
    attachments = _save_uploads(draft_dir, uploaded_files)
    document_result = _run_document_extraction(draft_dir, attachments)
    image_attachment_paths = _image_attachment_paths(draft_dir, attachments)
    material_tags = _material_tags_from_context(raw_text, attachments, document_result)
    # 是否批量只看用户是否勾选“批量开具发票”：
    # - 未勾选：所有上传材料都视为同一张发票的材料，图片进入视觉 LLM 识别链路。
    # - 已勾选且图片不超过 5 张：按“一张图一个业务单元”逐张走视觉 LLM。
    # - 已勾选且图片超过 5 张：回到 OCR/规则/待补全混合模式，控制等待时间和模型成本。
    vision_image_paths = [] if force_batch else image_attachment_paths
    batch_vision_enabled = force_batch and 0 < len(image_attachment_paths) <= BATCH_LLM_MAX_ATTACHMENTS
    ocr_result = _run_draft_ocr(
        draft_dir,
        attachments,
        defer_to_vision=bool(vision_image_paths) or batch_vision_enabled,
    )
    early_parse_source = _compose_parse_source(raw_text, document_result.combined_text, ocr_result.combined_text)
    note = _merge_extracted_note(note, _extract_invoice_note_from_context(early_parse_source))
    invoice_profile = _infer_invoice_profile(early_parse_source, note=note)
    reissue_units = _extract_reissue_draft_units(
        draft_dir=draft_dir,
        attachments=attachments,
        instruction_text=_compose_parse_source(raw_text, "", ocr_result.combined_text),
        fallback_context=early_parse_source,
        company_name=company_name,
    )
    if reissue_units:
        return _create_reissue_drafts(
            batch_id=draft_id,
            case_id=case_id,
            draft_dir=draft_dir,
            company_name=company_name,
            raw_text=raw_text,
            note=note,
            attachments=attachments,
            document_result=document_result,
            ocr_result=ocr_result,
            fallback_invoice_profile=invoice_profile,
            units=reissue_units,
        )
    if force_batch:
        platform_ocr_result = ocr_result
        if image_attachment_paths and not platform_ocr_result.combined_text.strip():
            platform_ocr_result = _run_draft_ocr(draft_dir, attachments, defer_to_vision=False)
        platform_ocr_text = platform_ocr_result.combined_text
        platform_history_units = _extract_platform_history_draft_units(
            draft_dir=draft_dir,
            attachments=attachments,
            ocr_text=platform_ocr_text,
            company_name=company_name,
        )
        if not platform_history_units and image_attachment_paths:
            fallback_platform_ocr = _run_platform_screenshot_ocr(draft_dir, attachments)
            if fallback_platform_ocr:
                platform_ocr_text = "\n\n".join(part for part in [platform_ocr_text, fallback_platform_ocr] if part.strip())
                platform_history_units = _extract_platform_history_draft_units(
                    draft_dir=draft_dir,
                    attachments=attachments,
                    ocr_text=platform_ocr_text,
                    company_name=company_name,
                )
                try:
                    platform_ocr_result = replace(platform_ocr_result, combined_text=platform_ocr_text)
                except TypeError:
                    pass
        if platform_history_units:
            return _create_platform_history_drafts(
                batch_id=draft_id,
                case_id=case_id,
                draft_dir=draft_dir,
                company_name=company_name,
                raw_text=raw_text,
                note=note,
                attachments=attachments,
                document_result=document_result,
                ocr_result=platform_ocr_result,
                fallback_invoice_profile=invoice_profile,
                units=platform_history_units,
            )
    if force_batch:
        workbook_units = _extract_workbook_invoice_units(draft_dir, attachments)
        if len(workbook_units) >= 1:
            buyer_extraction = extract_invoice_structured_data(
                raw_text=raw_text,
                note=note,
                document_text=document_result.combined_text,
                ocr_text=ocr_result.combined_text,
                image_paths=image_attachment_paths[:BATCH_LLM_MAX_ATTACHMENTS],
                force_llm_review=bool(image_attachment_paths),
                material_tags=material_tags,
            )
            batch_buyer = _enrich_buyer_from_sheet_context(company_name, buyer_extraction.buyer, buyer_extraction.parse_source)
            batch_buyer = _enrich_buyer_from_history_profile(company_name, batch_buyer, buyer_extraction.parse_source)
            return _create_workbook_draft_batch(
                batch_id=draft_id,
                case_id=case_id,
                draft_dir=draft_dir,
                company_name=company_name,
                raw_text=raw_text,
                note=_merge_extracted_note(note, buyer_extraction.extracted_note),
                buyer=batch_buyer,
                attachments=attachments,
                document_result=document_result,
                ocr_result=ocr_result,
                invoice_profile=invoice_profile,
                workbook_units=workbook_units,
                extract_warnings=buyer_extraction.warnings,
                llm_provider=buyer_extraction.llm_provider,
            )
    platform_requests: list[PlatformInvoiceRequest] = []
    batch_extract_strategy = "platform_screenshot_batch"
    batch_llm_provider = ""
    batch_extract_warnings: list[str] = []
    if force_batch and batch_vision_enabled:
        platform_requests, batch_llm_provider, batch_extract_warnings = _batch_vision_requests_from_uploaded_images(
            draft_dir=draft_dir,
            attachments=attachments,
            raw_text=raw_text,
            note=note,
            document_text=document_result.combined_text,
        )
        if platform_requests:
            platform_requests = _ensure_requests_cover_uploaded_images(platform_requests, attachments)
            batch_extract_strategy = "rules_plus_batch_vision"
    if force_batch and not platform_requests:
        platform_requests = _ensure_requests_cover_uploaded_images(extract_platform_invoice_requests(early_parse_source), attachments)
    if force_batch and platform_requests:
        line_profile = _blank_batch_line_profile() if batch_extract_strategy == "rules_plus_batch_vision" else (seller_default_line_profile(company_name) or _blank_batch_line_profile())
        return _create_platform_screenshot_draft_batch(
            batch_id=draft_id,
            case_id=case_id,
            draft_dir=draft_dir,
            company_name=company_name,
            raw_text=raw_text,
            note=note,
            attachments=attachments,
            document_result=document_result,
            ocr_result=ocr_result,
            invoice_profile=invoice_profile,
            requests=platform_requests,
            line_profile=line_profile,
            extract_strategy=batch_extract_strategy,
            llm_provider=batch_llm_provider,
            extract_warnings=batch_extract_warnings,
        )
    extraction = extract_invoice_structured_data(
        raw_text=raw_text,
        note=note,
        document_text=document_result.combined_text,
        ocr_text=ocr_result.combined_text,
        image_paths=vision_image_paths,
        force_llm_review=False,
        material_tags=material_tags,
    )
    parse_source = extraction.parse_source
    draft_note = _merge_extracted_note(note, extraction.extracted_note)
    buyer = extraction.buyer
    buyer = _enrich_buyer_from_sheet_context(company_name, buyer, parse_source)
    buyer = _enrich_buyer_from_history_profile(company_name, buyer, parse_source)
    lines = _apply_history_profile_to_lines(
        extraction.lines,
        company_name=company_name,
        buyer=buyer,
        parse_source=parse_source,
    )
    invoice_profile = _infer_invoice_profile(parse_source, note=draft_note)
    platform_requests = []
    if force_batch:
        platform_requests = _ensure_requests_cover_uploaded_images(extract_platform_invoice_requests(parse_source), attachments)
    if force_batch and platform_requests:
        line_profile = seller_default_line_profile(company_name) or _blank_batch_line_profile()
        return _create_platform_screenshot_draft_batch(
            batch_id=draft_id,
            case_id=case_id,
            draft_dir=draft_dir,
            company_name=company_name,
            raw_text=raw_text,
            note=note,
            attachments=attachments,
            document_result=document_result,
            ocr_result=ocr_result,
            invoice_profile=invoice_profile,
            requests=platform_requests,
            line_profile=line_profile,
            extract_strategy=extraction.strategy,
            llm_provider=extraction.llm_provider,
            extract_warnings=extraction.warnings,
        )
    split_lines = _build_amount_split_lines(
        company_name=company_name,
        parse_source=parse_source,
        buyer=buyer,
        lines=lines,
        invoice_profile=invoice_profile,
    )
    if force_batch and split_lines:
        batch = _create_split_draft_batch(
            batch_id=draft_id,
            case_id=case_id,
            draft_dir=draft_dir,
            company_name=company_name,
            raw_text=raw_text,
            note=note,
            buyer=buyer,
            attachments=attachments,
            document_result=document_result,
            ocr_result=ocr_result,
            invoice_profile=invoice_profile,
            split_lines=split_lines,
            extract_strategy=extraction.strategy,
            llm_provider=extraction.llm_provider,
            extract_warnings=extraction.warnings,
        )
        return batch

    lines = enrich_invoice_lines(lines, raw_text=parse_source, note=draft_note)

    issues = _build_draft_issues(
        company_name=company_name,
        raw_text=raw_text,
        attachments=attachments,
        buyer=buyer,
        lines=lines,
        special_business=invoice_profile["special_business"],
        document_status=document_result.status,
        document_note=document_result.note,
        ocr_status=ocr_result.status,
        ocr_note=ocr_result.note,
    )

    draft = InvoiceDraft(
        draft_id=draft_id,
        case_id=case_id,
        company_name=company_name.strip(),
        buyer=buyer,
        lines=lines,
        raw_text=raw_text,
        note=draft_note,
        issues=issues,
        source_images=attachments,
        workbook_name="开票明细表.xlsx",
        created_at=datetime.now().isoformat(timespec="seconds"),
        invoice_kind=invoice_profile["invoice_kind"],
        invoice_medium=invoice_profile["invoice_medium"],
        special_business=invoice_profile["special_business"],
        ocr_status=ocr_result.status,
        ocr_engine=ocr_result.engine,
        ocr_text=ocr_result.combined_text,
        ocr_note=ocr_result.note,
        source_doc_status=document_result.status,
        source_doc_text=document_result.combined_text,
        source_doc_note=document_result.note,
        extract_strategy=extraction.strategy,
        llm_provider=extraction.llm_provider,
        extract_warnings=extraction.warnings,
        material_tags=material_tags,
    )
    save_draft(draft)
    record_case_event(
        case_id=draft.case_id,
        draft_id=draft.draft_id,
        event_type="draft_created",
        payload={
            **draft_snapshot(draft),
            "attachment_count": len(attachments),
            "source_doc_status": document_result.status,
            "ocr_status": ocr_result.status,
            "llm_metrics": extraction.llm_metrics,
        },
    )
    return draft


def update_draft_from_form(
    draft_id: str,
    *,
    company_name: str,
    raw_text: str,
    note: str,
    buyer: BuyerInfo,
    lines: list[InvoiceLine],
    invoice_kind: str,
    invoice_medium: str,
    special_business: str,
    uploaded_files: list[FileStorage],
) -> InvoiceDraft:
    existing = load_draft(draft_id)
    if existing is None:
        raise FileNotFoundError(draft_id)

    manual_input_lines = bool(lines)
    has_new_uploads = any((getattr(file, "filename", "") or "").strip() for file in uploaded_files)
    learned_rule_rows = []
    llm_metrics = []
    draft_note = (note or "").strip()

    if manual_input_lines and not has_new_uploads:
        # 保存草稿上的人工编辑时，不重新解析原材料/OCR/调用 LLM；否则“保存修改”会被材料识别链路拖慢。
        attachments = existing.source_images
        parse_source = raw_text or existing.source_doc_text or existing.ocr_text
        inferred_profile = {
            "invoice_kind": existing.invoice_kind or "普通发票",
            "invoice_medium": existing.invoice_medium or "电子发票",
            "special_business": existing.special_business or "",
        }
        resolved_buyer = BuyerInfo(
            name=buyer.name or existing.buyer.name,
            tax_id=buyer.tax_id or existing.buyer.tax_id,
            address=buyer.address or existing.buyer.address,
            phone=buyer.phone or existing.buyer.phone,
            bank_name=buyer.bank_name or existing.buyer.bank_name,
            bank_account=buyer.bank_account or existing.buyer.bank_account,
        )
        resolved_lines = lines
        _mark_manual_coding_changes(resolved_lines, existing.lines)
        learned_rule_rows = write_learned_rules_from_manual_update(
            before_lines=existing.lines,
            after_lines=resolved_lines,
            case_id=existing.case_id or draft_id,
            draft_id=draft_id,
            company_name=company_name,
        )
        document_status = existing.source_doc_status
        document_note = existing.source_doc_note
        document_text = existing.source_doc_text
        ocr_status = existing.ocr_status
        ocr_note = existing.ocr_note
        ocr_text = existing.ocr_text
        ocr_engine = existing.ocr_engine
        extract_strategy = existing.extract_strategy
        llm_provider = existing.llm_provider
        extract_warnings = existing.extract_warnings
        material_tags = existing.material_tags
    else:
        attachments = [*existing.source_images, *_save_uploads(draft_directory(draft_id), uploaded_files)]
        document_result = _run_document_extraction(draft_directory(draft_id), attachments)
        image_attachment_paths = _image_attachment_paths(draft_directory(draft_id), attachments)
        material_tags = _material_tags_from_context(raw_text, attachments, document_result)
        ocr_result = _run_draft_ocr(draft_directory(draft_id), attachments, defer_to_vision=bool(image_attachment_paths))
        extraction = extract_invoice_structured_data(
            raw_text=raw_text,
            note=note,
            document_text=document_result.combined_text,
            ocr_text=ocr_result.combined_text,
            image_paths=image_attachment_paths,
            force_llm_review=False,
            material_tags=material_tags,
        )
        parse_source = extraction.parse_source
        draft_note = _merge_extracted_note(note, extraction.extracted_note)
        inferred_buyer = extraction.buyer
        inferred_buyer = _enrich_buyer_from_sheet_context(company_name, inferred_buyer, parse_source)
        inferred_buyer = _enrich_buyer_from_history_profile(company_name, inferred_buyer, parse_source)
        inferred_lines = _apply_history_profile_to_lines(
            extraction.lines,
            company_name=company_name,
            buyer=inferred_buyer,
            parse_source=parse_source,
        )
        inferred_profile = _infer_invoice_profile(parse_source, note=draft_note)
        resolved_buyer = BuyerInfo(
            name=buyer.name or inferred_buyer.name,
            tax_id=buyer.tax_id or inferred_buyer.tax_id,
            address=buyer.address or inferred_buyer.address,
            phone=buyer.phone or inferred_buyer.phone,
            bank_name=buyer.bank_name or inferred_buyer.bank_name,
            bank_account=buyer.bank_account or inferred_buyer.bank_account,
        )
        resolved_lines = lines or inferred_lines
        if manual_input_lines:
            _mark_manual_coding_changes(resolved_lines, existing.lines)
        resolved_lines = enrich_invoice_lines(
            resolved_lines,
            raw_text=parse_source,
            note=draft_note,
            preserve_existing_tax_rate=manual_input_lines,
        )
        if manual_input_lines:
            learned_rule_rows = write_learned_rules_from_manual_update(
                before_lines=existing.lines,
                after_lines=resolved_lines,
                case_id=existing.case_id or draft_id,
                draft_id=draft_id,
                company_name=company_name,
            )
        document_status = document_result.status
        document_note = document_result.note
        document_text = document_result.combined_text
        ocr_status = ocr_result.status
        ocr_note = ocr_result.note
        ocr_text = ocr_result.combined_text
        ocr_engine = ocr_result.engine
        extract_strategy = extraction.strategy
        llm_provider = extraction.llm_provider
        extract_warnings = extraction.warnings
        llm_metrics = extraction.llm_metrics
    issues = _build_draft_issues(
        company_name=company_name,
        raw_text=raw_text,
        attachments=attachments,
        buyer=resolved_buyer,
        lines=resolved_lines,
        special_business=(special_business or existing.special_business or inferred_profile["special_business"]),
        document_status=document_status,
        document_note=document_note,
        ocr_status=ocr_status,
        ocr_note=ocr_note,
    )

    draft = InvoiceDraft(
        draft_id=draft_id,
        case_id=existing.case_id or draft_id,
        company_name=company_name.strip(),
        buyer=resolved_buyer,
        lines=resolved_lines,
        raw_text=raw_text,
        note=draft_note,
        issues=issues,
        source_images=attachments,
        workbook_name=existing.workbook_name or "开票明细表.xlsx",
        created_at=existing.created_at,
        invoice_kind=invoice_kind or existing.invoice_kind or inferred_profile["invoice_kind"],
        invoice_medium=invoice_medium or existing.invoice_medium or inferred_profile["invoice_medium"],
        special_business=special_business or existing.special_business or inferred_profile["special_business"],
        ocr_status=ocr_status,
        ocr_engine=ocr_engine,
        ocr_text=ocr_text,
        ocr_note=ocr_note,
        source_doc_status=document_status,
        source_doc_text=document_text,
        source_doc_note=document_note,
        extract_strategy=extract_strategy,
        llm_provider=llm_provider,
        extract_warnings=extract_warnings,
        material_tags=material_tags,
    )
    save_draft(draft)
    edit_diffs = diff_drafts(existing, draft)
    record_case_event(
        case_id=draft.case_id,
        draft_id=draft.draft_id,
        event_type="draft_updated",
        payload={
            **draft_snapshot(draft),
            "diff_count": len(edit_diffs),
            "llm_metrics": llm_metrics,
        },
    )
    if edit_diffs:
        record_case_event(
            case_id=draft.case_id,
            draft_id=draft.draft_id,
            event_type="manual_edits_recorded",
            payload={"diffs": edit_diffs},
        )
    if learned_rule_rows:
        record_case_event(
            case_id=draft.case_id,
            draft_id=draft.draft_id,
            event_type="local_learned_rules_saved",
            payload={
                "rule_count": len(learned_rule_rows),
                "rules": learned_rule_rows,
            },
        )
    return draft


def _mark_manual_coding_changes(current_lines: list[InvoiceLine], previous_lines: list[InvoiceLine]) -> None:
    for index, current in enumerate(current_lines):
        if index >= len(previous_lines):
            continue
        previous = previous_lines[index]
        changes = []
        for field_name, label in (
            ("tax_category", "赋码大类"),
            ("tax_code", "税收编码"),
            ("tax_rate", "税率"),
        ):
            old_value = _manual_compare_value(previous, field_name)
            new_value = _manual_compare_value(current, field_name)
            if old_value == new_value:
                continue
            changes.append(f"{label}: {old_value or '空'} -> {new_value or '空'}")
        if not changes:
            continue
        current_reference = current.coding_reference.strip()
        if current_reference.startswith("人工修正赋码"):
            current.coding_reference = "人工修正赋码: " + "；".join(changes)
            continue
        origin = f"；原依据: {current_reference}" if current_reference else ""
        current.coding_reference = "人工修正赋码: " + "；".join(changes) + origin


def _manual_compare_value(line: InvoiceLine, field_name: str) -> str:
    if field_name == "tax_rate":
        return line.normalized_tax_rate()
    return str(getattr(line, field_name, "") or "").strip()


def _build_field_review_reasons(draft: InvoiceDraft) -> dict[str, list[str]]:
    reasons: dict[str, list[str]] = {}

    def add(field: str, message: str) -> None:
        cleaned = str(message or "").strip()
        if not cleaned:
            return
        bucket = reasons.setdefault(field, [])
        if cleaned not in bucket:
            bucket.append(cleaned)

    tags_text = " / ".join(draft.material_tags or [])
    warnings_text = "\n".join(draft.extract_warnings or [])
    has_image_material = any(tag in tags_text for tag in ["群聊截图", "图片材料", "车辆/现场照片"])
    has_ocr_risk = draft.ocr_status not in {"not_requested", "success", ""} or "OCR" in warnings_text

    if not draft.company_name.strip():
        add("销售方", "销售方为空；现场提交前必须确认当前税局登录主体和材料主体一致。")
    if "多个公司" in warnings_text:
        add("销售方", "材料中出现多个公司名；需确认哪个是销售主体，避免把样票买方当成咱们客户。")
        add("购买方名称", "材料中出现多个公司名；需确认购买方没有和销售主体串位。")

    if not draft.buyer.name.strip():
        add("购买方名称", "购买方名称未可靠识别；请从客户开票信息、样票或聊天截图核对。")
    if not draft.buyer.tax_id.strip():
        add("购买方税号", "购买方税号缺失；上传税局前必须补全或确认客户不需要税号。")
    if has_ocr_risk or has_image_material:
        add("购买方税号", "图片/OCR 容易把发票号、二维码数字或车架号误当税号；请重点核对。")

    if not draft.lines:
        add("开票明细", "本地规则未形成明细；需要人工补充或等待智能识别结果。")
        add("金额", "未形成可用金额；请从本次文字、Excel 或聊天截图确认。")
    else:
        missing_project = sum(1 for line in draft.lines if not line.project_name.strip())
        missing_amount = sum(1 for line in draft.lines if not line.resolved_amount_with_tax())
        missing_code = sum(1 for line in draft.lines if not line.tax_category.strip() or not line.tax_code.strip())
        low_conf_lines = [line for line in draft.lines if _line_needs_field_review(line)]
        if missing_project:
            add("开票明细", f"{missing_project} 行缺项目名称；请核对客户材料中的项目/服务内容。")
        if missing_amount:
            add("金额", f"{missing_amount} 行缺含税金额；请核对本次开票金额，不要沿用样票旧金额。")
        if missing_code:
            add("税收编码", f"{missing_code} 行缺税收编码或赋码大类；建议先用一键智能赋码/人工复核。")
        if low_conf_lines:
            add("开票明细", f"{len(low_conf_lines)} 行来自兜底、OCR、异常表或需人工复核来源；请逐行看命中来源。")

    if "样票 PDF" in tags_text or "样票" in warnings_text:
        add("金额", "样票里的金额可能是历史金额；以本次文字/聊天/Excel 明确金额为准。")
    if "多个金额" in warnings_text:
        add("金额", "材料中出现多个金额；需确认哪个是本次开票金额，哪些只是样票或历史金额。")
    if "财务流水/余额线索" in tags_text:
        add("开票明细", "财务流水/余额表只作线索，不应直接转成发票明细。")
        add("金额", "财务流水/余额表金额可能不是本次开票金额；请人工确认。")
    if "压缩包" in tags_text:
        add("附件", "压缩包材料可能只解析了部分文件；请确认附件内是否还有未识别材料。")
    if has_image_material:
        add("附件", "图片/截图已保留为附件；请对照原图核对系统识别出的字段。")

    return {field: messages[:4] for field, messages in reasons.items()}



def _line_needs_field_review(line: InvoiceLine) -> bool:
    reference = f"{line.coding_reference} {line.project_name} {line.specification}"
    markers = ["兜底", "异常", "需人工复核", "未命中", "OCR", "视觉", "样票", "历史", "智能赋码调用失败", "智能赋码返回结果未命中"]
    return any(marker in reference for marker in markers)



def save_draft(draft: InvoiceDraft) -> None:
    draft.field_review_reasons = _build_field_review_reasons(draft)
    draft_dir = draft_directory(draft.draft_id)
    draft_dir.mkdir(parents=True, exist_ok=True)
    payload = {
        "draft_id": draft.draft_id,
        "case_id": draft.case_id,
        "company_name": draft.company_name,
        "buyer": asdict(draft.buyer),
        "lines": [asdict(line) for line in draft.lines],
        "raw_text": draft.raw_text,
        "note": draft.note,
        "issues": draft.issues,
        "source_images": [asdict(image) for image in draft.source_images],
        "workbook_name": draft.workbook_name,
        "created_at": draft.created_at,
        "invoice_kind": draft.invoice_kind,
        "invoice_medium": draft.invoice_medium,
        "special_business": draft.special_business,
        "ocr_status": draft.ocr_status,
        "ocr_engine": draft.ocr_engine,
        "ocr_text": draft.ocr_text,
        "ocr_note": draft.ocr_note,
        "source_doc_status": draft.source_doc_status,
        "source_doc_text": draft.source_doc_text,
        "source_doc_note": draft.source_doc_note,
        "extract_strategy": draft.extract_strategy,
        "llm_provider": draft.llm_provider,
        "extract_warnings": draft.extract_warnings,
        "material_tags": draft.material_tags,
        "field_review_reasons": draft.field_review_reasons,
    }
    (draft_dir / "draft.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    (draft_dir / "raw_text.txt").write_text(draft.raw_text, encoding="utf-8")
    (draft_dir / "source_docs_text.txt").write_text(draft.source_doc_text, encoding="utf-8")
    (draft_dir / "ocr_text.txt").write_text(draft.ocr_text, encoding="utf-8")
    (draft_dir / "source_combined.txt").write_text(draft.combined_source_text(), encoding="utf-8")
    (draft_dir / "ocr_meta.json").write_text(
        json.dumps(
            {
                "status": draft.ocr_status,
                "engine": draft.ocr_engine,
                "note": draft.ocr_note,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    _write_workbook(draft_dir / draft.workbook_name, draft)
    sync_draft_to_ledger(draft)


def save_draft_batch(batch: DraftBatch) -> None:
    batch_dir = draft_directory(batch.batch_id)
    batch_dir.mkdir(parents=True, exist_ok=True)
    payload = {
        "batch_id": batch.batch_id,
        "case_id": batch.case_id,
        "company_name": batch.company_name,
        "created_at": batch.created_at,
        "items": [asdict(item) for item in batch.items],
        "raw_text": batch.raw_text,
        "note": batch.note,
        "issues": batch.issues,
        "source_images": [asdict(image) for image in batch.source_images],
        "invoice_kind": batch.invoice_kind,
        "invoice_medium": batch.invoice_medium,
        "special_business": batch.special_business,
        "extract_strategy": batch.extract_strategy,
        "llm_provider": batch.llm_provider,
        "extract_warnings": batch.extract_warnings,
        "material_tags": batch.material_tags,
    }
    (batch_dir / "batch.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    (batch_dir / "raw_text.txt").write_text(batch.raw_text, encoding="utf-8")


def load_draft_batch(batch_id: str) -> DraftBatch | None:
    batch_path = draft_directory(batch_id) / "batch.json"
    if not batch_path.exists():
        return None
    payload = json.loads(batch_path.read_text(encoding="utf-8"))
    return DraftBatch(
        batch_id=payload["batch_id"],
        case_id=payload.get("case_id", payload["batch_id"]),
        company_name=payload.get("company_name", ""),
        created_at=payload.get("created_at", ""),
        items=[DraftBatchItem(**item) for item in payload.get("items", [])],
        raw_text=payload.get("raw_text", ""),
        note=payload.get("note", ""),
        issues=payload.get("issues", []),
        source_images=[DraftAttachment(**item) for item in payload.get("source_images", [])],
        invoice_kind=payload.get("invoice_kind", "普通发票"),
        invoice_medium=payload.get("invoice_medium", "电子发票"),
        special_business=payload.get("special_business", ""),
        extract_strategy=payload.get("extract_strategy", "rules_only"),
        llm_provider=payload.get("llm_provider", ""),
        extract_warnings=payload.get("extract_warnings", []),
        material_tags=payload.get("material_tags", []),
    )


def load_draft(draft_id: str) -> InvoiceDraft | None:
    draft_path = draft_directory(draft_id) / "draft.json"
    if not draft_path.exists():
        return None
    payload = json.loads(draft_path.read_text(encoding="utf-8"))
    return InvoiceDraft(
        draft_id=payload["draft_id"],
        case_id=payload.get("case_id", payload["draft_id"]),
        company_name=payload.get("company_name", ""),
        buyer=BuyerInfo(**payload.get("buyer", {})),
        lines=[InvoiceLine(**line) for line in payload.get("lines", [])],
        raw_text=payload.get("raw_text", ""),
        note=payload.get("note", ""),
        issues=payload.get("issues", []),
        source_images=[DraftAttachment(**item) for item in payload.get("source_images", [])],
        workbook_name=payload.get("workbook_name", ""),
        created_at=payload.get("created_at", ""),
        invoice_kind=payload.get("invoice_kind", "普通发票"),
        invoice_medium=payload.get("invoice_medium", "电子发票"),
        special_business=payload.get("special_business", ""),
        ocr_status=payload.get("ocr_status", "not_requested"),
        ocr_engine=payload.get("ocr_engine", ""),
        ocr_text=payload.get("ocr_text", ""),
        ocr_note=payload.get("ocr_note", ""),
        source_doc_status=payload.get("source_doc_status", "not_requested"),
        source_doc_text=payload.get("source_doc_text", ""),
        source_doc_note=payload.get("source_doc_note", ""),
        extract_strategy=payload.get("extract_strategy", "rules_only"),
        llm_provider=payload.get("llm_provider", ""),
        extract_warnings=payload.get("extract_warnings", []),
        material_tags=payload.get("material_tags", []),
        field_review_reasons=payload.get("field_review_reasons", {}),
    )


def draft_directory(draft_id: str) -> Path:
    return WORKBENCH_ROOT / draft_id


def _requests_from_uploaded_images(attachments: list[DraftAttachment]) -> list[PlatformInvoiceRequest]:
    image_attachments = _uploaded_image_attachments(attachments)
    if len(image_attachments) < 2:
        return []
    return [_blank_request_from_attachment(attachment) for attachment in image_attachments]


def _run_platform_screenshot_ocr(draft_dir: Path, attachments: list[DraftAttachment]) -> str:
    command = shutil.which("tesseract") or shutil.which("tesseract.exe")
    if not command:
        return ""
    parts: list[str] = []
    for attachment in _uploaded_image_attachments(attachments):
        image_path = draft_dir / attachment.stored_name
        if not image_path.exists():
            continue
        try:
            completed = subprocess.run(
                [command, str(image_path), "stdout", "-l", "chi_sim+eng", "--psm", "11"],
                capture_output=True,
                timeout=10,
                check=False,
            )
        except Exception:  # noqa: BLE001
            continue
        if completed.returncode != 0:
            continue
        text = _decode_subprocess_text(completed.stdout).strip()
        if text:
            source_name = attachment.original_name or Path(attachment.stored_name).name
            parts.append(f"[{source_name}]\n{text}")
    return "\n\n".join(parts)


def _decode_subprocess_text(value: bytes) -> str:
    for encoding in ("utf-8", "gb18030", "latin-1"):
        try:
            return value.decode(encoding)
        except UnicodeDecodeError:
            continue
    return value.decode("utf-8", errors="ignore")


def _ensure_requests_cover_uploaded_images(
    requests: list[PlatformInvoiceRequest],
    attachments: list[DraftAttachment],
) -> list[PlatformInvoiceRequest]:
    """强制批量模式下，业务单元是上传图片，不是 OCR 成功块。

    OCR/LLM 可能只给 5 张图里的 2 张返回结构化文本；这种情况下也必须保留另外 3 张，
    让它们进入“待补全 / 异常项”，避免现场出现“5 张图只生成 2 张发票”。

    重要：OCR/LLM 返回的块名有时使用保存后的文件名，例如原图 `02.jpg` 保存为
    `uploads/02_02.jpg` 后，LLM 文本块会写成 `[02_02.jpg]`。强制批量时仍必须以
    原始上传图片为唯一业务单元，不能把 `[02_02.jpg]` 和 `02.jpg` 生成两张票。
    """
    image_attachments = _uploaded_image_attachments(attachments)
    if not image_attachments:
        return requests

    remaining = list(requests)
    merged: list[PlatformInvoiceRequest] = []
    for attachment in image_attachments:
        matched = _pop_matching_request_for_attachment(attachment, remaining)
        if matched is None:
            merged.append(_blank_request_from_attachment(attachment))
        else:
            # 对助理展示始终使用原始上传名，避免出现 02_02.jpg / 03_03.jpg 这类内部保存名。
            merged.append(replace(matched, source_name=attachment.original_name or Path(attachment.stored_name).name))
    return merged


def _batch_vision_requests_from_uploaded_images(
    *,
    draft_dir: Path,
    attachments: list[DraftAttachment],
    raw_text: str,
    note: str,
    document_text: str,
) -> tuple[list[PlatformInvoiceRequest], str, list[str]]:
    """少量批量图片按“一张图一个业务单元”逐张走视觉 LLM。

    这里故意不把 3-5 张图片一次性交给模型合并处理，避免模型把不同图片的购买方、税号、金额串单。
    """
    image_attachments = _uploaded_image_attachments(attachments)
    if not image_attachments or len(image_attachments) > BATCH_LLM_MAX_ATTACHMENTS:
        return [], "", []

    from .llm_adapter import get_llm_adapter

    if not get_llm_adapter().is_enabled:
        return [], "", []

    requests: list[PlatformInvoiceRequest] = []
    provider = ""
    warnings: list[str] = []
    for attachment in image_attachments:
        source_name = attachment.original_name or Path(attachment.stored_name).name
        image_path = draft_dir / attachment.stored_name
        extraction = extract_invoice_structured_data(
            raw_text=raw_text,
            note=f"{note}\n来源图片：{source_name}".strip(),
            document_text=document_text,
            ocr_text="",
            image_paths=[image_path],
            force_llm_review=True,
            material_tags=["群聊截图"],
        )
        if extraction.llm_provider:
            provider = extraction.llm_provider
        elif extraction.llm_metrics:
            provider = str(extraction.llm_metrics[-1].get("provider") or provider)
        warnings.extend(f"{source_name}: {warning}" for warning in extraction.warnings)
        line = extraction.lines[0] if extraction.lines else InvoiceLine(project_name="", amount_with_tax="")
        source_excerpt = extraction.parse_source.strip()
        if not source_excerpt:
            source_excerpt = f"来源图片：{source_name}\n视觉大模型已按单张图片尝试提取开票信息。"
        requests.append(
            PlatformInvoiceRequest(
                source_name=source_name,
                buyer=extraction.buyer,
                amount_with_tax=line.resolved_amount_with_tax(),
                source_excerpt=source_excerpt,
                project_name=line.project_name,
                tax_rate=line.normalized_tax_rate() if line.tax_rate else "",
                tax_category=line.tax_category,
                tax_code=line.tax_code,
                specification=line.specification,
                unit=line.unit,
                quantity=line.quantity,
            )
        )
    return requests, provider, warnings



def _uploaded_image_attachments(attachments: list[DraftAttachment]) -> list[DraftAttachment]:
    image_suffixes = {".png", ".jpg", ".jpeg", ".bmp", ".gif", ".webp"}
    return [
        attachment
        for attachment in attachments
        if Path(attachment.stored_name or attachment.original_name).suffix.lower() in image_suffixes
    ]


def _blank_request_from_attachment(attachment: DraftAttachment) -> PlatformInvoiceRequest:
    source_name = attachment.original_name or Path(attachment.stored_name).name
    return PlatformInvoiceRequest(
        source_name=source_name,
        buyer=BuyerInfo(name="", tax_id=""),
        amount_with_tax="",
        source_excerpt=f"来源图片：{source_name}\nOCR 未可靠识别为完整发票信息，请在批量工作表中人工补全。",
    )


def _pop_matching_request_for_attachment(
    attachment: DraftAttachment,
    requests: list[PlatformInvoiceRequest],
) -> PlatformInvoiceRequest | None:
    if not requests:
        return None
    attachment_names = _attachment_source_name_candidates(attachment)
    attachment_keys = {_normalize_source_name(name) for name in attachment_names}
    for index, request in enumerate(requests):
        if _normalize_source_name(request.source_name) in attachment_keys:
            return requests.pop(index)

    attachment_numeric_key = _source_numeric_key(attachment.original_name or Path(attachment.stored_name).name)
    if attachment_numeric_key:
        for index, request in enumerate(requests):
            request_numeric_key = _source_numeric_key(request.source_name)
            if request_numeric_key and request_numeric_key == attachment_numeric_key:
                return requests.pop(index)
    return None


def _attachment_source_name_candidates(attachment: DraftAttachment) -> list[str]:
    names = [attachment.original_name, Path(attachment.stored_name or "").name]
    return [name for name in names if name]


def _normalize_source_name(value: str) -> str:
    return Path(value or "").name.strip().lower()


def _source_numeric_key(value: str) -> str:
    stem = Path(value or "").stem.lower()
    tokens = [token.lstrip("0") or "0" for token in re.findall(r"\d+", stem)]
    return tokens[-1] if tokens else ""


def _blank_batch_line_profile() -> LineHistoryMatch:
    return LineHistoryMatch(
        project_name="",
        tax_category="",
        tax_code="",
        tax_rate="",
        specification="",
        unit="项",
        quantity="1",
        matched_source="批量模式待人工补全",
        confidence="pending_review",
    )


def _batch_line_coding_reference(line_profile: LineHistoryMatch) -> str:
    if not line_profile.project_name and not line_profile.tax_code:
        return "批量模式待人工补全，需人工复核"
    return (
        "销售主体历史开票档案推荐，需人工复核: "
        f"{line_profile.matched_source} -> {line_profile.tax_category or '未记录大类'}"
        f" / {line_profile.tax_code or '未记录编码'}"
    )


def _create_platform_screenshot_draft_batch(
    *,
    batch_id: str,
    case_id: str,
    draft_dir: Path,
    company_name: str,
    raw_text: str,
    note: str,
    attachments: list[DraftAttachment],
    document_result,
    ocr_result,
    invoice_profile: dict[str, str],
    requests: list[PlatformInvoiceRequest],
    line_profile,
    extract_strategy: str = "platform_screenshot_batch",
    llm_provider: str = "",
    extract_warnings: list[str] | None = None,
) -> DraftBatch:
    material_tags = _material_tags_from_context(raw_text, attachments, document_result)
    items: list[DraftBatchItem] = []
    batch_issues: list[str] = []
    for request in requests:
        child_draft_id = uuid4().hex[:10]
        buyer = request.buyer
        request_has_line = bool(request.project_name.strip() or request.amount_with_tax.strip() or request.tax_rate.strip() or request.tax_code.strip())
        line = InvoiceLine(
            project_name=request.project_name or line_profile.project_name,
            amount_with_tax=request.amount_with_tax,
            tax_rate=request.tax_rate or line_profile.tax_rate or ("1%" if line_profile.tax_code else ""),
            tax_category=request.tax_category or line_profile.tax_category,
            tax_code=request.tax_code or line_profile.tax_code,
            specification=request.specification or line_profile.specification,
            unit=request.unit or line_profile.unit or "项",
            quantity=request.quantity or "1",
            coding_reference="视觉大模型识别，需人工复核" if request_has_line else _batch_line_coding_reference(line_profile),
        )
        child_note_parts = [note.strip(), f"来源图片：{request.source_name}"]
        if request.order_no:
            child_note_parts.append(f"订单号：{request.order_no}")
        if request.email:
            child_note_parts.append(f"邮箱：{request.email}")
        child_note = "；".join(part for part in child_note_parts if part)
        if request_has_line:
            line = enrich_invoice_lines(
                [line],
                raw_text=request.source_excerpt or raw_text,
                note=child_note,
                preserve_existing_tax_rate=True,
            )[0]
        child_issues = _build_draft_issues(
            company_name=company_name,
            raw_text=request.source_excerpt or raw_text,
            attachments=attachments,
            buyer=buyer,
            lines=[line],
            special_business=invoice_profile["special_business"],
            document_status=document_result.status,
            document_note=document_result.note,
            ocr_status=ocr_result.status,
            ocr_note=ocr_result.note,
        )
        if not buyer.name.strip():
            child_issues.append(f"{request.source_name} 截图抬头不完整；请人工补全购买方名称。")
        if not buyer.tax_id.strip():
            child_issues.append(f"{request.source_name} 未可靠识别购买方税号；请人工补全。")
        if not request.amount_with_tax.strip():
            child_issues.append(f"{request.source_name} 未可靠识别开票金额；请人工补全。")
        child_draft = InvoiceDraft(
            draft_id=child_draft_id,
            case_id=case_id,
            company_name=company_name.strip(),
            buyer=buyer,
            lines=[line],
            raw_text=request.source_excerpt or raw_text,
            note=child_note,
            issues=child_issues,
            source_images=attachments,
            workbook_name="开票明细表.xlsx",
            created_at=datetime.now().isoformat(timespec="seconds"),
            invoice_kind=invoice_profile["invoice_kind"],
            invoice_medium=invoice_profile["invoice_medium"],
            special_business=invoice_profile["special_business"],
            ocr_status=ocr_result.status,
            ocr_engine=ocr_result.engine,
            ocr_text=request.source_excerpt,
            ocr_note=ocr_result.note,
            source_doc_status=document_result.status,
            source_doc_text=document_result.combined_text,
            source_doc_note=document_result.note,
            extract_strategy=extract_strategy,
            llm_provider=llm_provider,
            extract_warnings=list(extract_warnings or []),
            material_tags=material_tags,
        )
        save_draft(child_draft)
        record_case_event(
            case_id=case_id,
            draft_id=child_draft_id,
            batch_id=batch_id,
            event_type="platform_screenshot_child_draft_created",
            payload=draft_snapshot(child_draft),
        )
        items.append(
            DraftBatchItem(
                draft_id=child_draft_id,
                buyer_name=buyer.name or "待补全购买方名称",
                invoice_kind=invoice_profile["invoice_kind"],
                amount_total=line.resolved_amount_with_tax(),
                project_summary=line.project_name,
                line_count=1,
                issue_summary=child_issues[0] if child_issues else "",
            )
        )
        batch_issues.extend(child_issues)

    batch = DraftBatch(
        batch_id=batch_id,
        case_id=case_id,
        company_name=company_name.strip(),
        created_at=datetime.now().isoformat(timespec="seconds"),
        items=items,
        raw_text=raw_text,
        note=note.strip(),
        issues=batch_issues,
        source_images=attachments,
        invoice_kind=invoice_profile["invoice_kind"],
        invoice_medium=invoice_profile["invoice_medium"],
        special_business=invoice_profile["special_business"],
        extract_strategy=extract_strategy,
        llm_provider=llm_provider,
        extract_warnings=list(extract_warnings or []),
        material_tags=material_tags,
    )
    save_draft_batch(batch)
    record_case_event(
        case_id=case_id,
        batch_id=batch_id,
        event_type="platform_screenshot_draft_batch_created",
        payload=batch_snapshot(batch),
    )
    return batch



def _create_workbook_draft_batch(
    *,
    batch_id: str,
    case_id: str,
    draft_dir: Path,
    company_name: str,
    raw_text: str,
    note: str,
    buyer: BuyerInfo,
    attachments: list[DraftAttachment],
    document_result,
    ocr_result,
    invoice_profile: dict[str, str],
    workbook_units: list[WorkbookInvoiceUnit],
    extract_warnings: list[str],
    llm_provider: str,
) -> DraftBatch:
    material_tags = _material_tags_from_context(raw_text, attachments, document_result)
    batch_issues = [
        "当前材料命中了“Excel 明细 -> 草稿”规则；系统已按每个 Excel 文件生成一张待复核草稿。",
        "每个 Excel 文件对应一张草稿，Excel 内多行会作为这张发票的多行明细；请确认客户是否确实要求按这些表分别开票。",
        "Excel 明细由本地表格规则解析，并会优先参考客户历史档案赋码；发票图片/补充文字只作为购买方、票种、税点和备注参考。",
    ]
    if not buyer.name.strip():
        batch_issues.append("未可靠识别购买方名称；请在批量草稿中补全后再执行。")
    if not buyer.tax_id.strip():
        batch_issues.append("未可靠识别购买方税号；请在批量草稿中补全后再执行。")

    items: list[DraftBatchItem] = []
    for unit in workbook_units:
        child_draft_id = uuid4().hex[:10]
        child_dir = draft_directory(child_draft_id)
        child_dir.mkdir(parents=True, exist_ok=True)
        child_attachments = _clone_attachments_to_directory(
            source_dir=draft_dir,
            target_dir=child_dir,
            attachments=attachments,
        )
        parse_source = f"{raw_text}\n{unit.source_excerpt}"
        history_lines = _apply_history_profile_to_lines(
            unit.lines,
            company_name=company_name,
            buyer=buyer,
            parse_source=parse_source,
        )
        enriched_lines = enrich_invoice_lines(
            history_lines,
            raw_text=parse_source,
            note=note,
            preserve_existing_tax_rate=True,
        )
        child_issues = _build_draft_issues(
            company_name=company_name,
            raw_text=unit.source_excerpt or raw_text,
            attachments=child_attachments,
            buyer=buyer,
            lines=enriched_lines,
            special_business=invoice_profile["special_business"],
            document_status=document_result.status,
            document_note=document_result.note,
            ocr_status=ocr_result.status,
            ocr_note=ocr_result.note,
        )
        child_issues.insert(0, f"本草稿由 Excel 明细 `{unit.source_name}` 自动生成；请复核该表是否对应一张发票。")
        child_draft = InvoiceDraft(
            draft_id=child_draft_id,
            case_id=case_id,
            company_name=company_name.strip(),
            buyer=buyer,
            lines=enriched_lines,
            raw_text=unit.source_excerpt or raw_text,
            note="；".join(part for part in [note.strip(), f"来源 Excel：{unit.source_name}"] if part),
            issues=child_issues,
            source_images=child_attachments,
            workbook_name="开票明细表.xlsx",
            created_at=datetime.now().isoformat(timespec="seconds"),
            invoice_kind=invoice_profile["invoice_kind"],
            invoice_medium=invoice_profile["invoice_medium"],
            special_business=invoice_profile["special_business"],
            ocr_status=ocr_result.status,
            ocr_engine=ocr_result.engine,
            ocr_text=ocr_result.combined_text,
            ocr_note=ocr_result.note,
            source_doc_status=document_result.status,
            source_doc_text=unit.source_excerpt,
            source_doc_note=document_result.note,
            extract_strategy="rules_plus_workbook_batch",
            llm_provider=llm_provider,
            extract_warnings=list(extract_warnings),
            material_tags=material_tags,
        )
        save_draft(child_draft)
        record_case_event(
            case_id=case_id,
            draft_id=child_draft_id,
            batch_id=batch_id,
            event_type="workbook_child_draft_created",
            payload=draft_snapshot(child_draft),
        )
        items.append(
            DraftBatchItem(
                draft_id=child_draft_id,
                buyer_name=buyer.name or "待补全购买方名称",
                invoice_kind=invoice_profile["invoice_kind"],
                amount_total=_sum_line_amounts(enriched_lines),
                project_summary=_summarize_projects(enriched_lines),
                line_count=len(enriched_lines),
                issue_summary=child_issues[0] if child_issues else "",
            )
        )
        batch_issues.extend(child_issues)

    batch = DraftBatch(
        batch_id=batch_id,
        case_id=case_id,
        company_name=company_name.strip(),
        created_at=datetime.now().isoformat(timespec="seconds"),
        items=items,
        raw_text=raw_text,
        note=note.strip(),
        issues=batch_issues,
        source_images=attachments,
        invoice_kind=invoice_profile["invoice_kind"],
        invoice_medium=invoice_profile["invoice_medium"],
        special_business=invoice_profile["special_business"],
        extract_strategy="rules_plus_workbook_batch",
        llm_provider=llm_provider,
        extract_warnings=list(extract_warnings),
        material_tags=material_tags,
    )
    save_draft_batch(batch)
    record_case_event(
        case_id=case_id,
        batch_id=batch_id,
        event_type="workbook_draft_batch_created",
        payload=batch_snapshot(batch),
    )
    return batch


def _create_split_draft_batch(
    *,
    batch_id: str,
    case_id: str,
    draft_dir: Path,
    company_name: str,
    raw_text: str,
    note: str,
    buyer: BuyerInfo,
    attachments: list[DraftAttachment],
    document_result,
    ocr_result,
    invoice_profile: dict[str, str],
    split_lines: list[InvoiceLine],
    extract_strategy: str,
    llm_provider: str,
    extract_warnings: list[str],
) -> DraftBatch:
    material_tags = _material_tags_from_context(raw_text, attachments, document_result)
    batch_issues = [
        "当前材料命中了“一份输入 -> 多张草稿”规则；系统已按金额自动拆成多张待复核草稿。",
        "这类草稿通常来自聊天里只给买方资料、税点和多笔金额。请重点复核每张草稿的项目名称、税率和票种。",
    ]
    items: list[DraftBatchItem] = []

    for line in split_lines:
        child_draft_id = uuid4().hex[:10]
        child_dir = draft_directory(child_draft_id)
        child_dir.mkdir(parents=True, exist_ok=True)
        child_attachments = _clone_attachments_to_directory(
            source_dir=draft_dir,
            target_dir=child_dir,
            attachments=attachments,
        )
        enriched_lines = enrich_invoice_lines([line], raw_text=_compose_parse_source(raw_text, document_result.combined_text, ocr_result.combined_text), note=note)
        child_issues = _build_draft_issues(
            company_name=company_name,
            raw_text=raw_text,
            attachments=child_attachments,
            buyer=buyer,
            lines=enriched_lines,
            special_business=invoice_profile["special_business"],
            document_status=document_result.status,
            document_note=document_result.note,
            ocr_status=ocr_result.status,
            ocr_note=ocr_result.note,
        )
        child_issues.insert(0, "本草稿由金额拆票规则自动生成；请复核项目名称、税率和是否确实需要分两张票开具。")
        child_draft = InvoiceDraft(
            draft_id=child_draft_id,
            case_id=case_id,
            company_name=company_name.strip(),
            buyer=buyer,
            lines=enriched_lines,
            raw_text=raw_text,
            note=note.strip(),
            issues=child_issues,
            source_images=child_attachments,
            workbook_name="开票明细表.xlsx",
            created_at=datetime.now().isoformat(timespec="seconds"),
            invoice_kind=invoice_profile["invoice_kind"],
            invoice_medium=invoice_profile["invoice_medium"],
            special_business=invoice_profile["special_business"],
            ocr_status=ocr_result.status,
            ocr_engine=ocr_result.engine,
            ocr_text=ocr_result.combined_text,
            ocr_note=ocr_result.note,
            source_doc_status=document_result.status,
            source_doc_text=document_result.combined_text,
            source_doc_note=document_result.note,
            extract_strategy=extract_strategy,
            llm_provider=llm_provider,
            extract_warnings=list(extract_warnings),
            material_tags=material_tags,
        )
        save_draft(child_draft)
        record_case_event(
            case_id=case_id,
            draft_id=child_draft_id,
            batch_id=batch_id,
            event_type="split_child_draft_created",
            payload=draft_snapshot(child_draft),
        )
        items.append(
            DraftBatchItem(
                draft_id=child_draft_id,
                buyer_name=buyer.name,
                invoice_kind=invoice_profile["invoice_kind"],
                amount_total=line.resolved_amount_with_tax(),
                project_summary=_summarize_projects(enriched_lines),
                line_count=len(enriched_lines),
                issue_summary=child_issues[0] if child_issues else "",
            )
        )

    batch = DraftBatch(
        batch_id=batch_id,
        case_id=case_id,
        company_name=company_name.strip(),
        created_at=datetime.now().isoformat(timespec="seconds"),
        items=items,
        raw_text=raw_text,
        note=note.strip(),
        issues=batch_issues,
        source_images=attachments,
        invoice_kind=invoice_profile["invoice_kind"],
        invoice_medium=invoice_profile["invoice_medium"],
        special_business=invoice_profile["special_business"],
        extract_strategy=extract_strategy,
        llm_provider=llm_provider,
        extract_warnings=list(extract_warnings),
        material_tags=material_tags,
    )
    save_draft_batch(batch)
    record_case_event(
        case_id=case_id,
        batch_id=batch_id,
        event_type="draft_batch_created",
        payload=batch_snapshot(batch),
    )
    return batch




def _extract_platform_history_draft_units(
    *,
    draft_dir: Path,
    attachments: list[DraftAttachment],
    ocr_text: str,
    company_name: str,
) -> list[ReissueDraftUnit]:
    platform_blocks = _extract_platform_invoice_blocks(ocr_text)
    if not platform_blocks:
        return []
    records: list[HistoryInvoiceRecord] = []
    for attachment in attachments:
        suffix = Path(attachment.stored_name).suffix.lower()
        if suffix not in {".xlsx", ".xls"}:
            continue
        source_path = draft_dir / attachment.stored_name
        try:
            records.extend(_history_invoice_records_from_workbook(source_path, attachment.original_name))
        except Exception:  # noqa: BLE001
            continue
    if not records:
        return _platform_units_from_blocks_with_profile(platform_blocks, company_name=company_name)
    candidates = [record for record in records if record.is_positive and record.lines]
    if not candidates:
        return _platform_units_from_blocks_with_profile(platform_blocks, company_name=company_name)

    units: list[ReissueDraftUnit] = []
    used_invoice_numbers: set[str] = set()
    for block in platform_blocks:
        candidate = _match_platform_block_to_history_record(
            block,
            candidates,
            company_name=company_name,
            used_invoice_numbers=used_invoice_numbers,
        )
        if candidate is None:
            return []
        used_invoice_numbers.add(candidate.invoice_no)
        bill_id = _extract_platform_bill_id(block[1])
        note = candidate.note or bill_id
        units.append(
            ReissueDraftUnit(
                source_name=block[0],
                buyer=candidate.buyer,
                invoice_kind=candidate.invoice_kind,
                note=note,
                lines=[replace(line) for line in candidate.lines],
                target_amount=candidate.total_amount,
                source_excerpt=(
                    f"平台开票截图匹配历史正数发票 {candidate.invoice_no}；"
                    f"来源图片 {block[0]}；含税合计 {candidate.total_amount}。"
                ),
            )
        )
    return units


def _platform_units_from_blocks_with_profile(
    platform_blocks: list[tuple[str, str]],
    *,
    company_name: str,
) -> list[ReissueDraftUnit]:
    units: list[ReissueDraftUnit] = []
    for source_name, text in platform_blocks:
        buyer = BuyerInfo(
            name=_extract_platform_buyer_name(text),
            tax_id=_extract_platform_buyer_tax_id(text),
        )
        amount = _extract_platform_amount(text)
        bill_id = _extract_platform_bill_id(text)
        history_record = resolve_invoice_record_from_history(company_name=company_name, buyer=buyer, bill_id=bill_id, amount_with_tax=amount)
        if history_record is not None:
            units.append(
                ReissueDraftUnit(
                    source_name=source_name,
                    buyer=history_record.buyer,
                    invoice_kind=history_record.invoice_kind,
                    note=history_record.note or bill_id,
                    lines=[replace(line) for line in history_record.lines],
                    target_amount=history_record.amount_with_tax,
                    source_excerpt=(
                        f"平台开票截图匹配客户档案历史发票 {history_record.invoice_no}；"
                        f"来源图片 {source_name}；含税合计 {history_record.amount_with_tax}。"
                    ),
                )
            )
            continue
        line = _platform_line_from_profile(text, amount=amount, company_name=company_name, buyer=buyer)
        units.append(
            ReissueDraftUnit(
                source_name=source_name,
                buyer=buyer,
                invoice_kind=_infer_platform_invoice_kind(text),
                note=bill_id,
                lines=[line],
                target_amount=amount,
                source_excerpt=(
                    f"平台开票截图直接生成草稿；来源图片 {source_name}；"
                    f"含税金额 {amount or '待补全'}；客户档案用于补充常用项目/税码。"
                ),
            )
        )
    return units


def _platform_line_from_profile(
    text: str,
    *,
    amount: str,
    company_name: str,
    buyer: BuyerInfo,
) -> InvoiceLine:
    project_name = _extract_platform_project_name(text) or "服务费"
    tax_rate = _extract_platform_tax_rate(text) or ("6%" if _infer_platform_invoice_kind(text) == "增值税专用发票" else "")
    if _infer_platform_invoice_kind(text) == "增值税专用发票" and tax_rate in {"", "3%", "1%", "免税"} and "6" in text:
        tax_rate = "6%"
    line = InvoiceLine(
        project_name=project_name,
        amount_with_tax=amount,
        tax_rate=tax_rate,
        tax_category="现代服务" if "服务" in project_name or "现代服务" in text else "",
        unit="项",
        quantity="1",
        coding_reference="平台开票截图识别生成，金额/项目需人工复核；客户档案可补充历史税码。",
    )
    history_context = "\n".join(part for part in [project_name, tax_rate] if part)
    enriched = apply_line_history_hints([line], company_name=company_name, buyer=buyer, raw_text=history_context)
    return enriched[0] if enriched else line


def _extract_platform_project_name(text: str) -> str:
    compact = re.sub(r"\s+", "", str(text or ""))
    for matched in re.finditer(r"[\*+?]现代服务[\*\"]?([\u4e00-\u9fffA-Za-z0-9]{2,20}?服务费)", compact):
        value = matched.group(1).strip()
        if "支持服务费" in value:
            return "服务费"
        return value
    if "服务费" in compact:
        return "服务费"
    return ""


def _extract_platform_tax_rate(text: str) -> str:
    compact = re.sub(r"\s+", "", str(text or "")).replace("％", "%")
    for rate in ["6%", "3%", "1%", "免税"]:
        if rate in compact:
            return rate
    return ""


def _infer_platform_invoice_kind(text: str) -> str:
    compact = re.sub(r"\s+", "", str(text or ""))
    if "专票" in compact or "增值税专票" in compact:
        return "增值税专用发票"
    if "普票" in compact or "普通发票" in compact:
        return "普通发票"
    return "增值税专用发票"


def _extract_platform_invoice_blocks(ocr_text: str) -> list[tuple[str, str]]:
    blocks = _split_ocr_image_blocks(ocr_text)
    result: list[tuple[str, str]] = []
    for name, text in blocks:
        compact = re.sub(r"\s+", "", text)
        if not compact:
            continue
        score = 0
        for token in ["购买方信息", "销售方信息", "开票信息", "纳税人识别号", "发票类型", "备注", "账单ID"]:
            if token in compact:
                score += 1
        has_buyer_tax = bool(_extract_platform_buyer_tax_id(text))
        has_buyer_name = bool(_extract_platform_buyer_name(text))
        has_seller_tax = bool(re.search(r"92[0-9A-Z]{16}", compact))
        if score >= 4 and (has_buyer_tax or has_buyer_name) and has_seller_tax:
            result.append((name, text))
    return result


def _split_ocr_image_blocks(ocr_text: str) -> list[tuple[str, str]]:
    matches = list(re.finditer(r"(?m)^\[(?P<name>[^\]]+\.(?:jpg|jpeg|png|webp|bmp))\]\s*$", str(ocr_text or ""), re.IGNORECASE))
    if not matches:
        return []
    blocks: list[tuple[str, str]] = []
    for index, matched in enumerate(matches):
        start = matched.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(ocr_text)
        blocks.append((_display_source_name(matched.group("name")), ocr_text[start:end].strip()))
    return blocks


def _display_source_name(name: str) -> str:
    base = Path(name or "").name
    return re.sub(r"^\d{2}_", "", base)


def _match_platform_block_to_history_record(
    block: tuple[str, str],
    records: list[HistoryInvoiceRecord],
    *,
    company_name: str,
    used_invoice_numbers: set[str],
) -> HistoryInvoiceRecord | None:
    _source_name, text = block
    buyer_tax_id = _extract_platform_buyer_tax_id(text)
    buyer_name = _extract_platform_buyer_name(text)
    amount = _extract_platform_amount(text)
    candidates: list[HistoryInvoiceRecord] = []
    for record in records:
        if record.invoice_no in used_invoice_numbers:
            continue
        if company_name.strip() and record.seller_name.strip() and company_name.strip() not in record.seller_name.strip() and record.seller_name.strip() not in company_name.strip():
            continue
        if buyer_tax_id and record.buyer.tax_id != buyer_tax_id:
            continue
        if buyer_name and record.buyer.name and buyer_name not in record.buyer.name and record.buyer.name not in buyer_name:
            continue
        if amount and _decimal_from_text(record.total_amount) != _decimal_from_text(amount):
            continue
        candidates.append(record)
    if not candidates:
        return None
    candidates.sort(key=lambda item: item.issued_at, reverse=True)
    return candidates[0]


def _extract_platform_buyer_tax_id(text: str) -> str:
    compact = re.sub(r"\s+", "", str(text or "").upper()).replace("Ｏ", "O").replace("Ｉ", "I")
    candidates = [candidate for candidate in re.findall(r"[0-9][0-9A-Z]{17}", compact) if candidate.startswith(("91", "92"))]
    if not candidates:
        return ""
    # 平台截图通常先出现购买方税号，再出现销售方税号；销售方个体工商户常以 92 开头。
    for candidate in candidates:
        if candidate.startswith("91"):
            return candidate[:18]
    return ""


def _extract_platform_buyer_name(text: str) -> str:
    for line in str(text or "").splitlines():
        matched = re.search(r"(北京[^\s|｜>《》]{2,40}(?:网络技术有限公司|有限公司))", line)
        if matched:
            return matched.group(1).strip()
    matched = re.search(r"([\u4e00-\u9fffA-Za-z0-9（）()]{4,60}(?:网络技术有限公司|有限公司))", str(text or ""))
    return matched.group(1).strip() if matched else ""


def _extract_platform_amount(text: str) -> str:
    candidates: list[str] = []
    for matched in re.finditer(r"[¥￥Y]\s*(\d{1,6}(?:\.\d{1,3})?)", str(text or ""), re.IGNORECASE):
        amount = _money_text(matched.group(1))
        if amount:
            candidates.append(amount)
    return candidates[-1] if candidates else ""


def _extract_platform_bill_id(text: str) -> str:
    compact = re.sub(r"\s+", "", str(text or ""))
    matched = re.search(r"(?:账单ID|单ID)[）)]?(20\d{14,18})", compact)
    if matched:
        return _normalize_platform_bill_id(matched.group(1))
    ids = re.findall(r"20\d{14,18}", compact)
    return _normalize_platform_bill_id(ids[-1]) if ids else ""


def _normalize_platform_bill_id(value: str) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    if digits.startswith("20") and len(digits) > 18:
        return digits[:18]
    return digits


def _create_platform_history_drafts(
    *,
    batch_id: str,
    case_id: str,
    draft_dir: Path,
    company_name: str,
    raw_text: str,
    note: str,
    attachments: list[DraftAttachment],
    document_result,
    ocr_result,
    fallback_invoice_profile: dict[str, str],
    units: list[ReissueDraftUnit],
) -> InvoiceDraft | DraftBatch:
    if len(units) == 1:
        return _create_platform_history_single_draft(
            draft_id=batch_id,
            case_id=case_id,
            draft_dir=draft_dir,
            company_name=company_name,
            raw_text=raw_text,
            note=note,
            attachments=attachments,
            document_result=document_result,
            ocr_result=ocr_result,
            fallback_invoice_profile=fallback_invoice_profile,
            unit=units[0],
        )
    material_tags = _material_tags_from_context(raw_text, attachments, document_result)
    items: list[DraftBatchItem] = []
    batch_issues = [
        "当前材料命中“平台开票截图 + 历史导出”规则；系统已按实际平台开票截图生成草稿，群聊说明截图不会单独生成发票。",
        "请重点复核每张草稿的金额、购买方、项目、备注账单 ID、税率和税收编码。",
    ]
    for unit in units:
        child_draft_id = uuid4().hex[:10]
        child_dir = draft_directory(child_draft_id)
        child_dir.mkdir(parents=True, exist_ok=True)
        child_attachments = _clone_attachments_to_directory(source_dir=draft_dir, target_dir=child_dir, attachments=attachments)
        merged_note = _merge_extracted_note(note, unit.note)
        child_issues = _build_draft_issues(
            company_name=company_name,
            raw_text=unit.source_excerpt or raw_text,
            attachments=child_attachments,
            buyer=unit.buyer,
            lines=unit.lines,
            special_business=fallback_invoice_profile.get("special_business", ""),
            document_status=document_result.status,
            document_note=document_result.note,
            ocr_status=ocr_result.status,
            ocr_note=ocr_result.note,
        )
        child_issues.insert(0, f"本草稿由平台截图 `{unit.source_name}` 匹配历史导出生成；请复核截图金额和备注账单 ID。")
        child_draft = InvoiceDraft(
            draft_id=child_draft_id,
            case_id=case_id,
            company_name=company_name.strip(),
            buyer=unit.buyer,
            lines=unit.lines,
            raw_text=unit.source_excerpt or raw_text,
            note=merged_note,
            issues=child_issues,
            source_images=child_attachments,
            workbook_name="开票明细表.xlsx",
            created_at=datetime.now().isoformat(timespec="seconds"),
            invoice_kind=unit.invoice_kind or fallback_invoice_profile.get("invoice_kind", "普通发票"),
            invoice_medium=fallback_invoice_profile.get("invoice_medium", "电子发票"),
            special_business=fallback_invoice_profile.get("special_business", ""),
            ocr_status=ocr_result.status,
            ocr_engine=ocr_result.engine,
            ocr_text=ocr_result.combined_text,
            ocr_note=ocr_result.note,
            source_doc_status=document_result.status,
            source_doc_text=unit.source_excerpt,
            source_doc_note=document_result.note,
            extract_strategy="rules_plus_platform_history_batch",
            extract_warnings=["平台截图金额/账单 ID 来自 OCR 和历史导出匹配，需人工复核。"],
            material_tags=material_tags,
        )
        save_draft(child_draft)
        record_case_event(case_id=case_id, draft_id=child_draft_id, batch_id=batch_id, event_type="platform_history_child_draft_created", payload=draft_snapshot(child_draft))
        items.append(
            DraftBatchItem(
                draft_id=child_draft_id,
                buyer_name=unit.buyer.name or "待补全购买方名称",
                invoice_kind=unit.invoice_kind or fallback_invoice_profile.get("invoice_kind", "普通发票"),
                amount_total=_sum_line_amounts(unit.lines),
                project_summary=_summarize_projects(unit.lines),
                line_count=len(unit.lines),
                issue_summary=child_issues[0] if child_issues else "",
            )
        )
        batch_issues.extend(child_issues)
    batch = DraftBatch(
        batch_id=batch_id,
        case_id=case_id,
        company_name=company_name.strip(),
        created_at=datetime.now().isoformat(timespec="seconds"),
        items=items,
        raw_text=raw_text,
        note=note.strip(),
        issues=batch_issues,
        source_images=attachments,
        invoice_kind=_common_reissue_invoice_kind(units) or fallback_invoice_profile.get("invoice_kind", "普通发票"),
        invoice_medium=fallback_invoice_profile.get("invoice_medium", "电子发票"),
        special_business=fallback_invoice_profile.get("special_business", ""),
        extract_strategy="rules_plus_platform_history_batch",
        llm_provider="",
        extract_warnings=["平台开票截图 + 历史导出匹配生成草稿；请人工复核金额与备注。"],
        material_tags=material_tags,
    )
    save_draft_batch(batch)
    record_case_event(case_id=case_id, batch_id=batch_id, event_type="platform_history_draft_batch_created", payload=batch_snapshot(batch))
    return batch


def _create_platform_history_single_draft(
    *,
    draft_id: str,
    case_id: str,
    draft_dir: Path,
    company_name: str,
    raw_text: str,
    note: str,
    attachments: list[DraftAttachment],
    document_result,
    ocr_result,
    fallback_invoice_profile: dict[str, str],
    unit: ReissueDraftUnit,
) -> InvoiceDraft:
    merged_note = _merge_extracted_note(note, unit.note)
    issues = _build_draft_issues(
        company_name=company_name,
        raw_text=unit.source_excerpt or raw_text,
        attachments=attachments,
        buyer=unit.buyer,
        lines=unit.lines,
        special_business=fallback_invoice_profile.get("special_business", ""),
        document_status=document_result.status,
        document_note=document_result.note,
        ocr_status=ocr_result.status,
        ocr_note=ocr_result.note,
    )
    issues.insert(0, f"本草稿由平台截图 `{unit.source_name}` 匹配历史导出生成；请复核截图金额和备注账单 ID。")
    draft = InvoiceDraft(
        draft_id=draft_id,
        case_id=case_id,
        company_name=company_name.strip(),
        buyer=unit.buyer,
        lines=unit.lines,
        raw_text=unit.source_excerpt or raw_text,
        note=merged_note,
        issues=issues,
        source_images=list(attachments),
        workbook_name="开票明细表.xlsx",
        created_at=datetime.now().isoformat(timespec="seconds"),
        invoice_kind=unit.invoice_kind or fallback_invoice_profile.get("invoice_kind", "普通发票"),
        invoice_medium=fallback_invoice_profile.get("invoice_medium", "电子发票"),
        special_business=fallback_invoice_profile.get("special_business", ""),
        ocr_status=ocr_result.status,
        ocr_engine=ocr_result.engine,
        ocr_text=ocr_result.combined_text,
        ocr_note=ocr_result.note,
        source_doc_status=document_result.status,
        source_doc_text=unit.source_excerpt,
        source_doc_note=document_result.note,
        extract_strategy="rules_plus_platform_history",
        extract_warnings=["平台截图金额/账单 ID 来自 OCR 和历史导出匹配，需人工复核。"],
        material_tags=_material_tags_from_context(raw_text, attachments, document_result),
    )
    save_draft(draft)
    record_case_event(case_id=case_id, draft_id=draft_id, event_type="platform_history_draft_created", payload=draft_snapshot(draft))
    return draft

def _extract_reissue_draft_units(
    *,
    draft_dir: Path,
    attachments: list[DraftAttachment],
    instruction_text: str,
    fallback_context: str,
    company_name: str,
) -> list[ReissueDraftUnit]:
    if not _looks_like_reissue_instruction(instruction_text):
        return []
    invoice_numbers = _extract_invoice_numbers(instruction_text) or _extract_invoice_numbers(fallback_context)
    target_amounts = _extract_reissue_target_amounts(instruction_text)
    if not invoice_numbers or not target_amounts:
        return []
    records: list[HistoryInvoiceRecord] = []
    for attachment in attachments:
        suffix = Path(attachment.stored_name).suffix.lower()
        if suffix not in {".xlsx", ".xls"}:
            continue
        source_path = draft_dir / attachment.stored_name
        try:
            records.extend(_history_invoice_records_from_workbook(source_path, attachment.original_name))
        except Exception:  # noqa: BLE001
            continue
    if not records:
        return []
    by_invoice = {record.invoice_no: record for record in records if record.invoice_no}
    original = next((by_invoice[number] for number in invoice_numbers if number in by_invoice), None)
    if original is None:
        return []
    matched_records = _match_positive_history_records_for_amounts(
        records=records,
        target_amounts=target_amounts,
        original=original,
        company_name=company_name,
    )
    if matched_records:
        return [_reissue_unit_from_record(record, f"按客户重开金额 {amount} 匹配历史正数发票") for record, amount in zip(matched_records, target_amounts)]
    fallback_line_groups = _fallback_reissue_line_groups(original.lines, target_amounts)
    units: list[ReissueDraftUnit] = []
    for amount, lines in zip(target_amounts, fallback_line_groups):
        units.append(
            ReissueDraftUnit(
                source_name=f"原票 {original.invoice_no} 重开 {amount}",
                buyer=original.buyer,
                invoice_kind=original.invoice_kind,
                note=original.note,
                lines=lines,
                target_amount=amount,
                source_excerpt=(
                    f"红冲后重开：从历史发票 {original.invoice_no} 复用购买方、票种、税率、税收编码和明细；"
                    f"本次目标含税金额 {amount}。"
                ),
            )
        )
    return units


def _looks_like_reissue_instruction(text: str) -> bool:
    compact = re.sub(r"\s+", "", str(text or ""))
    if not compact:
        return False
    has_invoice_no = bool(_extract_invoice_numbers(compact))
    has_reissue_word = bool(re.search(r"重开|重新开|再开|分[两二三四五六七八九十0-9]+(?:笔|张)|作废|红冲|冲红", compact))
    return has_invoice_no and has_reissue_word


def _extract_invoice_numbers(text: str) -> list[str]:
    seen: set[str] = set()
    numbers: list[str] = []
    for matched in re.findall(r"(?<!\d)(\d{20})(?!\d)", str(text or "")):
        if matched not in seen:
            seen.add(matched)
            numbers.append(matched)
    return numbers


def _extract_reissue_target_amounts(text: str) -> list[str]:
    source = str(text or "")
    amount_texts: list[str] = []
    amount_list_match = re.search(r"金额\s*分别\s*(?:为|是)?\s*([0-9.,，、\s和及]+)", source)
    if amount_list_match:
        amount_texts.extend(re.findall(r"\d{2,}(?:\.\d{1,2})?", amount_list_match.group(1)))
    single_patterns = [
        r"开这个数\s*(\d{2,}(?:\.\d{1,2})?)",
        r"按(?:这个|新)?(?:金额|数)\s*(\d{2,}(?:\.\d{1,2})?)\s*重开",
        r"重开[^\d]{0,10}(\d{2,}(?:\.\d{1,2})?)",
    ]
    for pattern in single_patterns:
        for matched in re.findall(pattern, source):
            amount_texts.append(matched)
    amounts: list[str] = []
    seen: set[str] = set()
    for raw in amount_texts:
        parsed = _decimal_from_text(raw)
        if parsed is None or parsed <= 0 or parsed >= Decimal("100000000"):
            continue
        text_value = f"{parsed:.2f}"
        if text_value not in seen:
            seen.add(text_value)
            amounts.append(text_value)
    return amounts[:10]


def _history_invoice_records_from_workbook(path: Path, source_name: str) -> list[HistoryInvoiceRecord]:
    rows_by_sheet = _workbook_rows_by_sheet(path)
    base_rows: list[list[str]] = []
    line_rows: list[list[str]] = []
    for rows in rows_by_sheet:
        header_index, header = _find_history_header(rows, ["数电发票号码", "货物或应税劳务名称", "税收分类编码", "价税合计"])
        if header_index is not None and header is not None:
            line_rows = [header, *rows[header_index + 1 :]]
            continue
        header_index, header = _find_history_header(rows, ["数电发票号码", "购买方名称", "发票票种", "是否正数发票"])
        if header_index is not None and header is not None:
            base_rows = [header, *rows[header_index + 1 :]]
    if not base_rows or not line_rows:
        return []
    line_header = line_rows[0]
    line_index = {name: _first_header_index(line_header, [name]) for name in line_header if name}
    grouped_lines: dict[str, list[InvoiceLine]] = {}
    for row in line_rows[1:]:
        invoice_no = _row_value(row, line_index.get("数电发票号码"))
        if not invoice_no:
            continue
        line = _history_line_from_row(row, line_index)
        if line is None:
            continue
        grouped_lines.setdefault(invoice_no, []).append(line)
    base_header = base_rows[0]
    base_index = {name: _first_header_index(base_header, [name]) for name in base_header if name}
    records: list[HistoryInvoiceRecord] = []
    for row in base_rows[1:]:
        invoice_no = _row_value(row, base_index.get("数电发票号码"))
        if not invoice_no:
            continue
        total = _money_text(_row_value(row, base_index.get("价税合计")))
        positive_text = _row_value(row, base_index.get("是否正数发票"))
        kind_text = _row_value(row, base_index.get("发票票种"))
        records.append(
            HistoryInvoiceRecord(
                invoice_no=invoice_no,
                seller_name=_row_value(row, base_index.get("销方名称")),
                buyer=BuyerInfo(
                    name=_row_value(row, base_index.get("购买方名称")),
                    tax_id=_row_value(row, base_index.get("购方识别号")),
                ),
                invoice_kind="增值税专用发票" if "专用" in kind_text else "普通发票",
                total_amount=total,
                is_positive=positive_text != "否" and not str(total).startswith("-"),
                status=_row_value(row, base_index.get("发票状态")),
                issued_at=_row_value(row, base_index.get("开票日期")),
                note=_row_value(row, base_index.get("备注")),
                lines=grouped_lines.get(invoice_no, []),
                source_name=source_name,
            )
        )
    return records


def _find_history_header(rows: list[list[str]], required: list[str]) -> tuple[int | None, list[str] | None]:
    for index, row in enumerate(rows[:20]):
        normalized = [cell.strip() for cell in row]
        if all(_first_header_index(normalized, [name]) is not None for name in required):
            return index, normalized
    return None, None


def _history_line_from_row(row: list[str], index: dict[str, int | None]) -> InvoiceLine | None:
    amount = _money_text(_row_value(row, index.get("价税合计")))
    if not amount:
        return None
    raw_name = _row_value(row, index.get("货物或应税劳务名称"))
    tax_category, project_name = _split_history_item_name(raw_name)
    if not project_name:
        return None
    return InvoiceLine(
        project_name=project_name,
        tax_category=tax_category,
        specification=_row_value(row, index.get("规格型号")),
        unit=_row_value(row, index.get("单位")),
        quantity=_number_text(_row_value(row, index.get("数量"))),
        unit_price=_money_text(_row_value(row, index.get("单价"))),
        amount_with_tax=amount,
        tax_rate=_workbook_tax_rate_text(_row_value(row, index.get("税率"))) or "1%",
        tax_code=_row_value(row, index.get("税收分类编码")),
        coding_reference="历史发票明细复用：红冲后重开草稿按原票/历史正数票保留税收编码，需人工复核。",
    )


def _split_history_item_name(raw_name: str) -> tuple[str, str]:
    text = str(raw_name or "").strip()
    parts = [part.strip() for part in text.split("*") if part.strip()]
    if len(parts) >= 2:
        return parts[0], parts[-1]
    return "", text.strip("*")


def _match_positive_history_records_for_amounts(
    *,
    records: list[HistoryInvoiceRecord],
    target_amounts: list[str],
    original: HistoryInvoiceRecord,
    company_name: str,
) -> list[HistoryInvoiceRecord]:
    matched: list[HistoryInvoiceRecord] = []
    used: set[str] = set()
    for amount in target_amounts:
        amount_decimal = _decimal_from_text(amount)
        candidates = []
        for record in records:
            if record.invoice_no == original.invoice_no or record.invoice_no in used:
                continue
            if not record.is_positive or not record.lines:
                continue
            if company_name.strip() and record.seller_name.strip() and company_name.strip() not in record.seller_name.strip() and record.seller_name.strip() not in company_name.strip():
                continue
            if original.buyer.name and record.buyer.name != original.buyer.name:
                continue
            if amount_decimal is None or _decimal_from_text(record.total_amount) != amount_decimal:
                continue
            candidates.append(record)
        if not candidates:
            return []
        candidates.sort(key=lambda item: item.issued_at, reverse=True)
        selected = candidates[0]
        used.add(selected.invoice_no)
        matched.append(selected)
    return matched


def _fallback_reissue_line_groups(original_lines: list[InvoiceLine], target_amounts: list[str]) -> list[list[InvoiceLine]]:
    if not original_lines:
        return []
    if len(target_amounts) == 1:
        target = target_amounts[0]
        lines = [replace(line) for line in original_lines]
        if len(lines) == 1:
            lines[0].amount_with_tax = target
            lines[0].quantity = lines[0].quantity or "1"
            lines[0].unit_price = ""
        return [lines]
    groups: list[list[InvoiceLine]] = []
    cursor = 0
    for amount in target_amounts:
        target = _decimal_from_text(amount)
        running = Decimal("0")
        group: list[InvoiceLine] = []
        while cursor < len(original_lines):
            line = replace(original_lines[cursor])
            line_amount = _decimal_from_text(line.resolved_amount_with_tax()) or Decimal("0")
            running += abs(line_amount)
            group.append(line)
            cursor += 1
            if target is not None and abs(running - target) <= Decimal("0.01"):
                break
        if group:
            groups.append(group)
    return groups if len(groups) == len(target_amounts) else []


def _reissue_unit_from_record(record: HistoryInvoiceRecord, reason: str) -> ReissueDraftUnit:
    return ReissueDraftUnit(
        source_name=f"历史正数发票 {record.invoice_no}",
        buyer=record.buyer,
        invoice_kind=record.invoice_kind,
        note=record.note,
        lines=[replace(line) for line in record.lines],
        target_amount=record.total_amount,
        source_excerpt=(
            f"红冲后重开：{reason}；购买方 {record.buyer.name}；"
            f"票种 {record.invoice_kind}；含税合计 {record.total_amount}。"
        ),
    )


def _create_reissue_drafts(
    *,
    batch_id: str,
    case_id: str,
    draft_dir: Path,
    company_name: str,
    raw_text: str,
    note: str,
    attachments: list[DraftAttachment],
    document_result,
    ocr_result,
    fallback_invoice_profile: dict[str, str],
    units: list[ReissueDraftUnit],
) -> InvoiceDraft | DraftBatch:
    if len(units) == 1:
        return _create_single_reissue_draft(
            draft_id=batch_id,
            case_id=case_id,
            draft_dir=draft_dir,
            company_name=company_name,
            raw_text=raw_text,
            note=note,
            attachments=attachments,
            document_result=document_result,
            ocr_result=ocr_result,
            fallback_invoice_profile=fallback_invoice_profile,
            unit=units[0],
        )
    return _create_reissue_draft_batch(
        batch_id=batch_id,
        case_id=case_id,
        draft_dir=draft_dir,
        company_name=company_name,
        raw_text=raw_text,
        note=note,
        attachments=attachments,
        document_result=document_result,
        ocr_result=ocr_result,
        fallback_invoice_profile=fallback_invoice_profile,
        units=units,
    )


def _create_single_reissue_draft(
    *,
    draft_id: str,
    case_id: str,
    draft_dir: Path,
    company_name: str,
    raw_text: str,
    note: str,
    attachments: list[DraftAttachment],
    document_result,
    ocr_result,
    fallback_invoice_profile: dict[str, str],
    unit: ReissueDraftUnit,
) -> InvoiceDraft:
    child_attachments = list(attachments)
    merged_note = _merge_extracted_note(note, unit.note)
    issues = _build_draft_issues(
        company_name=company_name,
        raw_text=unit.source_excerpt or raw_text,
        attachments=child_attachments,
        buyer=unit.buyer,
        lines=unit.lines,
        special_business=fallback_invoice_profile.get("special_business", ""),
        document_status=document_result.status,
        document_note=document_result.note,
        ocr_status=ocr_result.status,
        ocr_note=ocr_result.note,
    )
    issues.insert(0, "本草稿由“红冲后重开”规则生成；系统只生成新的正数发票草稿，不处理红冲动作。")
    draft = InvoiceDraft(
        draft_id=draft_id,
        case_id=case_id,
        company_name=company_name.strip(),
        buyer=unit.buyer,
        lines=unit.lines,
        raw_text=unit.source_excerpt or raw_text,
        note=merged_note,
        issues=issues,
        source_images=child_attachments,
        workbook_name="开票明细表.xlsx",
        created_at=datetime.now().isoformat(timespec="seconds"),
        invoice_kind=unit.invoice_kind or fallback_invoice_profile.get("invoice_kind", "普通发票"),
        invoice_medium=fallback_invoice_profile.get("invoice_medium", "电子发票"),
        special_business=fallback_invoice_profile.get("special_business", ""),
        ocr_status=ocr_result.status,
        ocr_engine=ocr_result.engine,
        ocr_text=ocr_result.combined_text,
        ocr_note=ocr_result.note,
        source_doc_status=document_result.status,
        source_doc_text=unit.source_excerpt,
        source_doc_note=document_result.note,
        extract_strategy="rules_plus_reissue_history",
        extract_warnings=["检测到红冲后重开场景：红冲动作暂不自动处理；本页只生成新的正数发票草稿。"],
        material_tags=_material_tags_from_context(raw_text, attachments, document_result),
    )
    save_draft(draft)
    record_case_event(case_id=case_id, draft_id=draft_id, event_type="reissue_draft_created", payload=draft_snapshot(draft))
    return draft


def _common_reissue_invoice_kind(units: list[ReissueDraftUnit]) -> str:
    kinds = {unit.invoice_kind for unit in units if unit.invoice_kind}
    return kinds.pop() if len(kinds) == 1 else ""


def _create_reissue_draft_batch(
    *,
    batch_id: str,
    case_id: str,
    draft_dir: Path,
    company_name: str,
    raw_text: str,
    note: str,
    attachments: list[DraftAttachment],
    document_result,
    ocr_result,
    fallback_invoice_profile: dict[str, str],
    units: list[ReissueDraftUnit],
) -> DraftBatch:
    material_tags = _material_tags_from_context(raw_text, attachments, document_result)
    items: list[DraftBatchItem] = []
    batch_issues = [
        "当前材料命中了“红冲后拆分重开”规则；系统只生成新的正数发票草稿，不处理红冲动作。",
        "请重点复核每张重开草稿的拆分金额、明细归属、购买方、备注、税率和税收编码。",
    ]
    for unit in units:
        child_draft_id = uuid4().hex[:10]
        child_dir = draft_directory(child_draft_id)
        child_dir.mkdir(parents=True, exist_ok=True)
        child_attachments = _clone_attachments_to_directory(source_dir=draft_dir, target_dir=child_dir, attachments=attachments)
        merged_note = _merge_extracted_note(note, unit.note)
        child_issues = _build_draft_issues(
            company_name=company_name,
            raw_text=unit.source_excerpt or raw_text,
            attachments=child_attachments,
            buyer=unit.buyer,
            lines=unit.lines,
            special_business=fallback_invoice_profile.get("special_business", ""),
            document_status=document_result.status,
            document_note=document_result.note,
            ocr_status=ocr_result.status,
            ocr_note=ocr_result.note,
        )
        child_issues.insert(0, f"本草稿由红冲后重开金额 `{unit.target_amount}` 匹配生成；请复核是否对应客户要求的其中一笔。")
        child_draft = InvoiceDraft(
            draft_id=child_draft_id,
            case_id=case_id,
            company_name=company_name.strip(),
            buyer=unit.buyer,
            lines=unit.lines,
            raw_text=unit.source_excerpt or raw_text,
            note=merged_note,
            issues=child_issues,
            source_images=child_attachments,
            workbook_name="开票明细表.xlsx",
            created_at=datetime.now().isoformat(timespec="seconds"),
            invoice_kind=unit.invoice_kind or fallback_invoice_profile.get("invoice_kind", "普通发票"),
            invoice_medium=fallback_invoice_profile.get("invoice_medium", "电子发票"),
            special_business=fallback_invoice_profile.get("special_business", ""),
            ocr_status=ocr_result.status,
            ocr_engine=ocr_result.engine,
            ocr_text=ocr_result.combined_text,
            ocr_note=ocr_result.note,
            source_doc_status=document_result.status,
            source_doc_text=unit.source_excerpt,
            source_doc_note=document_result.note,
            extract_strategy="rules_plus_reissue_history_batch",
            extract_warnings=["检测到红冲后拆分重开场景：红冲动作暂不自动处理；本页只生成新的正数发票草稿。"],
            material_tags=material_tags,
        )
        save_draft(child_draft)
        record_case_event(case_id=case_id, draft_id=child_draft_id, batch_id=batch_id, event_type="reissue_child_draft_created", payload=draft_snapshot(child_draft))
        items.append(
            DraftBatchItem(
                draft_id=child_draft_id,
                buyer_name=unit.buyer.name or "待补全购买方名称",
                invoice_kind=unit.invoice_kind or fallback_invoice_profile.get("invoice_kind", "普通发票"),
                amount_total=_sum_line_amounts(unit.lines),
                project_summary=_summarize_projects(unit.lines),
                line_count=len(unit.lines),
                issue_summary=child_issues[0] if child_issues else "",
            )
        )
        batch_issues.extend(child_issues)
    batch = DraftBatch(
        batch_id=batch_id,
        case_id=case_id,
        company_name=company_name.strip(),
        created_at=datetime.now().isoformat(timespec="seconds"),
        items=items,
        raw_text=raw_text,
        note=note.strip(),
        issues=batch_issues,
        source_images=attachments,
        invoice_kind=_common_reissue_invoice_kind(units) or fallback_invoice_profile.get("invoice_kind", "普通发票"),
        invoice_medium=fallback_invoice_profile.get("invoice_medium", "电子发票"),
        special_business=fallback_invoice_profile.get("special_business", ""),
        extract_strategy="rules_plus_reissue_history_batch",
        llm_provider="",
        extract_warnings=["检测到红冲后拆分重开场景：红冲动作暂不自动处理；本页只生成新的正数发票草稿。"],
        material_tags=material_tags,
    )
    save_draft_batch(batch)
    record_case_event(case_id=case_id, batch_id=batch_id, event_type="reissue_draft_batch_created", payload=batch_snapshot(batch))
    return batch

def _extract_workbook_invoice_units(draft_dir: Path, attachments: list[DraftAttachment]) -> list[WorkbookInvoiceUnit]:
    units: list[WorkbookInvoiceUnit] = []
    for attachment in attachments:
        suffix = Path(attachment.stored_name).suffix.lower()
        if suffix not in {".xls", ".xlsx"}:
            continue
        source_path = draft_dir / attachment.stored_name
        if not source_path.exists():
            continue
        try:
            lines = _purchase_request_lines_from_workbook(source_path)
        except Exception:  # noqa: BLE001
            lines = []
        if not lines:
            try:
                lines = _generic_invoice_lines_from_workbook(source_path)
            except Exception:  # noqa: BLE001
                lines = []
        if not lines:
            continue
        excerpt_rows = ["项目名称\t规格型号\t单位\t数量\t单价\t含税金额"]
        for line in lines[:80]:
            excerpt_rows.append(
                "\t".join(
                    [
                        line.project_name,
                        line.specification,
                        line.unit,
                        line.quantity,
                        line.unit_price,
                        line.amount_with_tax,
                    ]
                )
            )
        units.append(
            WorkbookInvoiceUnit(
                source_name=attachment.original_name or Path(attachment.stored_name).name,
                lines=lines,
                source_excerpt="\n".join(excerpt_rows),
            )
        )
    return units


def _purchase_request_lines_from_workbook(path: Path) -> list[InvoiceLine]:
    rows_by_sheet = _workbook_rows_by_sheet(path)
    lines: list[InvoiceLine] = []
    for rows in rows_by_sheet:
        header_index = _find_purchase_request_header(rows)
        if header_index is None:
            continue
        header = rows[header_index]
        index = {name: _find_header_index(header, name) for name in ["物资名称", "规格型号", "计量单位", "采购单价", "实收数量", "实际金额"]}
        if index["物资名称"] is None or index["实际金额"] is None:
            continue
        for row in rows[header_index + 1:]:
            project_name = _row_value(row, index["物资名称"])
            if not project_name or "合计" in project_name:
                continue
            amount = _money_text(_row_value(row, index["实际金额"]))
            if not amount:
                continue
            unit_price = _money_text(_row_value(row, index["采购单价"])) if index["采购单价"] is not None else ""
            quantity = _number_text(_row_value(row, index["实收数量"])) if index["实收数量"] is not None else ""
            line = InvoiceLine(
                project_name=project_name,
                specification=_row_value(row, index["规格型号"]),
                unit=_row_value(row, index["计量单位"]) or "项",
                quantity=quantity or "1",
                unit_price=unit_price,
                amount_with_tax=amount,
                tax_rate="1%",
                coding_reference="Excel 明细本地解析，需人工复核：按客户表格生成开票明细。",
            )
            lines.append(line)
    return lines


def _generic_invoice_lines_from_workbook(path: Path) -> list[InvoiceLine]:
    """Best-effort parser for customer-provided invoice/detail workbooks."""

    rows_by_sheet = _workbook_rows_by_sheet(path)
    vehicle_lines = _vehicle_missing_certificate_lines_from_rows(rows_by_sheet)
    if vehicle_lines:
        return vehicle_lines
    parsed: list[InvoiceLine] = []
    for rows in rows_by_sheet:
        text = _rows_to_tab_text(rows)
        if not text.strip():
            continue
        parsed.extend(parse_bulk_invoice_lines(text))
    if parsed:
        return _mark_workbook_lines(parsed)
    return _fallback_project_amount_lines_from_rows(rows_by_sheet)


def _vehicle_missing_certificate_lines_from_rows(rows_by_sheet: list[list[list[str]]]) -> list[InvoiceLine]:
    lines: list[InvoiceLine] = []
    for rows in rows_by_sheet:
        if not rows:
            continue
        header = [cell.strip() for cell in rows[0]]
        company_index = _first_header_index(header, ["公司名", "客户名称", "购买方名称"])
        error_index = _first_header_index(header, ["错误信息", "车辆信息", "商品信息"])
        if error_index is None or "发票行数" not in header:
            continue
        current_buyer = ""
        for row in rows[1:]:
            buyer = _row_value(row, company_index)
            if buyer:
                current_buyer = buyer
            info = _row_value(row, error_index)
            if not info:
                continue
            matched = re.search(r"(?P<model>.+?)\s*车架号\s*(?P<vin>[A-Za-z0-9]+)\s*价格\s*(?P<amount>\d+(?:\.\d{1,2})?)", info)
            if not matched:
                continue
            model = matched.group("model").strip()
            lines.append(
                InvoiceLine(
                    project_name=f"电动两轮摩托车 {model}".strip(),
                    tax_category="机动车",
                    specification=f"车架号 {matched.group('vin')}",
                    unit="辆",
                    quantity="1",
                    amount_with_tax=_money_text(matched.group("amount")),
                    tax_rate="1%",
                    coding_reference=(
                        "机动车合格证异常表解析，需人工复核："
                        f"按错误信息提取车型/车架号/价格；购买方线索：{current_buyer}。"
                    ),
                )
            )
    return _dedupe_workbook_lines(lines)



def _workbook_rows_by_sheet(path: Path) -> list[list[list[str]]]:
    suffix = path.suffix.lower()
    rows_by_sheet: list[list[list[str]]] = []
    if suffix == ".xls":
        import xlrd

        workbook = xlrd.open_workbook(path)
        for sheet in workbook.sheets():
            rows_by_sheet.append([[_cell_text(cell) for cell in sheet.row_values(row_index)] for row_index in range(sheet.nrows)])
    elif suffix == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                rows_by_sheet.append([[_cell_text(cell) for cell in row] for row in sheet.iter_rows(values_only=True)])
        finally:
            workbook.close()
    return rows_by_sheet


def _rows_to_tab_text(rows: list[list[str]]) -> str:
    lines: list[str] = []
    for row in rows[:500]:
        trimmed = [cell.strip() for cell in row]
        while trimmed and not trimmed[-1]:
            trimmed.pop()
        if any(trimmed):
            lines.append("\t".join(trimmed))
    return "\n".join(lines)


def _mark_workbook_lines(lines: list[InvoiceLine]) -> list[InvoiceLine]:
    marked: list[InvoiceLine] = []
    for line in lines:
        if not line.coding_reference:
            line.coding_reference = "Excel 表格通用解析，需人工复核：按客户材料生成开票明细。"
        marked.append(line)
    return marked


def _dedupe_workbook_lines(lines: list[InvoiceLine]) -> list[InvoiceLine]:
    seen: set[tuple[str, str, str, str, str]] = set()
    deduped: list[InvoiceLine] = []
    for line in lines:
        key = (
            line.project_name.strip(),
            line.specification.strip(),
            line.quantity.strip(),
            line.unit_price.strip(),
            line.resolved_amount_with_tax().strip() or line.amount_with_tax.strip(),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(line)
    return _mark_workbook_lines(deduped)


def _fallback_project_amount_lines_from_rows(rows_by_sheet: list[list[list[str]]]) -> list[InvoiceLine]:
    lines: list[InvoiceLine] = []
    for rows in rows_by_sheet:
        header_index, header = _find_generic_project_amount_header(rows)
        if header_index is None or header is None:
            continue
        project_index = _first_header_index(header, ["项目名称", "开票内容", "开票项目", "发票项目", "商品全名", "商品名称", "品名", "物资名称", "清单名称", "名称"])
        amount_index = _first_header_index(header, ["开票金额", "含税金额", "含税总价", "含税总价(元)", "价税合计", "金额", "实际金额", "合计金额"])
        if project_index is None or amount_index is None:
            continue
        spec_index = _first_header_index(header, ["规格型号", "规格/项目特征", "规格"])
        unit_index = _first_header_index(header, ["单位", "计量单位"])
        qty_index = _first_header_index(header, ["数量", "销售数量", "实收数量"])
        price_index = _first_header_index(header, ["单价", "含税单价", "含税单价(元)", "采购单价"])
        tax_rate_index = _first_header_index(header, ["税率", "税率(%)", "税率/征收率"])
        category_index = _first_header_index(header, ["赋码大类", "税收分类", "税目大类", "大类"])
        for row in rows[header_index + 1:]:
            project_name = _row_value(row, project_index)
            amount = _money_text(_row_value(row, amount_index))
            if not project_name or not amount:
                continue
            if "合计" in project_name or project_name in {"小计", "总计"}:
                continue
            lines.append(
                InvoiceLine(
                    project_name=project_name.strip("*"),
                    tax_category=_row_value(row, category_index),
                    specification=_row_value(row, spec_index),
                    unit=_row_value(row, unit_index) or "项",
                    quantity=_number_text(_row_value(row, qty_index)) if qty_index is not None else "1",
                    unit_price=_money_text(_row_value(row, price_index)) if price_index is not None else "",
                    amount_with_tax=amount,
                    tax_rate=_workbook_tax_rate_text(
                        _row_value(row, tax_rate_index),
                        _row_value(header, tax_rate_index),
                    ) or "3%",
                    coding_reference="Excel 表格兜底解析，需人工复核：按项目/金额列生成开票明细。",
                )
            )
    return _dedupe_workbook_lines(lines)


def _find_generic_project_amount_header(rows: list[list[str]]) -> tuple[int | None, list[str] | None]:
    for index, row in enumerate(rows[:50]):
        normalized = [cell.strip().replace("（", "(").replace("）", ")") for cell in row]
        has_project = _first_header_index(normalized, ["项目名称", "开票内容", "开票项目", "发票项目", "商品全名", "商品名称", "品名", "物资名称", "清单名称", "名称"]) is not None
        has_amount = _first_header_index(normalized, ["开票金额", "含税金额", "含税总价", "含税总价(元)", "价税合计", "金额", "实际金额", "合计金额"]) is not None
        if has_project and has_amount:
            return index, normalized
    return None, None


def _first_header_index(header: list[str], names: list[str]) -> int | None:
    wanted = {name.strip() for name in names}
    for index, value in enumerate(header):
        if value.strip() in wanted:
            return index
    return None


def _cell_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        try:
            return str(int(float(text)))
        except ValueError:
            return text
    return text


def _find_purchase_request_header(rows: list[list[str]]) -> int | None:
    for index, row in enumerate(rows[:30]):
        compact = "\t".join(row)
        if "物资名称" in compact and "实际金额" in compact:
            return index
    return None


def _find_header_index(header: list[str], name: str) -> int | None:
    for index, value in enumerate(header):
        if value.strip() == name:
            return index
    return None


def _row_value(row: list[str], index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    return row[index].strip()


def _decimal_from_text(value: str) -> Decimal | None:
    text = str(value or "").replace(",", "").replace("，", "").replace("¥", "").replace("￥", "").strip()
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _money_text(value: str) -> str:
    parsed = _decimal_from_text(value)
    return f"{parsed:.2f}" if parsed is not None else ""


def _number_text(value: str) -> str:
    parsed = _decimal_from_text(value)
    if parsed is None:
        return str(value or "").strip()
    text = f"{parsed:f}"
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def _workbook_tax_rate_text(value: str, header: str = "") -> str:
    text = str(value or "").strip().replace("％", "%")
    if not text:
        return ""
    if text.endswith("%") or text in {"免税", "不征税", "免征增值税"}:
        return text
    parsed = _decimal_from_text(text)
    if parsed is None:
        return text
    # 工程清单类 Excel 常见表头为“税率(%)”，单元格值 1 / 1.0 表示 1%，不是 100%。
    if "%" in str(header or ""):
        return f"{_number_text(text)}%"
    if parsed <= Decimal("1"):
        parsed *= Decimal("100")
    percent_text = f"{parsed:f}"
    if "." in percent_text:
        percent_text = percent_text.rstrip("0").rstrip(".")
    return f"{percent_text}%"


def _sum_line_amounts(lines: list[InvoiceLine]) -> str:
    total = Decimal("0")
    has_amount = False
    for line in lines:
        parsed = _decimal_from_text(line.resolved_amount_with_tax())
        if parsed is not None:
            total += parsed
            has_amount = True
    return f"{total:.2f}" if has_amount else ""


def _save_uploads(draft_dir: Path, uploaded_files: list[FileStorage]) -> list[DraftAttachment]:
    uploads_dir = draft_dir / "uploads"
    uploads_dir.mkdir(parents=True, exist_ok=True)
    attachments: list[DraftAttachment] = []
    next_index = len([path for path in uploads_dir.iterdir() if path.is_file()]) + 1

    for file in uploaded_files:
        if not file.filename:
            continue
        original_name = Path(file.filename).name
        safe_base = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff._-]+", "_", Path(original_name).stem).strip("_") or f"image_{next_index}"
        suffix = Path(original_name).suffix.lower() or ".bin"
        stored_name = f"{next_index:02d}_{safe_base}{suffix}"
        target = uploads_dir / stored_name
        file.save(target)
        attachments.append(
            DraftAttachment(
                original_name=original_name,
                stored_name=f"uploads/{stored_name}",
                mime_type=file.mimetype or "",
                size_bytes=target.stat().st_size,
            )
        )
        next_index += 1
    return attachments


def _clone_attachments_to_directory(*, source_dir: Path, target_dir: Path, attachments: list[DraftAttachment]) -> list[DraftAttachment]:
    cloned: list[DraftAttachment] = []
    for attachment in attachments:
        source_path = source_dir / attachment.stored_name
        target_path = target_dir / attachment.stored_name
        target_path.parent.mkdir(parents=True, exist_ok=True)
        if source_path.exists():
            shutil.copy2(source_path, target_path)
        cloned.append(
            DraftAttachment(
                original_name=attachment.original_name,
                stored_name=attachment.stored_name,
                mime_type=attachment.mime_type,
                size_bytes=attachment.size_bytes,
            )
        )
    return cloned


def _write_workbook(path: Path, draft: InvoiceDraft) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "开票明细"
    headers = [
        "项目名称",
        "赋码大类",
        "税收编码",
        "零件编码",
        "规格型号",
        "单位",
        "数量",
        "单价",
        "含税金额",
        "税率/征收率",
        "赋码说明",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCEFE9")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for line in draft.lines:
        sheet.append(
            [
                line.project_name,
                line.tax_category,
                line.tax_code,
                line.source_item_code,
                line.specification,
                line.unit,
                line.quantity,
                line.unit_price,
                line.resolved_amount_with_tax(),
                line.normalized_tax_rate(),
                line.coding_reference,
            ]
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:K{max(sheet.max_row, 1)}"
    widths = {
        "A": 28,
        "B": 18,
        "C": 24,
        "D": 18,
        "E": 18,
        "F": 10,
        "G": 10,
        "H": 12,
        "I": 14,
        "J": 12,
        "K": 34,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    meta = workbook.create_sheet("来源")
    meta["A1"] = "Case ID"
    meta["B1"] = draft.case_id
    meta["A2"] = "草稿ID"
    meta["B2"] = draft.draft_id
    meta["A3"] = "纳税主体"
    meta["B3"] = draft.company_name
    meta["A4"] = "购买方名称"
    meta["B4"] = draft.buyer.name
    meta["A5"] = "购买方税号"
    meta["B5"] = draft.buyer.tax_id
    meta["A6"] = "生成时间"
    meta["B6"] = draft.created_at
    meta["A7"] = "票种配置"
    meta["B7"] = f"{draft.invoice_medium} / {draft.invoice_kind}" + (f" / {draft.special_business}" if draft.special_business else "")
    meta["A8"] = "抽取策略"
    meta["B8"] = draft.extract_strategy
    meta["A9"] = "LLM Provider"
    meta["B9"] = draft.llm_provider
    meta["A10"] = "原始输入"
    meta["B10"] = draft.raw_text
    meta["A11"] = "备注"
    meta["B11"] = draft.note
    meta["A12"] = "附件"
    meta["B12"] = "\n".join(item.original_name for item in draft.source_images)
    meta["A13"] = "文档解析状态"
    meta["B13"] = draft.source_doc_status
    meta["A14"] = "文档解析备注"
    meta["B14"] = draft.source_doc_note
    meta["A15"] = "识别提醒"
    meta["B15"] = "\n".join(draft.issues + draft.extract_warnings)
    meta["A16"] = "OCR 状态"
    meta["B16"] = draft.ocr_status
    meta["A17"] = "OCR 引擎"
    meta["B17"] = draft.ocr_engine
    meta["A18"] = "OCR 备注"
    meta["B18"] = draft.ocr_note
    meta.column_dimensions["A"].width = 14
    meta.column_dimensions["B"].width = 90
    for row in range(1, 19):
        meta[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")

    workbook.save(path)


def _compose_parse_source(raw_text: str, document_text: str, ocr_text: str) -> str:
    return "\n\n".join(part for part in [raw_text.strip(), document_text.strip(), ocr_text.strip()] if part)


def _build_amount_split_lines(
    *,
    company_name: str,
    parse_source: str,
    buyer: BuyerInfo,
    lines: list[InvoiceLine],
    invoice_profile: dict[str, str],
) -> list[InvoiceLine]:
    if lines or not buyer.name or not buyer.tax_id:
        return []

    amounts = _extract_split_amounts(parse_source)
    if len(amounts) < 2:
        return []

    line_profile = _infer_split_line_profile(
        company_name=company_name,
        parse_source=parse_source,
        invoice_profile=invoice_profile,
    )
    if not line_profile:
        return []

    return [
        InvoiceLine(
            project_name=line_profile["project_name"],
            amount_with_tax=amount,
            tax_rate=line_profile["tax_rate"],
            tax_category=line_profile["tax_category"],
            coding_reference=line_profile["coding_reference"],
        )
        for amount in amounts
    ]


def _extract_split_amounts(parse_source: str) -> list[str]:
    lines = [line.strip() for line in parse_source.splitlines() if line.strip()]
    staged: list[str] = []
    capture = False

    for line in lines:
        compact = line.replace(" ", "")
        if compact in {"金额", "开票金额", "金额：", "金额:"}:
            capture = True
            continue
        if capture:
            if _looks_like_split_amount_line(compact):
                staged.append(_normalize_amount_text(compact))
                continue
            if staged:
                break

    if len(staged) >= 2:
        return _dedupe_preserving_order(staged)

    fallbacks: list[str] = []
    for line in lines:
        compact = line.replace(" ", "")
        if _looks_like_split_amount_line(compact):
            fallbacks.append(_normalize_amount_text(compact))
    return _dedupe_preserving_order(fallbacks)


def _looks_like_split_amount_line(value: str) -> bool:
    if not re.fullmatch(r"\d+(?:\.\d{1,2})?", value):
        return False
    if "." not in value and len(value) > 6:
        return False
    return True


def _normalize_amount_text(value: str) -> str:
    compact = value.replace(",", "").replace("，", "").replace(" ", "")
    if "." in compact:
        whole, decimal = compact.split(".", 1)
        return f"{whole}.{decimal[:2].ljust(2, '0')}"
    return f"{compact}.00"


def _dedupe_preserving_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    deduped: list[str] = []
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        deduped.append(value)
    return deduped


def _infer_split_line_profile(
    *,
    company_name: str,
    parse_source: str,
    invoice_profile: dict[str, str],
) -> dict[str, str] | None:
    compact = f"{company_name}\n{parse_source}".replace(" ", "")
    tax_rate = _infer_split_tax_rate(compact)
    if tax_rate != "1%":
        return None

    if re.search(r"(服务|科技|咨询|传媒|调试|运营|出租汽车|网约车)", compact):
        return {
            "project_name": "服务费",
            "tax_category": "现代服务",
            "tax_rate": tax_rate,
            "coding_reference": "金额拆票规则：聊天只给买方资料、多个金额和一个点票种，系统按现代服务/服务费自动拆成多张草稿。",
        }

    if invoice_profile.get("invoice_kind") == "增值税专用发票":
        return {
            "project_name": "服务费",
            "tax_category": "现代服务",
            "tax_rate": tax_rate,
            "coding_reference": "金额拆票规则：当前材料未给出明确项目名，先按通用服务费口径拆成多张草稿，待人工确认。",
        }
    return None


def _infer_split_tax_rate(compact_text: str) -> str:
    if re.search(r"(一个点|1个点|一%|1%)", compact_text):
        return "1%"
    return "3%"


def _summarize_projects(lines: list[InvoiceLine]) -> str:
    names = [line.project_name for line in lines if line.project_name]
    if not names:
        return "待人工补充"
    return " / ".join(names[:3])


def _infer_invoice_profile(parse_source: str, *, note: str = "") -> dict[str, str]:
    context = f"{parse_source}\n{note}".strip()
    compact = context.replace(" ", "")

    invoice_kind = "普通发票"
    invoice_medium = "电子发票"
    special_business = ""

    labeled_type = re.search(
        r"(?:发票类型|票种|开票类型)[：:\t ]?(增值税专用发票|专用发票|专票|增票|增值税普通发票|普通发票|普票)",
        compact,
    )
    checked_special = re.search(r"(?:增值税专用发票|专用发票)(?:[（(]?[√✓✔][）)]?)", compact)
    checked_normal = re.search(r"(?:增值税普通发票|普通发票)(?:[（(]?[√✓✔][）)]?)", compact)
    explicit_special = bool(re.search(r"(开专票|专票|电子专票|增票)", compact))
    explicit_normal = bool(re.search(r"(开普票|普票|普通发票（?√|普通发票√|增值税普通发票（?√|增值税普通发票√)", compact))

    if labeled_type and labeled_type.group(1) in {"普通发票", "增值税普通发票", "普票"}:
        invoice_kind = "普通发票"
        explicit_normal = True
    elif labeled_type and labeled_type.group(1) in {"增值税专用发票", "专用发票", "专票", "增票"}:
        invoice_kind = "增值税专用发票"
        explicit_special = True
    elif checked_normal and not checked_special:
        invoice_kind = "普通发票"
    elif checked_special and not checked_normal:
        invoice_kind = "增值税专用发票"
    elif explicit_normal and not explicit_special:
        invoice_kind = "普通发票"
    elif explicit_special and not explicit_normal:
        invoice_kind = "增值税专用发票"
    elif "增值税专用发票" in compact and "增值税普通发票" not in compact and "普通发票" not in compact:
        invoice_kind = "增值税专用发票"

    if re.search(r"(纸质发票|纸票)", compact):
        invoice_medium = "纸质发票"

    if invoice_kind == "普通发票" and not explicit_normal:
        has_tax_id = bool(re.search(r"(税务登记号|税号|纳税人识别号|统一社会信用代码)", compact))
        has_address = bool(re.search(r"(单位地址|购买方地址|购方地址|地址)", compact))
        has_phone = bool(re.search(r"(电话号码|购买方电话|购方电话|电话)", compact))
        has_bank = bool(re.search(r"(开户行|开户银行)", compact))
        has_account = bool(re.search(r"(银行账号|账号)", compact))
        if has_tax_id and has_bank and has_account and (has_address or has_phone):
            invoice_kind = "增值税专用发票"

    if re.search(r"(机动车|车架号|车辆识别代号|VIN|合格证|厂牌型号|发动机号|发动机号码)", compact, re.IGNORECASE):
        special_business = "机动车"

    return {
        "invoice_kind": invoice_kind,
        "invoice_medium": invoice_medium,
        "special_business": special_business,
    }


def _enrich_buyer_from_sheet_context(company_name: str, buyer: BuyerInfo, parse_source: str) -> BuyerInfo:
    resolved = BuyerInfo(
        name=buyer.name,
        tax_id=buyer.tax_id,
        address=buyer.address,
        phone=buyer.phone,
        bank_name=buyer.bank_name,
        bank_account=buyer.bank_account,
    )
    current_company = company_name.strip()
    inside_billing_block = False

    for raw_line in parse_source.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if line.startswith("需方") and not resolved.name:
            matched = re.search(r"需方\s+(.+?公司)", line)
            if matched:
                candidate = matched.group(1).strip()
                if candidate and candidate != current_company:
                    resolved.name = candidate
                    continue
        if "开票资料" in line:
            inside_billing_block = True
            continue
        if not inside_billing_block and not any(
            marker in line for marker in ("单位名称", "单位地址", "电话号码", "税务登记号", "开户行", "开户银行", "账号", "银行账户")
        ):
            continue

        key, value = _split_possible_key_value(line)
        if not key or not value:
            continue
        value = value.strip()
        if not value:
            continue

        if "单位名称" in key and not resolved.name and value != current_company:
            resolved.name = value
            continue
        if "税务登记号" in key and not resolved.tax_id:
            tax_id_match = re.search(r"([0-9A-Z]{15,20})", value.upper())
            resolved.tax_id = tax_id_match.group(1) if tax_id_match else value
            continue
        if "单位地址" in key and not resolved.address:
            resolved.address = value
            continue
        if "电话号码" in key and not resolved.phone:
            resolved.phone = value
            continue
        if ("开户行" in key or "开户银行" in key) and not resolved.bank_name:
            resolved.bank_name = value
            continue
        if key in {"账号", "银行账号", "银行账户"} and not resolved.bank_account:
            resolved.bank_account = value

    return resolved



def _enrich_buyer_from_history_profile(company_name: str, buyer: BuyerInfo, parse_source: str) -> BuyerInfo:
    if not company_name.strip():
        return buyer
    history_match = resolve_buyer_from_history(parse_source, company_name=company_name)
    if history_match is None:
        return buyer
    if buyer.name and buyer.tax_id and buyer.tax_id == history_match.buyer.tax_id:
        return buyer
    if buyer.name and not buyer.tax_id:
        normalized_name = buyer.name.replace(" ", "")
        normalized_history = history_match.buyer.name.replace(" ", "")
        if history_match.matched_alias not in normalized_name and normalized_name not in normalized_history:
            return buyer
    return BuyerInfo(
        name=history_match.buyer.name if not buyer.name or not buyer.tax_id else buyer.name,
        tax_id=buyer.tax_id or history_match.buyer.tax_id,
        address=buyer.address,
        phone=buyer.phone,
        bank_name=buyer.bank_name,
        bank_account=buyer.bank_account,
    )



def _apply_history_profile_to_lines(
    lines: list[InvoiceLine],
    *,
    company_name: str,
    buyer: BuyerInfo,
    parse_source: str,
) -> list[InvoiceLine]:
    return apply_line_history_hints(lines, company_name=company_name, buyer=buyer, raw_text=parse_source)



def _split_possible_key_value(line: str) -> tuple[str, str]:
    if "：" in line:
        left, right = line.split("：", 1)
        return left.strip(), right.strip()
    if ":" in line:
        left, right = line.split(":", 1)
        return left.strip(), right.strip()
    return "", ""


def _build_draft_issues(
    *,
    company_name: str,
    raw_text: str,
    attachments: list[DraftAttachment],
    buyer: BuyerInfo,
    lines: list[InvoiceLine],
    special_business: str,
    document_status: str,
    document_note: str,
    ocr_status: str,
    ocr_note: str,
) -> list[str]:
    issues: list[str] = []
    has_image_attachments = any(
        Path(item.stored_name).suffix.lower() in {".png", ".jpg", ".jpeg", ".bmp", ".gif", ".webp"}
        for item in attachments
    )
    if not company_name.strip():
        issues.append("当前还没填写纳税主体名称；执行前请补充当前已登录企业。")
    if not raw_text.strip() and not attachments:
        issues.append("请输入原始开票信息，或至少上传一份材料。")
    if document_note and document_status not in {"not_requested", "success"}:
        issues.append(document_note)
    if has_image_attachments and ocr_note:
        issues.append(ocr_note)
    if not buyer.name:
        issues.append("当前还没自动识别出购买方名称，请在草稿页人工补充。")
    if not lines:
        if attachments and ocr_status in {"success", "partial"}:
            issues.append("图片文字已经提取，但当前还没稳定抽出明细行；请在草稿页人工补充。")
        else:
            issues.append("当前还没自动识别出开票明细，请在草稿页人工补充。")
    formal_library_size = len(load_formal_coding_library())
    if lines and formal_library_size:
        unresolved_indexes = [
            str(index)
            for index, line in enumerate(lines, start=1)
            if not line.tax_category or not line.tax_code
        ]
        if unresolved_indexes:
            issues.append(
                "第 "
                + "、".join(unresolved_indexes)
                + " 行还没命中正式赋码库，请人工确认赋码大类 / 税率 / 税收编码。"
            )
        special_tax_indexes = [
            str(index)
            for index, line in enumerate(lines, start=1)
            if line.normalized_tax_rate() in {"免税", "不征税", "免征增值税"}
        ]
        if special_tax_indexes:
            issues.append(
                "第 "
                + "、".join(special_tax_indexes)
                + " 行命中了免税/不征税口径，请在执行前再次确认客户场景与票面口径一致。"
            )
        history_category_review_indexes = [
            str(index)
            for index, line in enumerate(lines, start=1)
            if "历史同项目还出现过" in (line.coding_reference or "")
        ]
        if history_category_review_indexes:
            issues.append(
                "第 "
                + "、".join(history_category_review_indexes)
                + " 行在客户历史档案中还出现过其它品类或编码；系统已按最高频历史口径推荐，请人工复核本次项目归类。"
            )
    if special_business == "机动车":
        issues.append("系统从材料中识别出机动车线索，建议在草稿里确认 `特定业务 = 机动车` 后再执行。")
    return issues


def _material_tags_from_context(raw_text: str, attachments: list[DraftAttachment], document_result) -> list[str]:
    tags: list[str] = []
    if str(raw_text or "").strip():
        compact = re.sub(r"\s+", "", raw_text)
        if re.search(r"(微信|群聊|聊天|麻烦|帮忙|开票|发票|税号|金额)", compact):
            tags.append("群文本/补充说明")
        else:
            tags.append("文本说明")
    for item in getattr(document_result, "document_results", []) or []:
        material_type = getattr(item, "material_type", "") or "文档材料"
        if material_type:
            tags.append(material_type)
    for attachment in attachments:
        suffix = Path(attachment.stored_name or attachment.original_name).suffix.lower()
        name = (attachment.original_name or attachment.stored_name or "").lower()
        if suffix not in {".png", ".jpg", ".jpeg", ".bmp", ".gif", ".webp"}:
            continue
        if re.search(r"(wechat|微信|群聊|聊天|screenshot|截图|longscreenshot)", name, re.IGNORECASE):
            tags.append("群聊截图")
        elif re.search(r"(车|牌|维修|事故|照片|photo|image)", name, re.IGNORECASE):
            tags.append("车辆/现场照片")
        else:
            tags.append("图片材料")
    return _dedupe_preserve_order(tags)[:8]



def _dedupe_preserve_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        cleaned = str(value or "").strip()
        if not cleaned or cleaned in seen:
            continue
        seen.add(cleaned)
        result.append(cleaned)
    return result



def _image_attachment_paths(draft_dir: Path, attachments: list[DraftAttachment]) -> list[Path]:
    return [
        draft_dir / item.stored_name
        for item in attachments
        if Path(item.stored_name).suffix.lower() in {".png", ".jpg", ".jpeg", ".bmp", ".gif", ".webp"}
    ]


def _run_draft_ocr(draft_dir: Path, attachments: list[DraftAttachment], *, defer_to_vision: bool = False):
    image_paths = _image_attachment_paths(draft_dir, attachments)
    if defer_to_vision and image_paths:
        from .llm_adapter import get_llm_adapter
        from .ocr import OptionalOcrResult

        if not get_llm_adapter().is_enabled:
            return run_optional_ocr(image_paths)
        return OptionalOcrResult(
            status="vision_deferred",
            note="图片材料将直接交给视觉大模型结构化识别，不先走本地 OCR。",
        )
    return run_optional_ocr(image_paths)


def _run_document_extraction(draft_dir: Path, attachments: list[DraftAttachment]):
    file_paths = [draft_dir / item.stored_name for item in attachments]
    result = extract_supported_documents(file_paths)
    (draft_dir / "source_doc_meta.json").write_text(serialize_document_results(result), encoding="utf-8")
    return result
