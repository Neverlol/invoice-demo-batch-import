from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .models import InvoiceDraft, InvoiceLine

LEDGER_ROOT = Path(__file__).resolve().parent.parent / "output" / "workbench" / "tax_invoice_demo"
LEDGER_CSV_PATH = LEDGER_ROOT / "累计发票明细表.csv"
LEDGER_XLSX_PATH = LEDGER_ROOT / "累计发票明细表.xlsx"
FEEDBACK_CSV_PATH = LEDGER_ROOT / "赋码反馈候选池.csv"

LEDGER_HEADERS = [
    "case_id",
    "draft_id",
    "line_no",
    "saved_at",
    "company_name",
    "buyer_name",
    "buyer_tax_id",
    "project_name",
    "tax_category",
    "tax_code",
    "source_item_code",
    "specification",
    "unit",
    "quantity",
    "unit_price",
    "amount_with_tax",
    "tax_rate",
    "coding_reference",
    "coding_state",
    "note",
]

FEEDBACK_HEADERS = [
    "case_id",
    "draft_id",
    "line_no",
    "saved_at",
    "candidate_status",
    "company_name",
    "buyer_name",
    "buyer_tax_id",
    "project_name",
    "tax_category",
    "tax_code",
    "source_item_code",
    "specification",
    "unit",
    "quantity",
    "unit_price",
    "amount_with_tax",
    "tax_rate",
    "coding_reference",
    "note",
]


def sync_draft_to_ledger(draft: InvoiceDraft) -> None:
    LEDGER_ROOT.mkdir(parents=True, exist_ok=True)
    current_rows = _read_csv(LEDGER_CSV_PATH)
    current_rows = [row for row in current_rows if row.get("draft_id") != draft.draft_id]
    saved_at = datetime.now().isoformat(timespec="seconds")
    for index, line in enumerate(draft.lines, start=1):
        current_rows.append(_line_to_ledger_row(draft, line, index, saved_at))
    _write_csv(LEDGER_CSV_PATH, LEDGER_HEADERS, current_rows)
    _write_ledger_workbook(LEDGER_XLSX_PATH, current_rows)

    feedback_rows = _read_csv(FEEDBACK_CSV_PATH)
    feedback_rows = [row for row in feedback_rows if row.get("draft_id") != draft.draft_id]
    for index, line in enumerate(draft.lines, start=1):
        candidate_status = _feedback_status(line)
        if not candidate_status:
            continue
        feedback_rows.append(
            {
                "draft_id": draft.draft_id,
                "case_id": draft.case_id,
                "line_no": str(index),
                "saved_at": saved_at,
                "candidate_status": candidate_status,
                "company_name": draft.company_name,
                "buyer_name": draft.buyer.name,
                "buyer_tax_id": draft.buyer.tax_id,
                "project_name": line.project_name,
                "tax_category": line.tax_category,
                "tax_code": line.tax_code,
                "source_item_code": line.source_item_code,
                "specification": line.specification,
                "unit": line.unit,
                "quantity": line.quantity,
                "unit_price": line.unit_price,
                "amount_with_tax": line.resolved_amount_with_tax(),
                "tax_rate": line.normalized_tax_rate(),
                "coding_reference": line.coding_reference,
                "note": draft.note,
            }
        )
    _write_csv(FEEDBACK_CSV_PATH, FEEDBACK_HEADERS, feedback_rows)


def _line_to_ledger_row(draft: InvoiceDraft, line: InvoiceLine, line_no: int, saved_at: str) -> dict[str, str]:
    return {
        "draft_id": draft.draft_id,
        "case_id": draft.case_id,
        "line_no": str(line_no),
        "saved_at": saved_at,
        "company_name": draft.company_name,
        "buyer_name": draft.buyer.name,
        "buyer_tax_id": draft.buyer.tax_id,
        "project_name": line.project_name,
        "tax_category": line.tax_category,
        "tax_code": line.tax_code,
        "source_item_code": line.source_item_code,
        "specification": line.specification,
        "unit": line.unit,
        "quantity": line.quantity,
        "unit_price": line.unit_price,
        "amount_with_tax": line.resolved_amount_with_tax(),
        "tax_rate": line.normalized_tax_rate(),
        "coding_reference": line.coding_reference,
        "coding_state": _coding_state(line),
        "note": draft.note,
    }


def _coding_state(line: InvoiceLine) -> str:
    if line.coding_reference.startswith("人工修正赋码"):
        return "manual_correction"
    if line.coding_reference.startswith("命中 "):
        return "auto_hit_formal"
    if line.coding_reference.startswith("官方分类候选 "):
        return "taxonomy_candidate"
    if line.tax_category:
        return "manual_or_external_fill"
    return "unresolved"


def _feedback_status(line: InvoiceLine) -> str:
    state = _coding_state(line)
    if state == "auto_hit_formal":
        return ""
    if state == "manual_correction":
        return "manual_correction"
    if state == "taxonomy_candidate":
        return "taxonomy_only"
    if state == "manual_or_external_fill":
        return "manual_fill_without_formal_hit"
    return "unresolved"


def _read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _write_csv(path: Path, headers: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def _write_ledger_workbook(path: Path, rows: list[dict[str, str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "累计发票明细"
    sheet.append(LEDGER_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCEFE9")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        sheet.append([row.get(header, "") for header in LEDGER_HEADERS])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:T{max(sheet.max_row, 1)}"
    widths = {
        "A": 14,
        "B": 14,
        "C": 8,
        "D": 22,
        "E": 24,
        "F": 24,
        "G": 22,
        "H": 28,
        "I": 16,
        "J": 22,
        "K": 16,
        "L": 16,
        "M": 10,
        "N": 10,
        "O": 10,
        "P": 12,
        "Q": 10,
        "R": 30,
        "S": 18,
        "T": 30,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    workbook.save(path)
