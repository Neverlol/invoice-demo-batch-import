from __future__ import annotations

import csv
import io
import json
import re
import tempfile
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from xml.etree import ElementTree as ET

import xlrd
from openpyxl import load_workbook
from pypdf import PdfReader


@dataclass
class SourceDocumentResult:
    file_name: str
    extracted_text: str = ""
    error: str = ""
    parser: str = ""
    material_type: str = ""


@dataclass
class SourceDocumentsExtraction:
    status: str
    combined_text: str = ""
    note: str = ""
    document_results: list[SourceDocumentResult] = field(default_factory=list)


def extract_supported_documents(file_paths: list[Path]) -> SourceDocumentsExtraction:
    supported = [path for path in file_paths if _is_supported_document(path)]
    if not supported:
        return SourceDocumentsExtraction(status="not_requested", note="当前草稿没有可解析的文档附件。")

    results: list[SourceDocumentResult] = []
    combined_parts: list[str] = []
    for path in supported:
        result = _extract_single_document(path)
        results.append(result)
        if result.extracted_text:
            combined_parts.append(result.extracted_text)

    success_count = sum(1 for item in results if item.extracted_text)
    if success_count == 0:
        return SourceDocumentsExtraction(
            status="empty",
            note="已检测到文档附件，但当前没有提取出稳定文本；请在草稿页人工补充。",
            document_results=results,
        )

    status = "success" if success_count == len(results) else "partial"
    note = "已从 PDF / 表格 / 文档附件中提取文字，结果会并入草稿解析。"
    if status == "partial":
        note = "部分文档已提取到文字，部分仍需人工补充；草稿已保留原始附件。"
    return SourceDocumentsExtraction(
        status=status,
        combined_text="\n\n".join(part for part in combined_parts if part.strip()),
        note=note,
        document_results=results,
    )


def _extract_single_document(path: Path) -> SourceDocumentResult:
    suffix = path.suffix.lower()
    try:
        if suffix == ".pdf":
            text = _extract_pdf_text(path)
            return _document_result(path, text, parser="pypdf")
        if suffix == ".xlsx":
            text = _extract_xlsx_text(path)
            return _document_result(path, text, parser="openpyxl")
        if suffix == ".xls":
            text = _extract_xls_text(path)
            return _document_result(path, text, parser="xlrd")
        if suffix in {".txt", ".csv", ".tsv", ".md"}:
            text = _extract_plain_text(path)
            return _document_result(path, text, parser="plain_text")
        if suffix == ".docx":
            text = _extract_docx_text(path)
            return _document_result(path, text, parser="docx_xml")
        if suffix == ".doc":
            text = _extract_doc_text(path)
            return _document_result(path, text, parser="doc_best_effort")
        if suffix == ".zip":
            text = _extract_zip_text(path)
            return _document_result(path, text, parser="zip_bundle")
        if suffix == ".7z":
            return SourceDocumentResult(file_name=path.name, error="7z 压缩包当前不能直接解析；请先解压后上传里面的 Excel/PDF/Word/图片。", parser="archive_notice", material_type="压缩包需解压")
    except Exception as exc:  # noqa: BLE001
        return SourceDocumentResult(file_name=path.name, error=f"{type(exc).__name__}: {exc}", material_type=_classify_material_type(path, ""))
    return SourceDocumentResult(file_name=path.name, error="当前附件格式暂未接入解析。", material_type="暂不支持材料")



def _document_result(path: Path, text: str, *, parser: str) -> SourceDocumentResult:
    return SourceDocumentResult(
        file_name=path.name,
        extracted_text=text,
        parser=parser,
        material_type=_classify_material_type(path, text, parser=parser),
    )



def _classify_material_type(path: Path, text: str, *, parser: str = "") -> str:
    name = path.name.lower()
    suffix = path.suffix.lower()
    compact = re.sub(r"\s+", "", text or "")
    if suffix == ".zip":
        return "压缩包材料"
    if suffix == ".7z":
        return "压缩包需解压"
    if suffix in {".xls", ".xlsx", ".csv", ".tsv"}:
        if re.search(r"(余额表|收入|流水|银行流水|银行明细|his|对账)", name + compact, re.IGNORECASE):
            return "财务流水/余额线索"
        if re.search(r"(车架号|车辆识别代号|合格证|机动车|VIN)", text or "", re.IGNORECASE):
            return "机动车异常表"
        if "开票联络函" in compact or "发票申请" in compact or "开发票" in compact:
            return "开票联络函/申请表"
        if "合同" in name or "合同" in compact:
            return "合同/清单 Excel"
        if _looks_like_tax_history_excel(compact):
            return "税局历史 Excel"
        return "本次开票样单 Excel"
    if suffix == ".pdf":
        if _looks_like_invoice_pdf(compact):
            return "样票 PDF"
        if "合同" in name or "合同" in compact:
            return "合同 PDF"
        return "PDF 材料"
    if suffix in {".doc", ".docx"}:
        if re.search(r"(开发票信息|开票信息|单位名称|纳税识别号|纳税人识别号|统一社会信用代码)", compact):
            return "客户开票信息 Word"
        return "Word 材料"
    if suffix in {".txt", ".md"}:
        if re.search(r"(微信|群聊|聊天|客户说|麻烦|帮忙|开票|发票)", compact):
            return "群文本/整理文本"
        return "文本材料"
    return parser or "文档材料"



def _looks_like_tax_history_excel(compact: str) -> bool:
    markers = ["发票号码", "开票日期", "销售方", "购买方", "税收分类编码", "价税合计"]
    return sum(1 for marker in markers if marker in compact) >= 4



def _looks_like_invoice_pdf(compact: str) -> bool:
    markers = ["发票号码", "价税合计", "购买方", "销售方", "税额", "电子发票"]
    return sum(1 for marker in markers if marker in compact) >= 3



def _extract_pdf_text(path: Path) -> str:
    reader = PdfReader(str(path))
    parts: list[str] = []
    for page in reader.pages:
        text = page.extract_text() or ""
        text = text.strip()
        if text:
            parts.append(text)
    return "\n\n".join(parts)


def _extract_xlsx_text(path: Path) -> str:
    workbook = load_workbook(path, data_only=True)
    parts: list[str] = []
    for sheet in workbook.worksheets:
        rows: list[tuple[int, list[str], bool]] = []
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            cells = ["" if cell is None else str(cell).strip() for cell in row]
            if not any(cells):
                continue
            rows.append((row_index, cells, bool(sheet.row_dimensions[row_index].hidden)))
        if not rows:
            continue

        visible_rows = [cells for _, cells, hidden in rows if not hidden]
        use_visible_only = any(hidden for _, _, hidden in rows) and len(visible_rows) >= 2 and len(visible_rows) < len(rows)
        selected_rows = visible_rows if use_visible_only else [cells for _, cells, _ in rows]
        lines = ["\t".join(cells) for cells in selected_rows[:200]]
        if lines:
            parts.append("\n".join(lines[:200]))
    return "\n\n".join(parts)


def _extract_xls_text(path: Path) -> str:
    workbook = xlrd.open_workbook(path)
    parts: list[str] = []
    for sheet in workbook.sheets():
        lines: list[str] = []
        for row_index in range(min(sheet.nrows, 200)):
            row = sheet.row_values(row_index)
            cells = ["" if cell is None else str(cell).strip() for cell in row]
            if not any(cells):
                continue
            lines.append("\t".join(cells))
        if lines:
            parts.append("\n".join(lines))
    return "\n\n".join(parts)


def _extract_plain_text(path: Path) -> str:
    raw = path.read_bytes()
    for encoding in ["utf-8", "utf-8-sig", "gb18030", "gbk"]:
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("utf-8", errors="ignore")

    if path.suffix.lower() == ".csv":
        reader = csv.reader(io.StringIO(text))
        return "\n".join("\t".join(cell.strip() for cell in row) for row in reader if any(cell.strip() for cell in row))
    return text


def _extract_zip_text(path: Path) -> str:
    parts: list[str] = []
    with tempfile.TemporaryDirectory(prefix="invoice-zip-docs-") as temp_dir_raw:
        temp_dir = Path(temp_dir_raw)
        with zipfile.ZipFile(path) as archive:
            members = [member for member in archive.infolist() if not member.is_dir()]
            for member in members[:30]:
                name = Path(member.filename).name
                if not name or name.startswith("."):
                    continue
                suffix = Path(name).suffix.lower()
                if suffix not in {".pdf", ".xlsx", ".xls", ".txt", ".csv", ".tsv", ".md", ".docx", ".doc"}:
                    continue
                target = temp_dir / name
                target.write_bytes(archive.read(member))
                result = _extract_single_document(target)
                if result.extracted_text:
                    parts.append(f"[{member.filename}]\n{result.extracted_text}")
                elif result.error:
                    parts.append(f"[{member.filename}]\n解析提示：{result.error}")
    return "\n\n".join(parts)



def _extract_docx_text(path: Path) -> str:
    with zipfile.ZipFile(path) as archive:
        xml_payload = archive.read("word/document.xml")
    root = ET.fromstring(xml_payload)
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs: list[str] = []
    for paragraph in root.findall(".//w:p", namespace):
        texts = [node.text for node in paragraph.findall(".//w:t", namespace) if node.text]
        if texts:
            paragraphs.append("".join(texts).strip())
    return "\n".join(part for part in paragraphs if part)



def _extract_doc_text(path: Path) -> str:
    # Some customer files are actually OOXML/ZIP documents with a .doc suffix.
    # Try the structured reader first; otherwise fall back to a conservative
    # printable-text harvest from legacy OLE Word binaries.
    try:
        return _extract_docx_text(path)
    except Exception:
        pass
    raw = path.read_bytes()
    candidates: list[str] = []
    for encoding in ["utf-16le", "gb18030", "gbk", "utf-8"]:
        decoded = raw.decode(encoding, errors="ignore")
        chunks = _readable_doc_chunks(decoded)
        if chunks:
            candidates.append("\n".join(chunks))
    if not candidates:
        return ""
    return max(candidates, key=_doc_text_score)



def _readable_doc_chunks(text: str) -> list[str]:
    import re

    chunks = []
    for chunk in re.findall(r"[\u4e00-\u9fffA-Za-z0-9（）()：:，,。./_\-\s]{4,}", text):
        cleaned = re.sub(r"[ \t\r\f\v]+", " ", chunk).strip()
        cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
        if not cleaned:
            continue
        if _looks_like_doc_binary_noise(cleaned):
            continue
        chunks.append(cleaned)
    return chunks[:120]



def _looks_like_doc_binary_noise(value: str) -> bool:
    noise_markers = ["Content_Types", "theme/theme", "xml version", "Microsoft Office Word", "Root Entry", "WordDocument"]
    if any(marker in value for marker in noise_markers):
        return True
    chinese_count = sum(1 for char in value if "\u4e00" <= char <= "\u9fff")
    digit_count = sum(1 for char in value if char.isdigit())
    return chinese_count == 0 and digit_count < 6 and len(value) > 30



def _doc_text_score(value: str) -> int:
    import re

    field_markers = ["单位名称", "纳税识别号", "纳税人识别号", "统一社会信用代码", "开户行", "账号", "电话", "地址", "开票"]
    marker_score = sum(10000 for marker in field_markers if marker in value)
    tax_id_score = 5000 if re.search(r"[0-9A-Z]{15,20}", value.upper()) else 0
    chinese_count = sum(1 for char in value if "\u4e00" <= char <= "\u9fff")
    digit_count = sum(1 for char in value if char.isdigit())
    noise_penalty = sum(3000 for marker in ["Content_Types", "theme/theme", "Root Entry", "WordDocument"] if marker in value)
    return marker_score + tax_id_score + chinese_count * 3 + digit_count - noise_penalty



def _is_supported_document(path: Path) -> bool:
    return path.suffix.lower() in {".pdf", ".xlsx", ".xls", ".txt", ".csv", ".tsv", ".md", ".docx", ".doc", ".zip", ".7z"}


def serialize_document_results(extraction: SourceDocumentsExtraction) -> str:
    return json.dumps(
        {
            "status": extraction.status,
            "note": extraction.note,
            "document_results": [
                {
                    "file_name": item.file_name,
                    "parser": item.parser,
                    "material_type": item.material_type,
                    "error": item.error,
                    "has_text": bool(item.extracted_text),
                }
                for item in extraction.document_results
            ],
        },
        ensure_ascii=False,
        indent=2,
    )
