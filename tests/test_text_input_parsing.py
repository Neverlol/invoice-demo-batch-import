import csv
import io
import json
import tempfile
import unittest
from pathlib import Path
import os
from unittest.mock import patch

from tax_invoice_demo.models import BuyerInfo, InvoiceDraft, InvoiceLine
from tax_invoice_demo.parsing import extract_buyer_info_from_text, extract_invoice_lines_from_text
import tax_invoice_demo.case_events as case_events_module
import tax_invoice_demo.coding_library as coding_library_module
import tax_invoice_demo.ledger as ledger_module
import tax_invoice_demo.tax_rule_engine as tax_rule_engine_module
import tax_invoice_demo.workbench as workbench_module
import tax_invoice_demo.customer_profiles as customer_profiles_module
import tax_invoice_batch_demo.lean_workbench as lean_workbench_module
from openpyxl import Workbook, load_workbook
from werkzeug.datastructures import FileStorage


SIMPLE_TEXT_INPUT = """购买方名称：黑龙江芃领飞象网络科技有限公司
纳税人识别号：91230109MAK8RY0867
购买方地址：黑龙江省哈尔滨市松北区创新一路733号哈尔滨国际金融大厦14层4号办公
购买方开户银行：招商银行股份有限公司哈尔滨松北支行
银行账号：45190935610000

发票类型：普通发票
税率：3%

开票明细：
1. 放学乐大菠萝，规格型号：40支/箱，单位：箱，数量：59，含税金额：3540
2. 放学乐大蜜瓜，规格型号：40支/箱，单位：箱，数量：62，含税金额：3801.16
3. 放学乐葚是喜欢，规格型号：32支/箱，单位：箱，数量：14，含税金额：840
"""

MINIMAL_TEXT_INPUT = """辽宁恒润电力科技有限公司
91210102MABWM3X12T
500
普票
代理记账和税务申报
"""

MINIMAL_TEXT_INPUT_WITH_TAX_CATEGORY = """辽宁恒润电力科技有限公司
91210102MABWM3X12T
500
普票
纳税申报代办*代理记账和税务申报
"""

MINIMAL_INLINE_TEXT_INPUT = "辽宁恒润电力科技有限公司 91210102MABWM3X12T 开普票 代理记账和税务申报 500元"

MINIMAL_LABELED_INLINE_TEXT_INPUT = "给辽宁恒润电力科技有限公司开普通发票，税号91210102MABWM3X12T，项目：代理记账和税务申报，金额500"

MINIMAL_INLINE_TEXT_INPUT_WITH_RATE = "辽宁恒润电力科技有限公司 91210102MABWM3X12T 开普票 代理记账和税务申报 500元 6%"


class TextInputParsingTest(unittest.TestCase):
    def setUp(self):
        coding_library_module.load_formal_coding_library.cache_clear()
        self.tempdir = tempfile.TemporaryDirectory()
        self.temp_path = Path(self.tempdir.name)
        self.old_workbench_root = workbench_module.WORKBENCH_ROOT
        self.old_ledger_paths = (
            ledger_module.LEDGER_ROOT,
            ledger_module.LEDGER_CSV_PATH,
            ledger_module.LEDGER_XLSX_PATH,
            ledger_module.FEEDBACK_CSV_PATH,
        )
        self.old_batch_output_root = lean_workbench_module.BATCH_OUTPUT_ROOT
        self.old_success_paths = (
            lean_workbench_module.SUCCESS_LEDGER_CSV,
            lean_workbench_module.SUCCESS_LEDGER_XLSX,
        )
        self.old_event_root = case_events_module.EVENT_ROOT
        self.old_learned_rules_path = tax_rule_engine_module.LEARNED_RULES_PATH
        self.old_tenant_rules_path = tax_rule_engine_module.TENANT_RULES_PATH
        self.old_llm_tax_code_cache_path = tax_rule_engine_module.LLM_TAX_CODE_CACHE_PATH
        self.old_profile_cache_path = customer_profiles_module.PROFILE_CACHE_PATH
        self.old_sync_endpoint = os.environ.get("TAX_INVOICE_SYNC_ENDPOINT")
        self.old_sync_token = os.environ.get("TAX_INVOICE_SYNC_TOKEN")
        self.old_sync_enabled = os.environ.get("TAX_INVOICE_SYNC_ENABLED")
        self.old_operator = os.environ.get("TAX_INVOICE_OPERATOR")

        workbench_module.WORKBENCH_ROOT = self.temp_path / "workbench"
        ledger_module.LEDGER_ROOT = self.temp_path / "ledger"
        ledger_module.LEDGER_CSV_PATH = ledger_module.LEDGER_ROOT / "累计发票明细表.csv"
        ledger_module.LEDGER_XLSX_PATH = ledger_module.LEDGER_ROOT / "累计发票明细表.xlsx"
        ledger_module.FEEDBACK_CSV_PATH = ledger_module.LEDGER_ROOT / "赋码反馈候选池.csv"
        lean_workbench_module.BATCH_OUTPUT_ROOT = self.temp_path / "batch_import_preview"
        lean_workbench_module.SUCCESS_LEDGER_CSV = lean_workbench_module.BATCH_OUTPUT_ROOT / "批量导入成功明细.csv"
        lean_workbench_module.SUCCESS_LEDGER_XLSX = lean_workbench_module.BATCH_OUTPUT_ROOT / "批量导入成功明细.xlsx"
        case_events_module.EVENT_ROOT = self.temp_path / "events"
        tax_rule_engine_module.LEARNED_RULES_PATH = self.temp_path / "ledger" / "本地即时学习赋码规则.csv"
        tax_rule_engine_module.TENANT_RULES_PATH = self.temp_path / "ledger" / "客户同步赋码规则.csv"
        tax_rule_engine_module.LLM_TAX_CODE_CACHE_PATH = self.temp_path / "ledger" / "智能税码推荐缓存.json"
        customer_profiles_module.PROFILE_CACHE_PATH = self.temp_path / "workbench" / "客户档案缓存.json"
        tax_rule_engine_module.load_tenant_coding_library.cache_clear()
        tax_rule_engine_module.load_learned_coding_library.cache_clear()
        os.environ.pop("TAX_INVOICE_SYNC_ENDPOINT", None)
        os.environ.pop("TAX_INVOICE_SYNC_TOKEN", None)
        os.environ["TAX_INVOICE_SYNC_ENABLED"] = "0"
        os.environ["TAX_INVOICE_OPERATOR"] = "seed-assistant"

    def tearDown(self):
        workbench_module.WORKBENCH_ROOT = self.old_workbench_root
        (
            ledger_module.LEDGER_ROOT,
            ledger_module.LEDGER_CSV_PATH,
            ledger_module.LEDGER_XLSX_PATH,
            ledger_module.FEEDBACK_CSV_PATH,
        ) = self.old_ledger_paths
        lean_workbench_module.BATCH_OUTPUT_ROOT = self.old_batch_output_root
        (
            lean_workbench_module.SUCCESS_LEDGER_CSV,
            lean_workbench_module.SUCCESS_LEDGER_XLSX,
        ) = self.old_success_paths
        case_events_module.EVENT_ROOT = self.old_event_root
        tax_rule_engine_module.LEARNED_RULES_PATH = self.old_learned_rules_path
        tax_rule_engine_module.TENANT_RULES_PATH = self.old_tenant_rules_path
        tax_rule_engine_module.LLM_TAX_CODE_CACHE_PATH = self.old_llm_tax_code_cache_path
        customer_profiles_module.PROFILE_CACHE_PATH = self.old_profile_cache_path
        tax_rule_engine_module.load_tenant_coding_library.cache_clear()
        tax_rule_engine_module.load_learned_coding_library.cache_clear()
        if self.old_sync_endpoint is None:
            os.environ.pop("TAX_INVOICE_SYNC_ENDPOINT", None)
        else:
            os.environ["TAX_INVOICE_SYNC_ENDPOINT"] = self.old_sync_endpoint
        if self.old_sync_token is None:
            os.environ.pop("TAX_INVOICE_SYNC_TOKEN", None)
        else:
            os.environ["TAX_INVOICE_SYNC_TOKEN"] = self.old_sync_token
        if self.old_sync_enabled is None:
            os.environ.pop("TAX_INVOICE_SYNC_ENABLED", None)
        else:
            os.environ["TAX_INVOICE_SYNC_ENABLED"] = self.old_sync_enabled
        if self.old_operator is None:
            os.environ.pop("TAX_INVOICE_OPERATOR", None)
        else:
            os.environ["TAX_INVOICE_OPERATOR"] = self.old_operator
        self.tempdir.cleanup()
        coding_library_module.load_formal_coding_library.cache_clear()

    def test_simple_text_input_extracts_buyer_and_detail_lines(self):
        buyer = extract_buyer_info_from_text(SIMPLE_TEXT_INPUT)
        lines = extract_invoice_lines_from_text(SIMPLE_TEXT_INPUT)

        self.assertEqual(buyer.name, "黑龙江芃领飞象网络科技有限公司")
        self.assertEqual(buyer.tax_id, "91230109MAK8RY0867")
        self.assertEqual(buyer.bank_name, "招商银行股份有限公司哈尔滨松北支行")
        self.assertEqual(buyer.bank_account, "45190935610000")
        self.assertEqual(len(lines), 3)
        self.assertEqual(lines[0].project_name, "放学乐大菠萝")
        self.assertEqual(lines[0].specification, "40支/箱")
        self.assertEqual(lines[0].unit, "箱")
        self.assertEqual(lines[0].quantity, "59")
        self.assertEqual(lines[0].amount_with_tax, "3540")
        self.assertEqual(lines[0].tax_category, "")
        self.assertEqual(lines[0].tax_code, "")
        self.assertEqual(lines[0].normalized_tax_rate(), "3%")

    def test_real_business_excel_style_table_extracts_buyer_lines_and_global_tax_code(self):
        text = """物料名称\t规格型号\t数量\t开票金额\t税率
一次性使用微针电极\tCONSUMABLE TIP\t150\t35625\t0.13
一次性使用微针电极\tJ18BS\t1\t433.03849999999994\t0.13
一次性使用微针电极\tJ25BM\t3282\t1061785.36\t0.13
一次性使用微针电极\tJ25BS\t1\t428.76349999999996\t0.13
公司名称\t河北雅之颜医药有限公司
税号\t91130101MA7MB4FA89
税收编码\t1090245030000000000
发票类型\t增票
"""

        buyer = extract_buyer_info_from_text(text)
        lines = extract_invoice_lines_from_text(text)
        profile = workbench_module._infer_invoice_profile(text)

        self.assertEqual(buyer.name, "河北雅之颜医药有限公司")
        self.assertEqual(buyer.tax_id, "91130101MA7MB4FA89")
        self.assertEqual(len(lines), 4)
        self.assertEqual(lines[0].project_name, "一次性使用微针电极")
        self.assertEqual(lines[0].specification, "CONSUMABLE TIP")
        self.assertEqual(lines[0].quantity, "150")
        self.assertEqual(lines[0].resolved_amount_with_tax(), "35625.00")
        self.assertEqual(lines[0].normalized_tax_rate(), "13%")
        self.assertTrue(all(line.tax_code == "1090245030000000000" for line in lines))
        self.assertEqual(profile["invoice_kind"], "增值税专用发票")

    def test_real_business_excel_upload_creates_usable_draft_without_llm(self):
        workbook_path = self.temp_path / "测试.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet3"
        sheet.append(["物料名称", "规格型号", "数量", "开票金额", "税率"])
        sheet.append(["一次性使用微针电极", "CONSUMABLE TIP", 150, 35625, 0.13])
        sheet.append(["一次性使用微针电极", "J18BS", 1, 433.0385, 0.13])
        sheet.append(["一次性使用微针电极", "J25BM", 3282, 1061785.36, 0.13])
        sheet.append(["一次性使用微针电极", "J25BS", 1, 428.7635, 0.13])
        sheet.append([])
        sheet.append(["公司名称", "河北雅之颜医药有限公司"])
        sheet.append(["税号", "91130101MA7MB4FA89"])
        sheet.append(["税收编码", "1090245030000000000"])
        sheet.append(["发票类型", "增票"])
        workbook.save(workbook_path)

        with workbook_path.open("rb") as handle:
            draft = workbench_module.create_draft_from_workbench(
                "吉林省风生水起商贸有限公司",
                "",
                "",
                [FileStorage(stream=handle, filename="测试.xlsx")],
            )

        self.assertEqual(draft.buyer.name, "河北雅之颜医药有限公司")
        self.assertEqual(draft.buyer.tax_id, "91130101MA7MB4FA89")
        self.assertEqual(draft.invoice_kind, "增值税专用发票")
        self.assertEqual(len(draft.lines), 4)
        self.assertEqual(draft.lines[0].project_name, "一次性使用微针电极")
        self.assertEqual(draft.lines[0].specification, "CONSUMABLE TIP")
        self.assertEqual(draft.lines[0].normalized_tax_rate(), "13%")
        self.assertTrue(all(line.tax_code == "1090245030000000000" for line in draft.lines))
        self.assertFalse(any("当前还没自动识别出开票明细" in issue for issue in draft.issues))

    def test_llm_tax_code_recommendation_writes_draft_fields_with_review_label(self):
        text = """购买方名称：河北雅之颜医药有限公司
税号：91130101MA7MB4FA89
发票类型：增票
税率：13%
开票明细：
1. 一次性使用微针电极，规格型号：CONSUMABLE TIP，数量：150，金额：35625
2. 一次性使用微针电极，规格型号：J25BM，数量：2，金额：200
"""

        fake_adapter = _FakeTaxCodeLLMAdapter()
        with patch.object(tax_rule_engine_module, "get_llm_adapter", return_value=fake_adapter):
            draft = workbench_module.create_draft_from_workbench("吉林省风生水起商贸有限公司", text, "", [])

        self.assertEqual([line.tax_category for line in draft.lines], ["医疗仪器器械", "医疗仪器器械"])
        self.assertEqual([line.tax_code for line in draft.lines], ["1090245030000000000", "1090245030000000000"])
        self.assertTrue(all(line.coding_reference.startswith("智能推荐，需人工复核") for line in draft.lines))
        self.assertEqual(fake_adapter.call_count, 1)
        self.assertTrue(tax_rule_engine_module.LLM_TAX_CODE_CACHE_PATH.exists())

        cached_fake_adapter = _FakeTaxCodeLLMAdapter()
        with patch.object(tax_rule_engine_module, "get_llm_adapter", return_value=cached_fake_adapter):
            cached_draft = workbench_module.create_draft_from_workbench("吉林省风生水起商贸有限公司", text, "", [])

        self.assertEqual(cached_fake_adapter.call_count, 0)
        self.assertEqual([line.tax_code for line in cached_draft.lines], ["1090245030000000000", "1090245030000000000"])

    def test_simple_text_input_is_enriched_by_backend_coding_library(self):
        draft = workbench_module.create_draft_from_workbench("吉林省风生水起商贸有限公司", SIMPLE_TEXT_INPUT, "", [])

        self.assertEqual(draft.case_id, draft.draft_id)
        self.assertEqual(len(draft.lines), 3)
        self.assertEqual([line.tax_category for line in draft.lines], ["冷冻饮品", "冷冻饮品", "冷冻饮品"])
        self.assertEqual(
            [line.tax_code for line in draft.lines],
            ["1030209990000000000", "1030209990000000000", "1030209990000000000"],
        )
        self.assertTrue(all(line.coding_reference.startswith("命中 ") for line in draft.lines))
        self.assertEqual([line.normalized_tax_rate() for line in draft.lines], ["3%", "3%", "3%"])

    def test_simple_text_input_builds_valid_batch_template(self):
        draft = workbench_module.create_draft_from_workbench("吉林省风生水起商贸有限公司", SIMPLE_TEXT_INPUT, "", [])
        export = lean_workbench_module.export_draft_template(draft)

        self.assertEqual(draft.invoice_kind, "普通发票")
        self.assertEqual(len(draft.lines), 3)
        self.assertEqual(export["error_count"], 0)

    def test_numbered_lines_without_detail_header_are_supported(self):
        text = """购买方名称：测试购买方有限公司
纳税人识别号：91230100MA00000000
发票类型：普通发票
税率：0.03
1、项目名称：技术服务费，单位：项，数量：1，单价：500元，价税合计：500元，税收分类编码：3049900000000000000
2、开票项目：咨询服务，数量：2项，金额：1,000元，商品和服务分类简称：现代服务，商品和服务税收编码：3049900000000000000
"""
        lines = extract_invoice_lines_from_text(text)

        self.assertEqual(len(lines), 2)
        self.assertEqual(lines[0].project_name, "技术服务费")
        self.assertEqual(lines[0].unit_price, "500")
        self.assertEqual(lines[0].amount_with_tax, "500")
        self.assertEqual(lines[0].tax_code, "3049900000000000000")
        self.assertEqual(lines[0].normalized_tax_rate(), "3%")
        self.assertEqual(lines[1].project_name, "咨询服务")
        self.assertEqual(lines[1].quantity, "2")
        self.assertEqual(lines[1].unit, "项")
        self.assertEqual(lines[1].amount_with_tax, "1000")
        self.assertEqual(lines[1].tax_category, "现代服务")

    def test_labeled_type_normal_invoice_overrides_special_invoice_heuristic(self):
        draft = workbench_module.create_draft_from_workbench("吉林省风生水起商贸有限公司", SIMPLE_TEXT_INPUT, "", [])

        self.assertEqual(draft.invoice_kind, "普通发票")

    def test_explicit_three_percent_is_exported_as_decimal_not_overridden_by_coding_library(self):
        draft = workbench_module.create_draft_from_workbench("吉林省风生水起商贸有限公司", SIMPLE_TEXT_INPUT, "", [])
        export = lean_workbench_module.export_draft_template(draft)
        workbook = load_workbook(export["output_path"], data_only=True)
        sheet = workbook["2-发票明细信息"]
        headers = [str(cell.value or "").strip() for cell in sheet[3]]
        rate_col = headers.index("税率") + 1
        tax_code_col = headers.index("商品和服务税收编码") + 1
        rates = [str(sheet.cell(row=row, column=rate_col).value) for row in range(4, 7)]
        tax_codes = [str(sheet.cell(row=row, column=tax_code_col).value) for row in range(4, 7)]
        workbook.close()

        self.assertEqual([line.normalized_tax_rate() for line in draft.lines], ["3%", "3%", "3%"])
        self.assertEqual(rates, ["0.03", "0.03", "0.03"])
        self.assertEqual(tax_codes, ["1030209990000000000", "1030209990000000000", "1030209990000000000"])

    def test_one_percent_preview_tax_amount_is_not_treated_as_hundred_percent(self):
        text = """购买方名称：辽宁恒润电力科技有限公司
纳税人识别号：91210102MABWM3X12T
发票类型：普通发票
税率：1%
开票明细：
1. 纸制文具及用品，规格型号：A4资料册，单位：批，数量：1，含税金额：300.00
2. 文件夹，规格型号：办公用，单位：批，数量：1，含税金额：120.00
3. 文件架，规格型号：办公用，单位：批，数量：1，含税金额：80.00
"""
        draft = workbench_module.create_draft_from_workbench("吉林省风生水起商贸有限公司", text, "", [])
        preview = lean_workbench_module.draft_preview(draft)

        self.assertEqual(preview["amount_total"], "500.00")
        self.assertEqual(preview["tax_total"], "4.95")
        self.assertEqual([row["tax_rate"] for row in preview["line_rows"]], ["1%", "1%", "1%"])

    def test_minimal_five_line_text_input_extracts_buyer_and_detail_line(self):
        buyer = extract_buyer_info_from_text(MINIMAL_TEXT_INPUT)
        lines = extract_invoice_lines_from_text(MINIMAL_TEXT_INPUT)

        self.assertEqual(buyer.name, "辽宁恒润电力科技有限公司")
        self.assertEqual(buyer.tax_id, "91210102MABWM3X12T")
        self.assertEqual(len(lines), 1)
        self.assertEqual(lines[0].project_name, "代理记账和税务申报")
        self.assertEqual(lines[0].amount_with_tax, "500")
        self.assertEqual(lines[0].unit, "项")
        self.assertEqual(lines[0].quantity, "1")
        self.assertEqual(lines[0].unit_price, "500")

        draft = workbench_module.create_draft_from_workbench("吉林省风生水起商贸有限公司", MINIMAL_TEXT_INPUT, "", [])

        self.assertEqual(draft.invoice_kind, "普通发票")
        self.assertEqual(draft.buyer.name, "辽宁恒润电力科技有限公司")
        self.assertEqual(draft.buyer.tax_id, "91210102MABWM3X12T")
        self.assertEqual(len(draft.lines), 1)
        self.assertEqual(draft.lines[0].project_name, "代理记账和税务申报")
        self.assertEqual(draft.lines[0].tax_category, "")
        self.assertEqual(draft.lines[0].tax_code, "")
        self.assertEqual(draft.lines[0].normalized_tax_rate(), "3%")

    def test_daily_wechat_text_extracts_office_items_and_ignores_delivery_note(self):
        old_provider = os.environ.get("TAX_INVOICE_LLM_PROVIDER")
        os.environ["TAX_INVOICE_LLM_PROVIDER"] = "off"
        text = """麻烦开个普票哈

辽宁恒润电力科技有限公司
91210102MABWM3X12T

这次是办公室用的一些东西，明细大概这样：
A4复印纸 10包 240元
蓝色文件夹 20个 160元
桌面文件架 5个 100元

总共500，按1个点开就行
备注不用写，电子票发我微信
"""
        try:
            draft = workbench_module.create_draft_from_workbench("吉林省风生水起商贸有限公司", text, "", [])
        finally:
            if old_provider is None:
                os.environ.pop("TAX_INVOICE_LLM_PROVIDER", None)
            else:
                os.environ["TAX_INVOICE_LLM_PROVIDER"] = old_provider

        self.assertEqual(draft.invoice_kind, "普通发票")
        self.assertEqual(draft.buyer.name, "辽宁恒润电力科技有限公司")
        self.assertEqual(draft.buyer.tax_id, "91210102MABWM3X12T")
        self.assertEqual([line.project_name for line in draft.lines], ["A4复印纸", "蓝色文件夹", "桌面文件架"])
        self.assertEqual([line.amount_with_tax for line in draft.lines], ["240", "160", "100"])
        self.assertEqual([line.tax_rate for line in draft.lines], ["1%", "1%", "1%"])
        self.assertNotIn("电子票发我微信", [line.project_name for line in draft.lines])
        self.assertEqual(draft.lines[0].tax_code, "1060105020000000000")
        self.assertEqual(draft.lines[1].tax_code, "1060401020000000000")
        self.assertEqual(draft.lines[2].tax_code, "1060401030000000000")

    def test_mimo_ocr_medical_screenshot_text_maps_special_invoice_and_safe_codes(self):
        old_provider = os.environ.get("TAX_INVOICE_LLM_PROVIDER")
        os.environ["TAX_INVOICE_LLM_PROVIDER"] = "off"
        text = """辽宁瑞康医疗科技有限公司
91210103MAC7R8K26X
增票
税率：13%
开票明细：
1、一次性使用无菌注射器，规格型号：5ml 带针，单位：支，数量：800，含税金额：2400.00元
2、一次性使用静脉输液针，规格型号：0.7mm，单位：支，数量：1200，含税金额：1800.00元
3、医用检查手套，规格型号：丁腈 M号，单位：盒，数量：50，含税金额：1250.00元
辽宁省沈阳市和平区南京北街88号 024-23886666
中国建设银行沈阳和平支行 21050123456789012345
"""
        try:
            draft = workbench_module.create_draft_from_workbench("", text, "", [])
        finally:
            if old_provider is None:
                os.environ.pop("TAX_INVOICE_LLM_PROVIDER", None)
            else:
                os.environ["TAX_INVOICE_LLM_PROVIDER"] = old_provider

        self.assertEqual(draft.invoice_kind, "增值税专用发票")
        self.assertEqual(draft.buyer.name, "辽宁瑞康医疗科技有限公司")
        self.assertEqual(draft.buyer.tax_id, "91210103MAC7R8K26X")
        self.assertEqual(len(draft.lines), 3)
        self.assertEqual(draft.lines[0].tax_code, "1090245030000000000")
        self.assertEqual(draft.lines[1].tax_code, "1090245030000000000")
        self.assertEqual(draft.lines[2].tax_code, "1070508010000000000")
        self.assertTrue(all("需人工复核" in line.coding_reference for line in draft.lines))

    def test_platform_screenshot_text_creates_batch_with_seller_history_profile(self):
        seller = "哈尔滨市道里区庆成记隆江猪脚饭店（个体工商户）"
        ledger_module.sync_draft_to_ledger(
            InvoiceDraft(
                draft_id="history-food",
                case_id="history-food",
                company_name=seller,
                buyer=BuyerInfo(name="历史购买方", tax_id="91230102MAEMEM2G2M"),
                lines=[
                    InvoiceLine(
                        project_name="餐费",
                        amount_with_tax="86.20",
                        tax_rate="1%",
                        tax_category="餐饮服务",
                        tax_code="3070401000000000000",
                        unit="项",
                        quantity="1",
                        coding_reference="税局历史明细导入，需人工复核",
                    )
                ],
                created_at="2026-04-30T10:00:00",
            )
        )
        text = """[01.jpg]
发票详情
税号 91230102MAEMEM2G2M
联系人邮箱 2029794338@qq.com
建议开票金额 14.80
订单号: 8086410250960608559

[02.jpg]
发票详情
抬头 黑龙江源速商贸有限公司
税号 91230102MA1CDKE47Y
联系人邮箱 jsls7@163.com
建议开票金额 13.80
订单号: 8011766126836281741

[03.jpg]
发票详情
公司抬头 被平台截断的公司...
联系人邮箱 missing@example.com
建议开票金额 15.80
订单号: 8011766126836281742
"""

        batch = workbench_module.create_draft_from_workbench(seller, text, "平台截图批量测试", [])

        self.assertEqual(batch.__class__.__name__, "DraftBatch")
        self.assertEqual(len(batch.items), 3)
        first = workbench_module.load_draft(batch.items[0].draft_id)
        second = workbench_module.load_draft(batch.items[1].draft_id)
        third = workbench_module.load_draft(batch.items[2].draft_id)
        self.assertEqual(first.lines[0].project_name, "餐费")
        self.assertEqual(first.lines[0].tax_category, "餐饮服务")
        self.assertEqual(first.lines[0].tax_code, "3070401000000000000")
        self.assertEqual(first.lines[0].amount_with_tax, "14.80")
        self.assertTrue(any("购买方名称" in issue for issue in first.issues))
        self.assertEqual(second.buyer.name, "黑龙江源速商贸有限公司")
        self.assertEqual(second.lines[0].amount_with_tax, "13.80")
        self.assertEqual(third.lines[0].project_name, "餐费")
        self.assertEqual(third.lines[0].tax_code, "3070401000000000000")
        self.assertEqual(third.lines[0].amount_with_tax, "15.80")
        self.assertTrue(any("购买方税号" in issue for issue in third.issues))
        self.assertEqual(batch.items[2].buyer_name, "待补全购买方名称")

    def test_force_batch_mode_creates_one_child_draft_per_uploaded_image_even_without_ocr_fields(self):
        files = [
            FileStorage(stream=io.BytesIO(b"fake image one"), filename="01.png", content_type="image/png"),
            FileStorage(stream=io.BytesIO(b"fake image two"), filename="02.png", content_type="image/png"),
        ]

        batch = workbench_module.create_draft_from_workbench("", "", "批量截图模式", files, force_batch=True)

        self.assertEqual(batch.__class__.__name__, "DraftBatch")
        self.assertEqual(len(batch.items), 2)
        self.assertEqual(batch.items[0].buyer_name, "待补全购买方名称")
        first = workbench_module.load_draft(batch.items[0].draft_id)
        self.assertTrue(any("购买方名称" in issue for issue in first.issues))
        self.assertTrue(any("购买方税号" in issue for issue in first.issues))
        self.assertTrue(any("开票金额" in issue for issue in first.issues))
        self.assertEqual(first.lines[0].coding_reference, "批量模式待人工补全，需人工复核")

    def test_weak_chat_text_uses_history_profile_for_buyer_and_project(self):
        _seed_history_profile_row(
            company_name="吉林省风生水起商贸有限公司",
            buyer_name="辽宁恒润电力科技有限公司",
            buyer_tax_id="91210102MABWM3X12T",
            project_name="代理记账和税务申报",
            tax_category="纳税申报代理",
            tax_code="3040802050000000000",
            tax_rate="3%",
        )
        text = "帮开个票，辽宁恒润那个公司，代理服务，五百，普票"

        draft = workbench_module.create_draft_from_workbench("吉林省风生水起商贸有限公司", text, "", [])

        self.assertEqual(draft.invoice_kind, "普通发票")
        self.assertEqual(draft.buyer.name, "辽宁恒润电力科技有限公司")
        self.assertEqual(draft.buyer.tax_id, "91210102MABWM3X12T")
        self.assertEqual(len(draft.lines), 1)
        self.assertEqual(draft.lines[0].project_name, "代理记账和税务申报")
        self.assertEqual(draft.lines[0].amount_with_tax, "500")
        self.assertEqual(draft.lines[0].unit, "项")
        self.assertEqual(draft.lines[0].quantity, "1")
        self.assertEqual(draft.lines[0].tax_category, "纳税申报代理")
        self.assertEqual(draft.lines[0].tax_code, "3040802050000000000")
        self.assertIn("历史开票档案推荐", draft.lines[0].coding_reference)

    def test_minimal_inline_text_input_extracts_buyer_and_detail_line(self):
        buyer = extract_buyer_info_from_text(MINIMAL_INLINE_TEXT_INPUT)
        lines = extract_invoice_lines_from_text(MINIMAL_INLINE_TEXT_INPUT)

        self.assertEqual(buyer.name, "辽宁恒润电力科技有限公司")
        self.assertEqual(buyer.tax_id, "91210102MABWM3X12T")
        self.assertEqual(len(lines), 1)
        self.assertEqual(lines[0].project_name, "代理记账和税务申报")
        self.assertEqual(lines[0].amount_with_tax, "500")
        self.assertEqual(lines[0].normalized_tax_rate(), "3%")

        draft = workbench_module.create_draft_from_workbench("吉林省风生水起商贸有限公司", MINIMAL_INLINE_TEXT_INPUT, "", [])
        self.assertEqual(draft.invoice_kind, "普通发票")
        self.assertEqual(draft.buyer.name, "辽宁恒润电力科技有限公司")
        self.assertEqual(draft.buyer.tax_id, "91210102MABWM3X12T")
        self.assertEqual(len(draft.lines), 1)
        self.assertEqual(draft.lines[0].project_name, "代理记账和税务申报")
        self.assertEqual(draft.lines[0].resolved_amount_with_tax(), "500.00")

    def test_minimal_labeled_inline_text_input_extracts_buyer_and_detail_line(self):
        buyer = extract_buyer_info_from_text(MINIMAL_LABELED_INLINE_TEXT_INPUT)
        lines = extract_invoice_lines_from_text(MINIMAL_LABELED_INLINE_TEXT_INPUT)

        self.assertEqual(buyer.name, "辽宁恒润电力科技有限公司")
        self.assertEqual(buyer.tax_id, "91210102MABWM3X12T")
        self.assertEqual(len(lines), 1)
        self.assertEqual(lines[0].project_name, "代理记账和税务申报")
        self.assertEqual(lines[0].amount_with_tax, "500")

    def test_minimal_inline_text_input_respects_explicit_tax_rate(self):
        lines = extract_invoice_lines_from_text(MINIMAL_INLINE_TEXT_INPUT_WITH_RATE)

        self.assertEqual(len(lines), 1)
        self.assertEqual(lines[0].project_name, "代理记账和税务申报")
        self.assertEqual(lines[0].normalized_tax_rate(), "6%")

    def test_minimal_text_with_invoice_face_category_keeps_category_pending_code(self):
        draft = workbench_module.create_draft_from_workbench(
            "吉林省风生水起商贸有限公司",
            MINIMAL_TEXT_INPUT_WITH_TAX_CATEGORY,
            "",
            [],
        )
        export = lean_workbench_module.export_draft_template(draft)

        self.assertEqual(draft.invoice_kind, "普通发票")
        self.assertEqual(draft.buyer.name, "辽宁恒润电力科技有限公司")
        self.assertEqual(draft.buyer.tax_id, "91210102MABWM3X12T")
        self.assertEqual(len(draft.lines), 1)
        self.assertEqual(draft.lines[0].project_name, "代理记账和税务申报")
        self.assertEqual(draft.lines[0].tax_category, "纳税申报代办")
        self.assertEqual(draft.lines[0].tax_code, "")
        self.assertEqual(draft.lines[0].normalized_tax_rate(), "3%")
        self.assertTrue(any("税收编码" in issue or "正式赋码库" in issue for issue in draft.issues))
        self.assertGreater(export["error_count"], 0)

        workbook = load_workbook(export["output_path"], data_only=True)
        try:
            detail_sheet = workbook["2-发票明细信息"]
            headers = [str(cell.value or "").strip() for cell in detail_sheet[3]]
            detail_values = {
                header: str(detail_sheet.cell(row=4, column=index + 1).value or "")
                for index, header in enumerate(headers)
                if header
            }
            self.assertEqual(detail_values["项目名称"], "代理记账和税务申报")
            self.assertEqual(detail_values["商品和服务税收编码"], "")
            self.assertEqual(detail_values["金额"], "500.00")
            self.assertEqual(detail_values["税率"], "0.03")
        finally:
            workbook.close()

    def test_save_manual_draft_edits_does_not_rerun_extraction_or_llm(self):
        draft = workbench_module.create_draft_from_workbench("吉林省风生水起商贸有限公司", MINIMAL_TEXT_INPUT, "", [])
        form = _form_from_draft(
            draft,
            tax_category="纳税申报代理",
            tax_code="3040802050000000000",
            tax_rate="3%",
        )

        with patch.object(workbench_module, "extract_invoice_structured_data", side_effect=AssertionError("should not re-extract")), \
             patch.object(workbench_module, "_run_document_extraction", side_effect=AssertionError("should not parse attachments")), \
             patch.object(workbench_module, "_run_draft_ocr", side_effect=AssertionError("should not rerun OCR")), \
             patch.object(tax_rule_engine_module, "get_llm_adapter", side_effect=AssertionError("should not call LLM")):
            saved = lean_workbench_module.save_lean_draft_from_form(draft.draft_id, form, [])

        self.assertEqual(saved.lines[0].tax_code, "3040802050000000000")

    def test_manual_coding_fix_is_learned_immediately_for_next_draft(self):
        draft = workbench_module.create_draft_from_workbench("吉林省风生水起商贸有限公司", MINIMAL_TEXT_INPUT, "", [])
        self.assertEqual(draft.lines[0].tax_category, "")
        self.assertEqual(draft.lines[0].tax_code, "")

        form = _form_from_draft(
            draft,
            tax_category="纳税申报代办",
            tax_code="3040802050000000000",
            tax_rate="3%",
        )
        saved = lean_workbench_module.save_lean_draft_from_form(draft.draft_id, form, [])
        self.assertEqual(saved.lines[0].tax_category, "纳税申报代办")
        self.assertEqual(saved.lines[0].tax_code, "3040802050000000000")
        self.assertTrue(tax_rule_engine_module.LEARNED_RULES_PATH.exists())

        next_draft = workbench_module.create_draft_from_workbench("吉林省风生水起商贸有限公司", MINIMAL_TEXT_INPUT, "", [])
        self.assertEqual(next_draft.lines[0].tax_category, "纳税申报代办")
        self.assertEqual(next_draft.lines[0].tax_code, "3040802050000000000")
        self.assertEqual(next_draft.lines[0].normalized_tax_rate(), "3%")
        self.assertIn("命中 本地即时规则", next_draft.lines[0].coding_reference)

        rows = _read_csv_rows(tax_rule_engine_module.LEARNED_RULES_PATH)
        self.assertEqual(rows[0]["status"], "ready")
        self.assertEqual(rows[0]["source_operator"], "seed-assistant")
        self.assertEqual(rows[0]["original_project_name"], "代理记账和税务申报")
        self.assertEqual(rows[0]["final_project_name"], "代理记账和税务申报")

    def test_conflicting_manual_coding_fix_is_marked_pending_review(self):
        draft = workbench_module.create_draft_from_workbench("吉林省风生水起商贸有限公司", MINIMAL_TEXT_INPUT, "", [])
        lean_workbench_module.save_lean_draft_from_form(
            draft.draft_id,
            _form_from_draft(
                draft,
                tax_category="纳税申报代办",
                tax_code="3040802050000000000",
                tax_rate="3%",
            ),
            [],
        )

        conflicting_draft = workbench_module.create_draft_from_workbench("吉林省风生水起商贸有限公司", MINIMAL_TEXT_INPUT, "", [])
        lean_workbench_module.save_lean_draft_from_form(
            conflicting_draft.draft_id,
            _form_from_draft(
                conflicting_draft,
                tax_category="错误分类待审核",
                tax_code="1000000000000000000",
                tax_rate="13%",
            ),
            [],
        )

        rows = _read_csv_rows(tax_rule_engine_module.LEARNED_RULES_PATH)
        self.assertEqual([row["status"] for row in rows], ["ready", "pending_review"])
        self.assertEqual(rows[1]["raw_alias"], "代理记账和税务申报")
        self.assertEqual(rows[1]["conflict_with_rule_id"], rows[0]["rule_id"])
        self.assertIn("冲突待审核", rows[1]["decision_basis"])

        next_draft = workbench_module.create_draft_from_workbench("吉林省风生水起商贸有限公司", MINIMAL_TEXT_INPUT, "", [])
        self.assertEqual(next_draft.lines[0].tax_category, "纳税申报代办")
        self.assertEqual(next_draft.lines[0].tax_code, "3040802050000000000")
        self.assertIn("命中 本地即时规则", next_draft.lines[0].coding_reference)

    def test_synced_tenant_rules_override_local_learned_rules(self):
        draft = workbench_module.create_draft_from_workbench("吉林省风生水起商贸有限公司", MINIMAL_TEXT_INPUT, "", [])
        saved = lean_workbench_module.save_lean_draft_from_form(
            draft.draft_id,
            _form_from_draft(
                draft,
                tax_category="本地错误分类",
                tax_code="1000000000000000000",
                tax_rate="3%",
            ),
            [],
        )
        self.assertEqual(saved.lines[0].tax_category, "本地错误分类")

        tax_rule_engine_module.write_tenant_rule_package(
            [
                {
                    "raw_alias": "代理记账和税务申报",
                    "normalized_invoice_name": "代理记账和税务申报",
                    "tax_category": "云端审核分类",
                    "tax_code": "3040802050000000000",
                    "tax_treatment_or_rate": "0.03",
                }
            ],
            package_id="rules-test",
            version="2026-04-24-a",
            tenant="seed",
        )

        next_draft = workbench_module.create_draft_from_workbench("吉林省风生水起商贸有限公司", MINIMAL_TEXT_INPUT, "", [])
        self.assertEqual(next_draft.lines[0].tax_category, "云端审核分类")
        self.assertEqual(next_draft.lines[0].tax_code, "3040802050000000000")
        self.assertIn("命中 客户规则", next_draft.lines[0].coding_reference)


    def test_cloud_profile_cache_supplies_seller_project_profile(self):
        customer_profiles_module.PROFILE_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
        customer_profiles_module.PROFILE_CACHE_PATH.write_text(
            json.dumps(
                [
                    {
                        "seller_name": "沈阳瑞渡酒店管理有限公司",
                        "seller_tax_id": "91210103MAEFNBQW91",
                        "project_profiles": [
                            {
                                "project_name": "管理服务费",
                                "tax_category": "企业管理服务",
                                "tax_code": "3040801990000000000",
                                "tax_rate": "1%",
                                "unit": "项",
                                "line_count": 6,
                            }
                        ],
                        "buyer_profiles": [
                            {"buyer_name": "沈阳测试购买方有限公司", "buyer_tax_id": "91210100TEST000001", "line_count": 3}
                        ],
                    }
                ],
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

        profile = customer_profiles_module.seller_default_line_profile("沈阳瑞渡酒店管理有限公司")
        buyer = customer_profiles_module.resolve_buyer_from_history("测试购买方给开管理费", company_name="沈阳瑞渡酒店管理有限公司")

        self.assertIsNotNone(profile)
        self.assertEqual(profile.project_name, "管理服务费")
        self.assertEqual(profile.tax_code, "3040801990000000000")
        self.assertIsNotNone(buyer)
        self.assertEqual(buyer.buyer.tax_id, "91210100TEST000001")
        summary = customer_profiles_module.profile_cache_summary()
        self.assertEqual(summary["seller_count"], 1)
        self.assertEqual(summary["project_profile_count"], 1)


def _seed_history_profile_row(
    *,
    company_name: str,
    buyer_name: str,
    buyer_tax_id: str,
    project_name: str,
    tax_category: str,
    tax_code: str,
    tax_rate: str,
) -> None:
    ledger_module.LEDGER_CSV_PATH.parent.mkdir(parents=True, exist_ok=True)
    with ledger_module.LEDGER_CSV_PATH.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=ledger_module.LEDGER_HEADERS)
        writer.writeheader()
        writer.writerow(
            {
                "case_id": "history-case",
                "draft_id": "history-draft",
                "line_no": "1",
                "saved_at": "2026-04-30T09:30:00",
                "company_name": company_name,
                "buyer_name": buyer_name,
                "buyer_tax_id": buyer_tax_id,
                "project_name": project_name,
                "tax_category": tax_category,
                "tax_code": tax_code,
                "source_item_code": "",
                "specification": "",
                "unit": "项",
                "quantity": "1",
                "unit_price": "500.00",
                "amount_with_tax": "500.00",
                "tax_rate": tax_rate,
                "coding_reference": "历史已确认记录",
                "coding_state": "manual_correction",
                "note": "",
            }
        )



def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))



def _form_from_draft(draft, *, tax_category: str, tax_code: str, tax_rate: str):
    from werkzeug.datastructures import MultiDict

    form = MultiDict(
        [
            ("company_name", draft.company_name),
            ("raw_text", draft.raw_text),
            ("note", draft.note),
            ("invoice_kind", draft.invoice_kind),
            ("special_business", draft.special_business),
            ("buyer_name", draft.buyer.name),
            ("buyer_tax_id", draft.buyer.tax_id),
            ("buyer_address", draft.buyer.address),
            ("buyer_phone", draft.buyer.phone),
            ("buyer_bank_name", draft.buyer.bank_name),
            ("buyer_bank_account", draft.buyer.bank_account),
        ]
    )
    for line in draft.lines:
        form.add("line_project_name", line.project_name)
        form.add("line_tax_category", tax_category)
        form.add("line_tax_code", tax_code)
        form.add("line_specification", line.specification)
        form.add("line_unit", line.unit)
        form.add("line_quantity", line.quantity)
        form.add("line_unit_price", line.unit_price)
        form.add("line_amount_with_tax", line.resolved_amount_with_tax())
        form.add("line_tax_rate", tax_rate)
        form.add("line_coding_reference", line.coding_reference)
    return form


class _FakeTaxCodeLLMResponse:
    parsed_json = {
        "候选分类": [
            {
                "分类名称": "注射穿刺器械",
                "税收编码": "1090245030000000000",
                "置信度": "0.82",
            }
        ]
    }


class _FakeTaxCodeLLMAdapter:
    is_enabled = True

    def __init__(self):
        self.call_count = 0

    def classify_tax_code(self, item_name, candidates):
        self.call_count += 1
        self.last_item_name = item_name
        self.last_candidates = candidates
        assert any("1090245030000000000" in candidate for candidate in candidates)
        return _FakeTaxCodeLLMResponse()




if __name__ == "__main__":
    unittest.main()
