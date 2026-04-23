import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook
from werkzeug.datastructures import FileStorage, MultiDict

import tax_invoice_demo.ledger as ledger_module
import tax_invoice_demo.workbench as workbench_module
import tax_invoice_batch_demo.lean_workbench as lean_workbench_module


WORKSPACE_ROOT = Path(__file__).resolve().parents[2]
CASE_ROOT = WORKSPACE_ROOT / "invoice-demo" / "案例库原始材料"


class CaseLibraryBatchImportTest(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.temp_path = Path(self.tempdir.name)
        self.old_workbench_root = workbench_module.WORKBENCH_ROOT
        self.old_ledger_paths = (
            ledger_module.LEDGER_ROOT,
            ledger_module.LEDGER_CSV_PATH,
            ledger_module.LEDGER_XLSX_PATH,
            ledger_module.FEEDBACK_CSV_PATH,
        )
        self.old_batch_output_root = lean_workbench_module.BATCH_OUTPUT_ROOT
        self.old_success_paths = (
            lean_workbench_module.SUCCESS_LEDGER_CSV,
            lean_workbench_module.SUCCESS_LEDGER_XLSX,
        )

        workbench_module.WORKBENCH_ROOT = self.temp_path / "workbench"
        ledger_module.LEDGER_ROOT = self.temp_path / "ledger"
        ledger_module.LEDGER_CSV_PATH = ledger_module.LEDGER_ROOT / "累计发票明细表.csv"
        ledger_module.LEDGER_XLSX_PATH = ledger_module.LEDGER_ROOT / "累计发票明细表.xlsx"
        ledger_module.FEEDBACK_CSV_PATH = ledger_module.LEDGER_ROOT / "赋码反馈候选池.csv"
        lean_workbench_module.BATCH_OUTPUT_ROOT = self.temp_path / "batch_import_preview"
        lean_workbench_module.SUCCESS_LEDGER_CSV = lean_workbench_module.BATCH_OUTPUT_ROOT / "批量导入成功明细.csv"
        lean_workbench_module.SUCCESS_LEDGER_XLSX = lean_workbench_module.BATCH_OUTPUT_ROOT / "批量导入成功明细.xlsx"

    def tearDown(self):
        workbench_module.WORKBENCH_ROOT = self.old_workbench_root
        (
            ledger_module.LEDGER_ROOT,
            ledger_module.LEDGER_CSV_PATH,
            ledger_module.LEDGER_XLSX_PATH,
            ledger_module.FEEDBACK_CSV_PATH,
        ) = self.old_ledger_paths
        lean_workbench_module.BATCH_OUTPUT_ROOT = self.old_batch_output_root
        (
            lean_workbench_module.SUCCESS_LEDGER_CSV,
            lean_workbench_module.SUCCESS_LEDGER_XLSX,
        ) = self.old_success_paths
        self.tempdir.cleanup()

    @unittest.skipUnless((CASE_ROOT / "5/source/开票.xlsx").exists(), "本地未发现案例库原始材料")
    def test_case5_excel_docx_generates_enriched_valid_batch_template(self):
        draft = self._draft_from_case_files(
            "吉林省风生水起商贸有限公司",
            "5/source/开票.xlsx",
            "5/source/芃领开票信息.docx",
        )
        export = lean_workbench_module.export_draft_template(draft)

        self.assertEqual(draft.buyer.name, "黑龙江芃领飞象网络科技有限公司")
        self.assertEqual(draft.invoice_kind, "增值税专用发票")
        self.assertEqual(len(draft.lines), 3)
        self.assertEqual(draft.total_amount_with_tax, "8181.16")
        self.assertEqual([line.project_name for line in draft.lines], ["放学乐大菠萝", "大蜜瓜", "葚是喜欢"])
        self.assertEqual([line.tax_category for line in draft.lines], ["冷冻饮品", "冷冻饮品", "冷冻饮品"])
        self.assertEqual(
            [line.tax_code for line in draft.lines],
            ["1030209990000000000", "1030209990000000000", "1030209990000000000"],
        )
        self.assertEqual([line.normalized_tax_rate() for line in draft.lines], ["13%", "13%", "13%"])
        self.assertEqual(export["error_count"], 0)

        detail_rows = _read_detail_rows(export["output_path"])
        self.assertEqual([row["税率"] for row in detail_rows], ["0.13", "0.13", "0.13"])
        self.assertEqual(
            [row["商品和服务税收编码"] for row in detail_rows],
            ["1030209990000000000", "1030209990000000000", "1030209990000000000"],
        )

    @unittest.skipUnless((CASE_ROOT / "5/source/开票.xlsx").exists(), "本地未发现案例库原始材料")
    def test_manual_invoice_kind_and_tax_rate_edits_survive_template_rebuild(self):
        draft = self._draft_from_case_files(
            "吉林省风生水起商贸有限公司",
            "5/source/开票.xlsx",
            "5/source/芃领开票信息.docx",
        )
        form = _form_from_draft(draft, invoice_kind="普通发票", tax_rate="3%")
        saved = lean_workbench_module.save_lean_draft_from_form(draft.draft_id, form, [])
        export = lean_workbench_module.export_draft_template(saved)

        self.assertEqual(saved.invoice_kind, "普通发票")
        self.assertEqual([line.normalized_tax_rate() for line in saved.lines], ["3%", "3%", "3%"])
        self.assertEqual([line.tax_category for line in saved.lines], ["冷冻饮品", "冷冻饮品", "冷冻饮品"])
        self.assertEqual(
            [line.tax_code for line in saved.lines],
            ["1030209990000000000", "1030209990000000000", "1030209990000000000"],
        )
        self.assertEqual(export["error_count"], 0)

        basic_rows = _read_basic_rows(export["output_path"])
        detail_rows = _read_detail_rows(export["output_path"])
        self.assertEqual(basic_rows[0]["发票类型"], "普通发票")
        self.assertEqual([row["税率"] for row in detail_rows], ["0.03", "0.03", "0.03"])

        feedback_rows = _read_csv_rows(ledger_module.FEEDBACK_CSV_PATH)
        self.assertEqual(len(feedback_rows), 3)
        self.assertEqual({row["candidate_status"] for row in feedback_rows}, {"manual_correction"})
        self.assertTrue(all("人工修正赋码" in row["coding_reference"] for row in feedback_rows))
        self.assertEqual([row["tax_rate"] for row in feedback_rows], ["3%", "3%", "3%"])
        self.assertEqual([row["amount_with_tax"] for row in feedback_rows], ["3540.00", "3801.16", "840.00"])

        lean_workbench_module.record_success_to_ledger(saved)
        success_rows = _read_csv_rows(lean_workbench_module.SUCCESS_LEDGER_CSV)
        self.assertEqual(len(success_rows), 3)
        self.assertIn("coding_reference", success_rows[0])
        self.assertTrue(all("人工修正赋码" in row["coding_reference"] for row in success_rows))

    @unittest.skipUnless((CASE_ROOT / "15/source/平安利顺 手机支架 雨伞 风扇.pdf").exists(), "本地未发现案例库原始材料")
    def test_case15_pdf_multi_category_generates_valid_batch_template(self):
        draft = self._draft_from_case_files(
            "吉林省风生水起商贸有限公司",
            "15/source/平安利顺 手机支架 雨伞 风扇.pdf",
        )
        export = lean_workbench_module.export_draft_template(draft)

        self.assertEqual(draft.buyer.name, "前锦网络信息技术（上海）有限公司")
        self.assertEqual(draft.invoice_kind, "普通发票")
        self.assertEqual(len(draft.lines), 3)
        self.assertEqual(draft.total_amount_with_tax, "5500.00")
        self.assertEqual([line.project_name for line in draft.lines], ["手机支架", "雨伞", "风扇"])
        self.assertEqual([line.tax_category for line in draft.lines], ["计算机配套产品", "日用杂品", "家用通风电器具"])
        self.assertEqual(
            [line.tax_code for line in draft.lines],
            ["1090512990000000000", "1060512990000000000", "1090416990000000000"],
        )
        self.assertEqual(export["error_count"], 0)

        detail_rows = _read_detail_rows(export["output_path"])
        self.assertEqual([row["税率"] for row in detail_rows], ["0.01", "0.01", "0.01"])
        self.assertEqual(
            [row["商品和服务税收编码"] for row in detail_rows],
            ["1090512990000000000", "1060512990000000000", "1090416990000000000"],
        )

    def _draft_from_case_files(self, company_name: str, *relatives: str):
        files = []
        handles = []
        try:
            for relative in relatives:
                path = CASE_ROOT / relative
                handle = path.open("rb")
                handles.append(handle)
                files.append(FileStorage(stream=handle, filename=path.name))
            return workbench_module.create_draft_from_workbench(
                company_name=company_name,
                raw_text="",
                note="",
                uploaded_files=files,
            )
        finally:
            for handle in handles:
                handle.close()


def _read_detail_rows(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, data_only=True)
    try:
        sheet = workbook["2-发票明细信息"]
        headers = [str(cell.value or "").strip() for cell in sheet[3]]
        rows = []
        for row_number in range(4, sheet.max_row + 1):
            values = {header: str(sheet.cell(row=row_number, column=index + 1).value or "") for index, header in enumerate(headers) if header}
            if any(values.values()):
                rows.append(values)
        return rows
    finally:
        workbook.close()


def _read_basic_rows(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, data_only=True)
    try:
        sheet = workbook["1-发票基本信息"]
        headers = [str(cell.value or "").strip() for cell in sheet[3]]
        rows = []
        for row_number in range(4, sheet.max_row + 1):
            values = {header: str(sheet.cell(row=row_number, column=index + 1).value or "") for index, header in enumerate(headers) if header}
            if any(values.values()):
                rows.append(values)
        return rows
    finally:
        workbook.close()


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _form_from_draft(draft, *, invoice_kind: str, tax_rate: str) -> MultiDict:
    form = MultiDict(
        [
            ("company_name", draft.company_name),
            ("raw_text", draft.raw_text),
            ("note", draft.note),
            ("invoice_kind", invoice_kind),
            ("special_business", draft.special_business),
            ("buyer_name", draft.buyer.name),
            ("buyer_tax_id", draft.buyer.tax_id),
            ("buyer_address", draft.buyer.address),
            ("buyer_phone", draft.buyer.phone),
            ("buyer_bank_name", draft.buyer.bank_name),
            ("buyer_bank_account", draft.buyer.bank_account),
        ]
    )
    for line in draft.lines:
        form.add("line_project_name", line.project_name)
        form.add("line_tax_category", line.tax_category)
        form.add("line_tax_code", line.tax_code)
        form.add("line_specification", line.specification)
        form.add("line_unit", line.unit)
        form.add("line_quantity", line.quantity)
        form.add("line_unit_price", line.unit_price)
        form.add("line_amount_with_tax", line.resolved_amount_with_tax())
        form.add("line_tax_rate", tax_rate)
        form.add("line_coding_reference", line.coding_reference)
    return form


if __name__ == "__main__":
    unittest.main()
