#!/usr/bin/env python3
"""Incrementally ingest seed-customer materials into local customer profiles.

P0 workflow:
1. Put new assistant/customer files into: 测试组客户档案储备/_收件箱/待处理
2. Run this script.
3. Parse tax-bureau history Excel files into stable profile CSV/JSON under _档案库.
4. Move processed inbox files into 已处理 / 重复文件 / 解析失败.

The active profile layer is rebuilt only from trusted tax-bureau history Excel files
stored in customer folders, not directly from raw OCR/image candidates.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import shutil
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - deployment diagnostic
    load_workbook = None


PROJECT_ROOT = Path(__file__).resolve().parents[1]


def default_profile_root() -> Path:
    """Return the runtime customer-profile workspace.

    Development historically keeps `测试组客户档案储备` next to the code project.
    Windows full-install packages should keep it inside the installed product
    folder, so the inbox/profile workflow is covered by the package and does not
    accidentally write to `C:\` when the zip is flattened into C:\InvoiceAssistant.
    """

    env_root = os.environ.get("TAX_INVOICE_PROFILE_ROOT", "").strip()
    if env_root:
        return Path(env_root).expanduser()
    bundled_root = PROJECT_ROOT / "测试组客户档案储备"
    legacy_sibling_root = PROJECT_ROOT.parent / "测试组客户档案储备"
    legacy_active = legacy_sibling_root / "_档案库" / "customer_profiles_active.json"
    # 本地开发环境已有完整历史档案源，优先继续使用；安装包环境则使用包内工作区。
    if legacy_active.exists():
        return legacy_sibling_root
    if bundled_root.exists():
        return bundled_root
    if legacy_sibling_root.exists():
        return legacy_sibling_root
    return bundled_root


DEFAULT_PROFILE_ROOT = default_profile_root()
HISTORY_SHEET_NAME = "信息汇总表"
BASIC_SHEET_NAME = "发票基础信息"
REQUIRED_HISTORY_HEADERS = {
    "销方识别号",
    "销方名称",
    "购方识别号",
    "购买方名称",
    "税收分类编码",
    "货物或应税劳务名称",
    "税率",
    "价税合计",
}
INBOX_DIR = "_收件箱"
PENDING_DIR = "待处理"
PROCESSED_DIR = "已处理"
FAILED_DIR = "解析失败"
DUPLICATE_DIR = "重复文件"
PROFILE_DB_DIR = "_档案库"
REVIEW_DIR = "_待确认"
EXCLUDED_REBUILD_DIRS = {INBOX_DIR, PROFILE_DB_DIR, REVIEW_DIR, "客户档案_整理_20260430"}
IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".webp", ".bmp"}
EXCEL_EXTS = {".xlsx"}


@dataclass
class HistoryRow:
    source_file: str
    serial: str
    seller_tax_id: str
    seller_name: str
    buyer_tax_id: str
    buyer_name: str
    invoice_date: str
    tax_code: str
    full_item_name: str
    tax_category: str
    project_name: str
    specification: str
    unit: str
    quantity: str
    unit_price: str
    amount: str
    tax_rate: str
    tax_amount: str
    amount_with_tax: str
    invoice_source: str
    invoice_kind: str
    invoice_status: str
    is_positive: str
    risk_level: str
    note: str = ""


def profile_root_from_args(value: str | None) -> Path:
    return Path(value).expanduser().resolve() if value else DEFAULT_PROFILE_ROOT.resolve()


def ensure_dirs(root: Path) -> None:
    for rel in [
        f"{INBOX_DIR}/{PENDING_DIR}",
        f"{INBOX_DIR}/{PROCESSED_DIR}",
        f"{INBOX_DIR}/{FAILED_DIR}",
        f"{INBOX_DIR}/{DUPLICATE_DIR}",
        PROFILE_DB_DIR,
        REVIEW_DIR,
    ]:
        (root / rel).mkdir(parents=True, exist_ok=True)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_manifest(root: Path) -> dict[str, dict[str, str]]:
    path = root / PROFILE_DB_DIR / "import_manifest.csv"
    if not path.exists():
        return {}
    with path.open("r", newline="", encoding="utf-8-sig") as f:
        return {row["sha256"]: row for row in csv.DictReader(f) if row.get("sha256")}


def write_manifest(root: Path, manifest: dict[str, dict[str, str]]) -> None:
    path = root / PROFILE_DB_DIR / "import_manifest.csv"
    fields = ["sha256", "original_name", "file_type", "status", "seller_name", "canonical_path", "processed_path", "message", "ingested_at"]
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in sorted(manifest.values(), key=lambda item: item.get("ingested_at", "")):
            writer.writerow({key: row.get(key, "") for key in fields})


def safe_filename(value: str) -> str:
    cleaned = re.sub(r"[\\/:*?\"<>|\s]+", "_", value).strip("_")
    return cleaned[:100] or "未命名"


def unique_destination(path: Path) -> Path:
    if not path.exists():
        return path
    stem, suffix = path.stem, path.suffix
    for index in range(2, 1000):
        candidate = path.with_name(f"{stem}_{index}{suffix}")
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"无法生成不重名路径: {path}")


def move_file(source: Path, target: Path) -> Path:
    target.parent.mkdir(parents=True, exist_ok=True)
    target = unique_destination(target)
    shutil.move(str(source), str(target))
    return target


def copy_file(source: Path, target: Path) -> Path:
    target.parent.mkdir(parents=True, exist_ok=True)
    target = unique_destination(target)
    shutil.copy2(source, target)
    return target


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_date(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value).strip()


def split_full_item_name(value: str) -> tuple[str, str]:
    text = value.strip()
    matched = re.match(r"^\*([^*]+)\*(.+)$", text)
    if matched:
        return matched.group(1).strip(), matched.group(2).strip()
    return "", text.strip("*")


def decimal_text(value: str) -> str:
    text = str(value or "").replace(",", "").strip()
    if not text:
        return "0.00"
    try:
        return f"{Decimal(text):.2f}"
    except InvalidOperation:
        return text


def decimal_value(value: str) -> Decimal:
    try:
        return Decimal(str(value or "0").replace(",", ""))
    except InvalidOperation:
        return Decimal("0")


def parse_history_excel(path: Path, *, source_root: Path) -> list[HistoryRow]:
    if load_workbook is None:
        raise RuntimeError("openpyxl 不可用，无法解析 Excel 历史明细。")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if HISTORY_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"缺少 Sheet：{HISTORY_SHEET_NAME}")
        sheet = workbook[HISTORY_SHEET_NAME]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError("信息汇总表为空")
        headers = [cell_text(value) for value in rows[0]]
        missing = REQUIRED_HISTORY_HEADERS - set(headers)
        if missing:
            raise ValueError("不是税局历史明细 Excel，缺少表头：" + "、".join(sorted(missing)))
        idx = {name: headers.index(name) for name in headers if name}

        def get(row: tuple[Any, ...], header: str) -> str:
            index = idx.get(header)
            if index is None or index >= len(row):
                return ""
            return cell_text(row[index])

        parsed: list[HistoryRow] = []
        for raw in rows[1:]:
            if not raw or not any(raw):
                continue
            seller_name = get(raw, "销方名称")
            seller_tax_id = get(raw, "销方识别号")
            if not seller_name or not seller_tax_id:
                continue
            full_item = get(raw, "货物或应税劳务名称")
            tax_category, project_name = split_full_item_name(full_item)
            parsed.append(
                HistoryRow(
                    source_file=str(path.relative_to(source_root)) if path.is_relative_to(source_root) else str(path),
                    serial=get(raw, "数电发票号码"),
                    seller_tax_id=seller_tax_id,
                    seller_name=seller_name,
                    buyer_tax_id=get(raw, "购方识别号"),
                    buyer_name=get(raw, "购买方名称"),
                    invoice_date=normalize_date(raw[idx["开票日期"]]) if "开票日期" in idx and idx["开票日期"] < len(raw) else "",
                    tax_code=get(raw, "税收分类编码"),
                    full_item_name=full_item,
                    tax_category=tax_category,
                    project_name=project_name,
                    specification=get(raw, "规格型号"),
                    unit=get(raw, "单位"),
                    quantity=get(raw, "数量"),
                    unit_price=get(raw, "单价"),
                    amount=get(raw, "金额"),
                    tax_rate=get(raw, "税率"),
                    tax_amount=get(raw, "税额"),
                    amount_with_tax=get(raw, "价税合计"),
                    invoice_source=get(raw, "发票来源"),
                    invoice_kind=get(raw, "发票票种"),
                    invoice_status=get(raw, "发票状态"),
                    is_positive=get(raw, "是否正数发票"),
                    risk_level=get(raw, "发票风险等级"),
                )
            )
        if not parsed:
            raise ValueError("未解析到有效历史明细行")
        return parsed
    finally:
        workbook.close()


def is_trusted_history_row(row: HistoryRow) -> bool:
    status = row.invoice_status.strip()
    positive = row.is_positive.strip()
    if status and status != "正常":
        return False
    if positive and positive != "是":
        return False
    if row.amount_with_tax.strip().startswith("-") or row.amount.strip().startswith("-"):
        return False
    text = " ".join([row.full_item_name, row.invoice_status, row.risk_level])
    if any(token in text for token in ["红字", "红冲", "作废"]):
        return False
    return True


def ingest_pending_files(root: Path, *, dry_run: bool = False) -> dict[str, int]:
    """Process inbox entries.

    Important convention: one subfolder under _收件箱/待处理 is treated as one
    customer material bundle. If the bundle contains a tax-bureau history Excel,
    that Excel identifies the seller, and all screenshots/notes in the same
    bundle are archived under that seller's 客户沟通材料 directory.
    """

    pending = root / INBOX_DIR / PENDING_DIR
    manifest = load_manifest(root)
    counters = Counter()
    today = datetime.now().strftime("%Y%m%d")
    review_rows: list[dict[str, str]] = []

    entries = sorted([entry for entry in pending.iterdir() if entry.name != ".DS_Store"], key=lambda item: item.name)
    for entry in entries:
        if entry.is_dir():
            _ingest_pending_bundle(root, entry, manifest, counters, review_rows, today=today, dry_run=dry_run)
        elif entry.is_file():
            _ingest_pending_file(
                root,
                entry,
                manifest,
                counters,
                review_rows,
                today=today,
                dry_run=dry_run,
                bundle_root=None,
                bundle_seller_name="",
                bundle_name="",
            )

    if not dry_run:
        write_manifest(root, manifest)
        if review_rows:
            append_review_rows(root / REVIEW_DIR / "图片材料待确认.csv", review_rows)
    return dict(counters)



def _ingest_pending_bundle(
    root: Path,
    bundle_dir: Path,
    manifest: dict[str, dict[str, str]],
    counters: Counter,
    review_rows: list[dict[str, str]],
    *,
    today: str,
    dry_run: bool,
) -> None:
    files = sorted([item for item in bundle_dir.rglob("*") if item.is_file() and item.name != ".DS_Store"])
    if not files:
        return
    counters["bundles_seen"] += 1

    seller_candidates: Counter[str] = Counter()
    for file_path in files:
        if file_path.suffix.lower() not in EXCEL_EXTS:
            continue
        file_hash = sha256_file(file_path)
        if file_hash in manifest and manifest[file_hash].get("seller_name"):
            seller_candidates[manifest[file_hash]["seller_name"]] += 1
            continue
        try:
            rows = parse_history_excel(file_path, source_root=root)
            seller_candidates[rows[0].seller_name] += 1
        except Exception:
            continue

    bundle_seller_name = seller_candidates.most_common(1)[0][0] if seller_candidates else ""
    if bundle_seller_name:
        counters["bundles_with_seller"] += 1
    else:
        counters["bundles_without_seller"] += 1
        _archive_unassigned_bundle(root, bundle_dir, files, manifest, counters, review_rows, today=today, dry_run=dry_run)
        if not dry_run:
            shutil.rmtree(bundle_dir, ignore_errors=True)
        return

    for file_path in files:
        _ingest_pending_file(
            root,
            file_path,
            manifest,
            counters,
            review_rows,
            today=today,
            dry_run=dry_run,
            bundle_root=bundle_dir,
            bundle_seller_name=bundle_seller_name,
            bundle_name=bundle_dir.name,
        )

    if not dry_run:
        try:
            bundle_dir.rmdir()
        except OSError:
            # Nested empty directories may remain; remove best-effort.
            shutil.rmtree(bundle_dir, ignore_errors=True)



def _archive_unassigned_bundle(
    root: Path,
    bundle_dir: Path,
    files: list[Path],
    manifest: dict[str, dict[str, str]],
    counters: Counter,
    review_rows: list[dict[str, str]],
    *,
    today: str,
    dry_run: bool,
) -> None:
    """Preserve a bundle without a tax-history seller as one review unit.

    Customer/assistant materials may include invoice samples, PDFs, Word docs,
    screenshots, or current detail workbooks without any tax-bureau history
    export. Those files are not active profile evidence, but they are not
    failures either. Keep them together for manual seller assignment.
    """

    bundle_name = bundle_dir.name
    review_bundle_dir = root / REVIEW_DIR / "未关联客户资料包" / f"{today}_{safe_filename(bundle_name)}"
    processed_base = root / INBOX_DIR / PROCESSED_DIR / today / safe_filename(bundle_name)
    for path in files:
        counters["unassigned_bundle_material"] += 1
        file_hash = sha256_file(path)
        rel_in_bundle = path.relative_to(bundle_dir)
        if file_hash in manifest:
            counters["duplicate"] += 1
            continue
        canonical = review_bundle_dir / rel_in_bundle
        if dry_run:
            processed = processed_base / rel_in_bundle
        else:
            canonical = copy_file(path, canonical)
            processed = move_file(path, processed_base / rel_in_bundle)
        suffix = path.suffix.lower()
        file_type = "image_candidate" if suffix in IMAGE_EXTS else "unassigned_bundle_material"
        manifest[file_hash] = {
            "sha256": file_hash,
            "original_name": path.name,
            "file_type": file_type,
            "status": "pending_seller_review",
            "seller_name": "",
            "canonical_path": str(canonical.relative_to(root)),
            "processed_path": str(processed.relative_to(root)),
            "message": f"无税局历史 Excel，按完整资料包待确认销售主体；资料包：{bundle_name}",
            "ingested_at": datetime.now().isoformat(timespec="seconds"),
        }
        review_rows.append(
            {
                "sha256": file_hash,
                "original_name": path.name,
                "seller_name": "",
                "relative_path": str(canonical.relative_to(root)),
                "candidate_type": file_type,
                "status": "pending_seller_review",
                "note": f"完整资料包待确认销售主体：{bundle_name}。不要把样票/PDF 文件名中的买方误判为销售主体。",
            }
        )



def _ingest_pending_file(
    root: Path,
    path: Path,
    manifest: dict[str, dict[str, str]],
    counters: Counter,
    review_rows: list[dict[str, str]],
    *,
    today: str,
    dry_run: bool,
    bundle_root: Path | None,
    bundle_seller_name: str,
    bundle_name: str,
) -> None:
    counters["seen"] += 1
    file_hash = sha256_file(path)
    rel_in_bundle = path.relative_to(bundle_root) if bundle_root else Path(path.name)
    processed_base = root / INBOX_DIR / PROCESSED_DIR / today / (safe_filename(bundle_name) if bundle_name else "")
    duplicate_base = root / INBOX_DIR / DUPLICATE_DIR / today / (safe_filename(bundle_name) if bundle_name else "")
    failed_base = root / INBOX_DIR / FAILED_DIR / today / (safe_filename(bundle_name) if bundle_name else "")

    if file_hash in manifest:
        counters["duplicate"] += 1
        if not dry_run:
            moved = move_file(path, duplicate_base / rel_in_bundle)
            manifest[file_hash]["processed_path"] = str(moved.relative_to(root))
        return

    suffix = path.suffix.lower()
    try:
        if suffix in EXCEL_EXTS:
            try:
                rows = parse_history_excel(path, source_root=root)
            except Exception:
                if bundle_seller_name:
                    # In a seller-identified bundle, an Excel that is not a tax-bureau
                    # history export is usually a current invoice sample/detail workbook.
                    # Archive it with the customer materials instead of treating it as
                    # a failed history import. It must not enter active profiles.
                    material_dir = root / bundle_seller_name / "客户沟通材料" / f"{today}_{safe_filename(bundle_name)}"
                    canonical = material_dir / rel_in_bundle
                    if dry_run:
                        processed = processed_base / rel_in_bundle
                    else:
                        canonical = copy_file(path, canonical)
                        processed = move_file(path, processed_base / rel_in_bundle)
                    manifest[file_hash] = {
                        "sha256": file_hash,
                        "original_name": path.name,
                        "file_type": "bundle_material_candidate",
                        "status": "candidate_only",
                        "seller_name": bundle_seller_name,
                        "canonical_path": str(canonical.relative_to(root)),
                        "processed_path": str(processed.relative_to(root)),
                        "message": f"同资料包归档到客户沟通材料；非税局历史 Excel，不进入 active 档案；资料包：{bundle_name}",
                        "ingested_at": datetime.now().isoformat(timespec="seconds"),
                    }
                    counters["bundle_material_archived"] += 1
                    review_rows.append(
                        {
                            "sha256": file_hash,
                            "original_name": path.name,
                            "seller_name": bundle_seller_name,
                            "relative_path": str(canonical.relative_to(root)),
                            "candidate_type": "bundle_material_candidate",
                            "status": "pending_review",
                            "note": f"来自同一资料包：{bundle_name}；非税局历史 Excel，按本次开票材料候选归档，不进入 active 档案。",
                        }
                    )
                    return
                raise
            seller_name = rows[0].seller_name
            target = root / seller_name / "历史开票明细" / path.name
            if dry_run:
                canonical = target
                processed = processed_base / rel_in_bundle
            else:
                canonical = copy_file(path, target)
                processed = move_file(path, processed_base / rel_in_bundle)
            manifest[file_hash] = {
                "sha256": file_hash,
                "original_name": path.name,
                "file_type": "tax_bureau_history_excel",
                "status": "imported",
                "seller_name": seller_name,
                "canonical_path": str(canonical.relative_to(root)),
                "processed_path": str(processed.relative_to(root)),
                "message": f"解析 {len(rows)} 行历史明细" + (f"；所属资料包：{bundle_name}" if bundle_name else ""),
                "ingested_at": datetime.now().isoformat(timespec="seconds"),
            }
            counters["history_excel_imported"] += 1
            return

        if bundle_seller_name:
            material_dir = root / bundle_seller_name / "客户沟通材料" / f"{today}_{safe_filename(bundle_name)}"
            canonical = material_dir / rel_in_bundle
            if dry_run:
                processed = processed_base / rel_in_bundle
            else:
                canonical = copy_file(path, canonical)
                processed = move_file(path, processed_base / rel_in_bundle)
            file_type = "image_candidate" if suffix in IMAGE_EXTS else "bundle_material_candidate"
            manifest[file_hash] = {
                "sha256": file_hash,
                "original_name": path.name,
                "file_type": file_type,
                "status": "candidate_only",
                "seller_name": bundle_seller_name,
                "canonical_path": str(canonical.relative_to(root)),
                "processed_path": str(processed.relative_to(root)),
                "message": f"同资料包归档到客户沟通材料；不进入 active 档案；资料包：{bundle_name}",
                "ingested_at": datetime.now().isoformat(timespec="seconds"),
            }
            counters["bundle_material_archived"] += 1
            review_rows.append(
                {
                    "sha256": file_hash,
                    "original_name": path.name,
                    "seller_name": bundle_seller_name,
                    "relative_path": str(canonical.relative_to(root)),
                    "candidate_type": file_type,
                    "status": "pending_review",
                    "note": f"来自同一资料包：{bundle_name}；默认不进入 active 档案。",
                }
            )
            return

        if suffix in IMAGE_EXTS:
            counters["image_candidate"] += 1
            review_rows.append(
                {
                    "sha256": file_hash,
                    "original_name": path.name,
                    "seller_name": "",
                    "relative_path": str(path.relative_to(root)),
                    "candidate_type": "image_material",
                    "status": "pending_review",
                    "note": "图片候选材料：待 OCR/人工归类；默认不进入 active 档案。若需关联客户，请和该客户历史明细放入同一个资料包文件夹。",
                }
            )
            if dry_run:
                processed = processed_base / "图片材料" / rel_in_bundle
            else:
                processed = move_file(path, processed_base / "图片材料" / rel_in_bundle)
            manifest[file_hash] = {
                "sha256": file_hash,
                "original_name": path.name,
                "file_type": "image_candidate",
                "status": "candidate_only",
                "seller_name": "",
                "canonical_path": "",
                "processed_path": str(processed.relative_to(root)),
                "message": "图片候选材料，未关联销售主体，未进入 active 档案",
                "ingested_at": datetime.now().isoformat(timespec="seconds"),
            }
            return

        counters["unsupported"] += 1
        if dry_run:
            failed = failed_base / rel_in_bundle
        else:
            failed = move_file(path, failed_base / rel_in_bundle)
        manifest[file_hash] = {
            "sha256": file_hash,
            "original_name": path.name,
            "file_type": suffix or "unknown",
            "status": "unsupported",
            "seller_name": "",
            "canonical_path": "",
            "processed_path": str(failed.relative_to(root)),
            "message": "暂不支持的文件类型；如需关联客户，请和该客户历史明细放入同一个资料包文件夹。",
            "ingested_at": datetime.now().isoformat(timespec="seconds"),
        }
    except Exception as exc:  # noqa: BLE001
        counters["failed"] += 1
        if dry_run:
            failed = failed_base / rel_in_bundle
        else:
            failed = move_file(path, failed_base / rel_in_bundle)
        manifest[file_hash] = {
            "sha256": file_hash,
            "original_name": path.name,
            "file_type": suffix or "unknown",
            "status": "failed",
            "seller_name": bundle_seller_name,
            "canonical_path": "",
            "processed_path": str(failed.relative_to(root)),
            "message": f"解析失败：{type(exc).__name__}: {exc}",
            "ingested_at": datetime.now().isoformat(timespec="seconds"),
        }



def append_review_rows(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = ["sha256", "original_name", "seller_name", "relative_path", "candidate_type", "status", "note"]
    existing_hashes = set()
    if path.exists():
        with path.open("r", newline="", encoding="utf-8-sig") as f:
            existing_hashes = {row.get("sha256", "") for row in csv.DictReader(f)}
    new_rows = [row for row in rows if row["sha256"] not in existing_hashes]
    if not new_rows:
        return
    write_header = not path.exists()
    with path.open("a", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        if write_header:
            writer.writeheader()
        writer.writerows(new_rows)


def collect_history_rows(root: Path) -> list[HistoryRow]:
    rows: list[HistoryRow] = []
    for path in sorted(root.rglob("*.xlsx")):
        rel_parts = path.relative_to(root).parts
        if rel_parts and rel_parts[0] in EXCLUDED_REBUILD_DIRS:
            continue
        try:
            rows.extend(parse_history_excel(path, source_root=root))
        except Exception:
            continue
    return [row for row in rows if is_trusted_history_row(row)]


def normalize_dedupe_text(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip().upper()


def normalize_dedupe_date(value: str) -> str:
    text = str(value or "").strip()
    return text[:10]


def history_row_dedupe_key(row: HistoryRow) -> tuple[str, ...]:
    """Stable line-level key for repeated exports of the same tax history.

    Re-exported customer history files often contain old invoice rows plus a few
    new rows. File name or file hash is not enough: a changed Excel can contain
    mostly duplicate lines. The active profile layer therefore deduplicates by
    invoice-line business identity, not by source file.
    """

    return (
        normalize_dedupe_text(row.seller_tax_id),
        normalize_dedupe_text(row.seller_name),
        normalize_dedupe_text(row.buyer_tax_id),
        normalize_dedupe_text(row.buyer_name),
        normalize_dedupe_text(row.serial),
        normalize_dedupe_date(row.invoice_date),
        normalize_dedupe_text(row.full_item_name or row.project_name),
        normalize_dedupe_text(row.tax_code),
        normalize_dedupe_text(row.tax_rate),
        decimal_text(row.amount),
        decimal_text(row.tax_amount),
        decimal_text(row.amount_with_tax),
    )


def dedupe_history_rows(rows: list[HistoryRow]) -> tuple[list[HistoryRow], int]:
    seen: set[tuple[str, ...]] = set()
    unique_rows: list[HistoryRow] = []
    duplicate_count = 0
    for row in rows:
        key = history_row_dedupe_key(row)
        if key in seen:
            duplicate_count += 1
            continue
        seen.add(key)
        unique_rows.append(row)
    return unique_rows, duplicate_count


def rebuild_profiles(root: Path) -> dict[str, int]:
    profile_dir = root / PROFILE_DB_DIR
    profile_dir.mkdir(parents=True, exist_ok=True)
    trusted_rows = collect_history_rows(root)
    rows, duplicate_history_rows = dedupe_history_rows(trusted_rows)
    seller_groups: dict[tuple[str, str], list[HistoryRow]] = defaultdict(list)
    for row in rows:
        seller_groups[(row.seller_name, row.seller_tax_id)].append(row)

    write_seller_profiles(profile_dir / "seller_profiles.csv", seller_groups)
    write_project_profiles(profile_dir / "seller_project_profiles.csv", seller_groups)
    write_buyer_profiles(profile_dir / "buyer_profiles.csv", seller_groups)
    write_material_inventory(root, profile_dir / "material_inventory.csv")
    write_active_json(profile_dir / "customer_profiles_active.json", seller_groups)
    write_overview_md(profile_dir / "客户档案总览.md", seller_groups)
    write_product_cache(PROJECT_ROOT / "output" / "workbench" / "tax_invoice_demo" / "客户档案缓存.json", seller_groups)
    return {
        "trusted_history_rows_raw": len(trusted_rows),
        "duplicate_history_rows": duplicate_history_rows,
        "trusted_history_rows": len(rows),
        "seller_count": len(seller_groups),
        "buyer_count": sum(len({(row.buyer_name, row.buyer_tax_id) for row in group}) for group in seller_groups.values()),
    }


def write_seller_profiles(path: Path, groups: dict[tuple[str, str], list[HistoryRow]]) -> None:
    fields = [
        "seller_name",
        "seller_tax_id",
        "history_rows",
        "invoice_count",
        "buyer_count",
        "date_min",
        "date_max",
        "total_amount_with_tax",
        "primary_project",
        "primary_tax_category",
        "primary_tax_code",
        "primary_tax_rate",
        "source_confidence",
    ]
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for (seller_name, seller_tax_id), rows in sorted(groups.items()):
            project_counter = Counter((r.project_name, r.tax_category, r.tax_code, r.tax_rate) for r in rows)
            top = project_counter.most_common(1)[0][0] if project_counter else ("", "", "", "")
            writer.writerow(
                {
                    "seller_name": seller_name,
                    "seller_tax_id": seller_tax_id,
                    "history_rows": len(rows),
                    "invoice_count": len({r.serial for r in rows if r.serial}),
                    "buyer_count": len({(r.buyer_name, r.buyer_tax_id) for r in rows}),
                    "date_min": min([r.invoice_date for r in rows if r.invoice_date], default=""),
                    "date_max": max([r.invoice_date for r in rows if r.invoice_date], default=""),
                    "total_amount_with_tax": f"{sum(decimal_value(r.amount_with_tax) for r in rows):.2f}",
                    "primary_project": top[0],
                    "primary_tax_category": top[1],
                    "primary_tax_code": top[2],
                    "primary_tax_rate": top[3],
                    "source_confidence": "official_history_export",
                }
            )


def write_project_profiles(path: Path, groups: dict[tuple[str, str], list[HistoryRow]]) -> None:
    fields = [
        "seller_name",
        "seller_tax_id",
        "full_item_name",
        "tax_category",
        "project_name",
        "tax_code",
        "tax_rate",
        "unit",
        "line_count",
        "total_amount_with_tax",
        "confidence",
        "profile_note",
        "source_confidence",
    ]
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for (seller_name, seller_tax_id), rows in sorted(groups.items()):
            grouped: dict[tuple[str, str, str, str, str, str], list[HistoryRow]] = defaultdict(list)
            for row in rows:
                grouped[(row.full_item_name, row.tax_category, row.project_name, row.tax_code, row.tax_rate, row.unit or "项")].append(row)
            for key, items in sorted(grouped.items(), key=lambda item: (-len(item[1]), item[0])):
                writer.writerow(
                    {
                        "seller_name": seller_name,
                        "seller_tax_id": seller_tax_id,
                        "full_item_name": key[0],
                        "tax_category": key[1],
                        "project_name": key[2],
                        "tax_code": key[3],
                        "tax_rate": key[4],
                        "unit": key[5],
                        "line_count": len(items),
                        "total_amount_with_tax": f"{sum(decimal_value(r.amount_with_tax) for r in items):.2f}",
                        "confidence": "high" if len(items) >= 2 else "medium",
                        "profile_note": "税局历史明细提取，作为推荐需人工复核",
                        "source_confidence": "official_history_export",
                    }
                )


def write_buyer_profiles(path: Path, groups: dict[tuple[str, str], list[HistoryRow]]) -> None:
    fields = [
        "seller_name",
        "seller_tax_id",
        "buyer_name",
        "buyer_tax_id",
        "line_count",
        "invoice_count_hint",
        "total_amount_with_tax",
        "first_date",
        "last_date",
        "common_projects",
        "source_confidence",
    ]
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for (seller_name, seller_tax_id), rows in sorted(groups.items()):
            grouped: dict[tuple[str, str], list[HistoryRow]] = defaultdict(list)
            for row in rows:
                grouped[(row.buyer_name, row.buyer_tax_id)].append(row)
            for (buyer_name, buyer_tax_id), items in sorted(grouped.items(), key=lambda item: (-len(item[1]), item[0][0])):
                projects = Counter(r.project_name for r in items if r.project_name)
                writer.writerow(
                    {
                        "seller_name": seller_name,
                        "seller_tax_id": seller_tax_id,
                        "buyer_name": buyer_name,
                        "buyer_tax_id": buyer_tax_id,
                        "line_count": len(items),
                        "invoice_count_hint": len({r.serial for r in items if r.serial}),
                        "total_amount_with_tax": f"{sum(decimal_value(r.amount_with_tax) for r in items):.2f}",
                        "first_date": min([r.invoice_date for r in items if r.invoice_date], default=""),
                        "last_date": max([r.invoice_date for r in items if r.invoice_date], default=""),
                        "common_projects": "；".join(f"{name}:{count}" for name, count in projects.most_common()),
                        "source_confidence": "official_history_export",
                    }
                )


def write_material_inventory(root: Path, path: Path) -> None:
    fields = ["relative_path", "type", "size_bytes", "sha256", "note"]
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for item in sorted(root.rglob("*")):
            if not item.is_file() or item.name == ".DS_Store":
                continue
            rel_parts = item.relative_to(root).parts
            if rel_parts and rel_parts[0] == PROFILE_DB_DIR:
                continue
            suffix = item.suffix.lower()
            if suffix in EXCEL_EXTS:
                file_type = "excel_history_or_candidate"
            elif suffix in IMAGE_EXTS:
                file_type = "image_material"
            else:
                file_type = "other"
            note = ""
            if INBOX_DIR in rel_parts:
                note = "收件箱归档文件"
            elif "历史开票明细" in rel_parts:
                note = "客户历史开票明细"
            elif suffix in IMAGE_EXTS:
                note = "图片/截图材料，默认需人工复核"
            writer.writerow(
                {
                    "relative_path": str(item.relative_to(root)),
                    "type": file_type,
                    "size_bytes": item.stat().st_size,
                    "sha256": sha256_file(item),
                    "note": note,
                }
            )


def write_active_json(path: Path, groups: dict[tuple[str, str], list[HistoryRow]]) -> None:
    payload = []
    for (seller_name, seller_tax_id), rows in sorted(groups.items()):
        project_groups: dict[tuple[str, str, str, str, str, str], list[HistoryRow]] = defaultdict(list)
        buyer_groups: dict[tuple[str, str], list[HistoryRow]] = defaultdict(list)
        for row in rows:
            project_groups[(row.full_item_name, row.tax_category, row.project_name, row.tax_code, row.tax_rate, row.unit or "项")].append(row)
            buyer_groups[(row.buyer_name, row.buyer_tax_id)].append(row)
        payload.append(
            {
                "seller_name": seller_name,
                "seller_tax_id": seller_tax_id,
                "source_confidence": "official_history_export",
                "updated_at": datetime.now().isoformat(timespec="seconds"),
                "project_profiles": [
                    {
                        "full_item_name": key[0],
                        "tax_category": key[1],
                        "project_name": key[2],
                        "tax_code": key[3],
                        "tax_rate": key[4],
                        "unit": key[5],
                        "line_count": len(items),
                        "confidence": "high" if len(items) >= 2 else "medium",
                    }
                    for key, items in sorted(project_groups.items(), key=lambda item: (-len(item[1]), item[0]))
                ],
                "buyer_profiles": [
                    {"buyer_name": key[0], "buyer_tax_id": key[1], "line_count": len(items)}
                    for key, items in sorted(buyer_groups.items(), key=lambda item: (-len(item[1]), item[0][0]))
                ],
            }
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_product_cache(path: Path, groups: dict[tuple[str, str], list[HistoryRow]]) -> None:
    # Same shape as active profile JSON, copied to workbench output for product integration.
    write_active_json(path, groups)


def write_overview_md(path: Path, groups: dict[tuple[str, str], list[HistoryRow]]) -> None:
    lines = [
        "# 客户档案总览",
        "",
        f"更新时间：{datetime.now().isoformat(timespec='seconds')}",
        "",
        "| 销售主体 | 税号 | 历史行/发票 | 购买方 | 主常用项目 | 税码 | 税率 |",
        "|---|---:|---:|---:|---|---:|---:|",
    ]
    for (seller_name, seller_tax_id), rows in sorted(groups.items()):
        projects = Counter((r.project_name, r.tax_category, r.tax_code, r.tax_rate) for r in rows)
        top = projects.most_common(1)[0][0] if projects else ("", "", "", "")
        lines.append(
            f"| {seller_name} | `{seller_tax_id}` | {len(rows)}/{len({r.serial for r in rows if r.serial})} | {len({(r.buyer_name, r.buyer_tax_id) for r in rows})} | {top[0]} / {top[1]} | `{top[2]}` | {top[3]} |"
        )
    lines += [
        "",
        "## 使用边界",
        "",
        "- active 档案仅来自正常、正数、未作废/未红冲的税局历史明细。",
        "- 图片/OCR/LLM 结果默认进入候选或待确认，不直接污染 active 档案。",
        "- 生成草稿时仍需人工复核项目、税码、税率和购买方。",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_run_report(
    root: Path,
    ingest_counts: dict[str, int],
    rebuild_counts: dict[str, int],
    *,
    dry_run: bool,
    cloud_sync: dict[str, str | int] | None = None,
) -> Path:
    report_dir = root / PROFILE_DB_DIR / "处理报告"
    report_dir.mkdir(parents=True, exist_ok=True)
    path = report_dir / f"profile_ingest_{datetime.now().strftime('%Y%m%d_%H%M%S')}.md"
    lines = [
        "# 客户档案收件箱处理报告",
        "",
        f"时间：{datetime.now().isoformat(timespec='seconds')}",
        f"模式：{'dry-run' if dry_run else 'execute'}",
        "",
        "## 本次收件箱处理",
        "",
    ]
    if ingest_counts:
        for key, value in sorted(ingest_counts.items()):
            lines.append(f"- {key}: {value}")
    else:
        lines.append("- 无新增待处理文件")
    lines += ["", "## active 档案重建", ""]
    for key, value in sorted(rebuild_counts.items()):
        lines.append(f"- {key}: {value}")
    if cloud_sync is not None:
        lines += ["", "## 云端同步", ""]
        for key, value in sorted(cloud_sync.items()):
            lines.append(f"- {key}: {value}")
    lines += [
        "",
        "## 下一步",
        "",
        "- 新文件继续放入 `_收件箱/待处理`。",
        "- 需要人工确认的图片/OCR 候选查看 `_待确认`。",
        "- 产品侧读取 `_档案库/customer_profiles_active.json` 或工作台输出缓存 `output/workbench/tax_invoice_demo/客户档案缓存.json`。",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def sync_profiles_to_cloud() -> dict[str, str | int]:
    sys.path.insert(0, str(PROJECT_ROOT))
    from tax_invoice_demo.sync_service import sync_customer_profiles

    result = sync_customer_profiles(PROJECT_ROOT / "output" / "workbench" / "tax_invoice_demo" / "客户档案缓存.json")
    return {
        "status": result.status,
        "seller_count": result.seller_count,
        "buyer_count": result.buyer_count,
        "line_profile_count": result.line_profile_count,
        "batch_id": result.batch_id,
        "endpoint": result.endpoint,
        "error": result.error,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="增量处理测试组客户档案收件箱")
    parser.add_argument("--root", default="", help="客户档案储备目录；默认使用项目同级的 测试组客户档案储备")
    parser.add_argument("--init-only", action="store_true", help="只创建目录结构，不处理文件")
    parser.add_argument("--rebuild-only", action="store_true", help="只重建 _档案库，不处理收件箱")
    parser.add_argument("--dry-run", action="store_true", help="只预演，不移动/写入收件箱处理状态")
    parser.add_argument("--sync-cloud", action="store_true", help="处理完成后同步结构化 active 客户档案到阿里云 sync center")
    args = parser.parse_args()

    root = profile_root_from_args(args.root)
    ensure_dirs(root)
    if args.init_only:
        print(f"已初始化目录：{root}")
        return 0

    ingest_counts: dict[str, int] = {}
    if not args.rebuild_only:
        ingest_counts = ingest_pending_files(root, dry_run=args.dry_run)
    rebuild_counts = rebuild_profiles(root)
    cloud_sync = sync_profiles_to_cloud() if args.sync_cloud and not args.dry_run else None
    report = write_run_report(root, ingest_counts, rebuild_counts, dry_run=args.dry_run, cloud_sync=cloud_sync)
    print(f"客户档案处理完成：{root}")
    print(f"报告：{report}")
    print(json.dumps({"ingest": ingest_counts, "rebuild": rebuild_counts, "cloud_sync": cloud_sync}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
