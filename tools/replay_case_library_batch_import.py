from __future__ import annotations

import json
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from werkzeug.datastructures import FileStorage


PACKAGE_ROOT = Path(__file__).resolve().parents[1]
WORKSPACE_ROOT = PACKAGE_ROOT.parent
for search_path in (str(PACKAGE_ROOT), str(WORKSPACE_ROOT)):
    if search_path in sys.path:
        sys.path.remove(search_path)
sys.path.insert(0, str(WORKSPACE_ROOT))
sys.path.insert(0, str(PACKAGE_ROOT))

from tax_invoice_demo.models import InvoiceDraft  # noqa: E402
from tax_invoice_demo.workbench import create_draft_from_workbench  # noqa: E402
from tax_invoice_batch_demo.lean_workbench import export_draft_template  # noqa: E402


CASE_ROOT = WORKSPACE_ROOT / "invoice-demo" / "案例库原始材料"
REPORT_ROOT = PACKAGE_ROOT / "output" / "case_replay"


@dataclass(frozen=True)
class Scenario:
    scenario_id: str
    label: str
    company_name: str
    files: tuple[str, ...]
    buyer_name: str
    invoice_kind: str
    line_count: int
    total_amount: str
    project_names: tuple[str, ...]
    tax_categories: tuple[str, ...]
    tax_rates: tuple[str, ...]
    tax_codes: tuple[str, ...]
    expected_template_valid: bool = True
    expected_template_error_substring: str = ""


SCENARIOS = (
    Scenario(
        scenario_id="case5",
        label="案例 5 / Excel + DOCX 多行冷冻饮品",
        company_name="吉林省风生水起商贸有限公司",
        files=("5/source/开票.xlsx", "5/source/芃领开票信息.docx"),
        buyer_name="黑龙江芃领飞象网络科技有限公司",
        invoice_kind="增值税专用发票",
        line_count=3,
        total_amount="8181.16",
        project_names=("放学乐大菠萝", "大蜜瓜", "葚是喜欢"),
        tax_categories=("冷冻饮品",),
        tax_rates=("13%",),
        tax_codes=("1030209990000000000",),
    ),
    Scenario(
        scenario_id="case4",
        label="案例 4 / Excel 单票 141 行刀具五金",
        company_name="吉林省风生水起商贸有限公司",
        files=("4/source/宋总开发票15672元-2026.04.12.xls",),
        buyer_name="沈阳株硬合金工具有限公司",
        invoice_kind="增值税专用发票",
        line_count=141,
        total_amount="15672.00",
        project_names=("刀盘",),
        tax_categories=("金属制品",),
        tax_rates=("13%",),
        tax_codes=("1080403990000000000",),
    ),
    Scenario(
        scenario_id="case15a",
        label="案例 15A / PDF 塑料杯",
        company_name="吉林省风生水起商贸有限公司",
        files=("15/source/塑料杯.pdf",),
        buyer_name="前锦网络信息技术（上海）有限公司",
        invoice_kind="普通发票",
        line_count=1,
        total_amount="1750.00",
        project_names=("塑料杯 ABF796",),
        tax_categories=("塑料制品",),
        tax_rates=("1%",),
        tax_codes=("1070601010200000000",),
    ),
    Scenario(
        scenario_id="case15b",
        label="案例 15B / PDF 手机支架雨伞风扇",
        company_name="吉林省风生水起商贸有限公司",
        files=("15/source/平安利顺 手机支架 雨伞 风扇.pdf",),
        buyer_name="前锦网络信息技术（上海）有限公司",
        invoice_kind="普通发票",
        line_count=3,
        total_amount="5500.00",
        project_names=("手机支架", "雨伞", "风扇"),
        tax_categories=("计算机配套产品", "日用杂品", "家用通风电器具"),
        tax_rates=("1%",),
        tax_codes=("1090512990000000000", "1060512990000000000", "1090416990000000000"),
    ),
    Scenario(
        scenario_id="case15c",
        label="案例 15C / PDF 帆布袋",
        company_name="吉林省风生水起商贸有限公司",
        files=("15/source/鹏新旭 帆布袋.pdf",),
        buyer_name="前锦网络信息技术（上海）有限公司",
        invoice_kind="普通发票",
        line_count=1,
        total_amount="5700.00",
        project_names=("帆布袋",),
        tax_categories=("纺织产品",),
        tax_rates=("1%",),
        tax_codes=("1040103990000000000",),
    ),
    Scenario(
        scenario_id="case14_material_gap",
        label="案例 14 / 合同木制品草稿可生成但缺专票税号",
        company_name="吉林省风生水起商贸有限公司",
        files=("14/source/沈阳市东海包装材料有限公司39900合同.xlsx",),
        buyer_name="沈阳市东海包装材料有限公司",
        invoice_kind="增值税专用发票",
        line_count=3,
        total_amount="35309.75",
        project_names=("托盘",),
        tax_categories=("木制品",),
        tax_rates=("13%",),
        tax_codes=("1050101990000000000",),
        expected_template_valid=False,
        expected_template_error_substring="专票必须填写购买方纳税人识别号",
    ),
)


def replay_scenario(scenario: Scenario) -> dict:
    draft = _draft_from_case_files(scenario)
    export = export_draft_template(draft)
    detail_rows = _read_detail_rows(export["output_path"])
    template_error_messages = [issue.get("message", "") for issue in export["validation_issues"]]
    template_expectation = export["error_count"] == 0
    if not scenario.expected_template_valid:
        template_expectation = (
            export["error_count"] > 0
            and any(scenario.expected_template_error_substring in message for message in template_error_messages)
        )

    checks = {
        "buyer_name": scenario.buyer_name in draft.buyer.name,
        "invoice_kind": draft.invoice_kind == scenario.invoice_kind,
        "line_count": len(draft.lines) == scenario.line_count,
        "total_amount": draft.total_amount_with_tax == scenario.total_amount,
        "project_names": all(name in [line.project_name for line in draft.lines] for name in scenario.project_names),
        "tax_category": _unique_values(line.tax_category for line in draft.lines) == sorted(scenario.tax_categories),
        "tax_rate": _unique_values(line.normalized_tax_rate() for line in draft.lines) == sorted(scenario.tax_rates),
        "draft_tax_code": _unique_values(line.tax_code for line in draft.lines) == sorted(scenario.tax_codes),
        "template_tax_rate_decimal": _unique_values(row.get("税率", "") for row in detail_rows)
        == sorted(_rate_to_template_decimal(rate) for rate in scenario.tax_rates),
        "template_tax_code": _unique_values(row.get("商品和服务税收编码", "") for row in detail_rows)
        == sorted(scenario.tax_codes),
        "template_validation": template_expectation,
    }
    return {
        "scenario_id": scenario.scenario_id,
        "label": scenario.label,
        "passed": all(checks.values()),
        "checks": checks,
        "draft_id": draft.draft_id,
        "template_path": str(Path(export["output_path"]).resolve()),
        "actual": {
            "buyer_name": draft.buyer.name,
            "invoice_kind": draft.invoice_kind,
            "line_count": len(draft.lines),
            "total_amount": draft.total_amount_with_tax,
            "project_names": [line.project_name for line in draft.lines[:8]],
            "tax_categories": sorted({line.tax_category for line in draft.lines}),
            "tax_rates": sorted({line.normalized_tax_rate() for line in draft.lines}),
            "tax_codes": sorted({line.tax_code for line in draft.lines}),
            "template_error_count": export["error_count"],
            "template_validation_issues": export["validation_issues"][:5],
            "expected_template_valid": scenario.expected_template_valid,
        },
    }


def _draft_from_case_files(scenario: Scenario) -> InvoiceDraft:
    files = []
    handles = []
    try:
        for relative in scenario.files:
            path = CASE_ROOT / relative
            handle = path.open("rb")
            handles.append(handle)
            files.append(FileStorage(stream=handle, filename=path.name))
        draft = create_draft_from_workbench(
            company_name=scenario.company_name,
            raw_text="",
            note="",
            uploaded_files=files,
        )
        if not isinstance(draft, InvoiceDraft):
            raise RuntimeError(f"{scenario.scenario_id} 生成了批量草稿，当前脚本只校验单张导入模板。")
        return draft
    finally:
        for handle in handles:
            handle.close()


def _read_detail_rows(path: str | Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, data_only=True)
    try:
        sheet = workbook["2-发票明细信息"]
        headers = [str(cell.value or "").strip() for cell in sheet[3]]
        rows = []
        for row_number in range(4, sheet.max_row + 1):
            values = {header: str(sheet.cell(row=row_number, column=index + 1).value or "") for index, header in enumerate(headers) if header}
            if any(values.values()):
                rows.append(values)
        return rows
    finally:
        workbook.close()


def _rate_to_template_decimal(rate: str) -> str:
    if rate == "免税":
        return "0"
    return str(float(rate.rstrip("%")) / 100).rstrip("0").rstrip(".")


def _unique_values(values) -> list[str]:
    return sorted({str(value) for value in values})


def render_report(results: list[dict]) -> str:
    passed = sum(1 for item in results if item["passed"])
    lines = [
        "# 新线批量导入案例库回放报告",
        "",
        f"- 生成时间：`{datetime.now().isoformat(timespec='seconds')}`",
        f"- 通过：`{passed}/{len(results)}`",
        "- 说明：`case14_material_gap` 是预期材料缺口，草稿正确但源合同缺专票税号，模板校验失败属于预期。",
        "",
    ]
    for item in results:
        symbol = "✅" if item["passed"] else "❌"
        lines.extend(
            [
                f"## {symbol} {item['label']}",
                "",
                f"- 场景 ID：`{item['scenario_id']}`",
                f"- 草稿 ID：`{item['draft_id']}`",
                f"- 导入模板：`{item['template_path']}`",
                "- 校验：",
            ]
        )
        for key, value in item["checks"].items():
            lines.append(f"  - `{key}`: {'ok' if value else 'fail'}")
        lines.append(f"- 实际结果：`{json.dumps(item['actual'], ensure_ascii=False)}`")
        lines.append("")
    return "\n".join(lines)


def main() -> int:
    REPORT_ROOT.mkdir(parents=True, exist_ok=True)
    results = [replay_scenario(scenario) for scenario in SCENARIOS]
    report_json = REPORT_ROOT / "新线批量导入案例库回放报告.json"
    report_md = REPORT_ROOT / "新线批量导入案例库回放报告.md"
    report_json.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
    report_md.write_text(render_report(results), encoding="utf-8")
    passed = sum(1 for item in results if item["passed"])
    print(f"case replay pass: {passed}/{len(results)}")
    print(f"report: {report_md}")
    return 0 if passed == len(results) else 1


if __name__ == "__main__":
    raise SystemExit(main())
