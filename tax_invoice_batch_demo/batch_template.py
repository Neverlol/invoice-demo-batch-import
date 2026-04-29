from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OFFICIAL_TEMPLATE_DIR = ROOT / "official_templates"
LEGACY_TEMPLATE_PATH = OFFICIAL_TEMPLATE_DIR / "(V251101版)批量开票-导入开票模板.xlsx"


def latest_official_template_path(template_dir: str | Path = OFFICIAL_TEMPLATE_DIR) -> Path:
    """Return the newest bundled tax bureau batch-import template.

    Tax bureaus can reject an otherwise valid workbook when the embedded official
    template version is outdated. Prefer the highest `(Vxxxxxx版)` template in
    `official_templates/` so dropping a newly downloaded official template into
    that directory upgrades exports without code changes.
    """

    directory = Path(template_dir)
    candidates = [path for path in directory.glob("*.xlsx") if not path.name.startswith("~$")]
    if not candidates:
        return LEGACY_TEMPLATE_PATH
    return max(candidates, key=_official_template_sort_key)


def _official_template_sort_key(path: Path) -> tuple[int, float, str]:
    match = re.search(r"V(\d+)", path.name, flags=re.IGNORECASE)
    version = int(match.group(1)) if match else 0
    try:
        mtime = path.stat().st_mtime
    except OSError:
        mtime = 0.0
    return version, mtime, path.name


DEFAULT_TEMPLATE_PATH = latest_official_template_path()


@dataclass
class TemplateBuyer:
    name: str
    tax_id: str = ""
    address: str = ""
    phone: str = ""
    bank_name: str = ""
    bank_account: str = ""
    email: str = ""


@dataclass
class TemplateLine:
    project_name: str
    amount: str
    tax_rate: str
    tax_category: str = ""
    tax_code: str = ""
    specification: str = ""
    unit: str = ""
    quantity: str = ""
    unit_price: str = ""
    discount_amount: str = ""
    preferential_policy_flag: str = "否"
    preferential_policy_type: str = ""
    immediate_refund_type: str = ""
    coal_type: str = ""


@dataclass
class TemplateInvoice:
    serial_no: str
    invoice_type: str
    buyer: TemplateBuyer
    lines: list[TemplateLine]
    special_business_type: str = ""
    price_includes_tax: str = "是"
    note: str = ""
    natural_person_flag: str = "否"
    buyer_id_type: str = ""
    buyer_id_number: str = ""
    buyer_country_or_region: str = ""
    buyer_email: str = ""
    payee: str = ""
    reviewer: str = ""
    extra_fields: dict[str, str] = field(default_factory=dict)


@dataclass(frozen=True)
class TaxonomyEntry:
    official_code: str
    official_name: str
    category_short_name: str


def _default_taxonomy_path() -> Path:
    bundled = ROOT / "tax_invoice_demo" / "data" / "taxonomy_master_v0.1.csv"
    if bundled.exists():
        return bundled
    return ROOT.parent / "tax_invoice_demo" / "data" / "taxonomy_master_v0.1.csv"


DEFAULT_TAXONOMY_PATH = _default_taxonomy_path()


def export_workbench_draft(draft: Any, output_path: str | Path) -> Path:
    invoice = invoice_from_workbench_draft(draft)
    return export_template_invoices([invoice], output_path)


def invoice_from_workbench_draft(draft: Any, *, serial_no: str | None = None) -> TemplateInvoice:
    buyer = TemplateBuyer(
        name=getattr(draft.buyer, "name", ""),
        tax_id=getattr(draft.buyer, "tax_id", ""),
        address=getattr(draft.buyer, "address", ""),
        phone=getattr(draft.buyer, "phone", ""),
        bank_name=getattr(draft.buyer, "bank_name", ""),
        bank_account=getattr(draft.buyer, "bank_account", ""),
    )
    lines: list[TemplateLine] = []
    for line in getattr(draft, "lines", []):
        lines.append(
            TemplateLine(
                project_name=getattr(line, "project_name", ""),
                amount=_resolved_amount_with_tax(line),
                tax_rate=_normalized_tax_rate(getattr(line, "tax_rate", "")),
                tax_category=getattr(line, "tax_category", ""),
                tax_code=getattr(line, "tax_code", ""),
                specification=getattr(line, "specification", ""),
                unit=getattr(line, "unit", ""),
                quantity=getattr(line, "quantity", ""),
                unit_price=getattr(line, "unit_price", ""),
            )
        )
    return TemplateInvoice(
        serial_no=serial_no or getattr(draft, "draft_id", "DRAFT001"),
        invoice_type=_normalize_invoice_type(getattr(draft, "invoice_kind", "")),
        special_business_type=(getattr(draft, "special_business", "") or "").strip(),
        price_includes_tax="是",
        buyer=buyer,
        lines=lines,
        note=(getattr(draft, "note", "") or "").strip(),
    )


def export_template_invoices(
    invoices: Iterable[TemplateInvoice],
    output_path: str | Path,
    *,
    template_path: str | Path = DEFAULT_TEMPLATE_PATH,
) -> Path:
    workbook = load_workbook(template_path)
    basic_sheet = workbook["1-发票基本信息"]
    detail_sheet = workbook["2-发票明细信息"]
    extra_sheet = workbook["4-附加要素信息"]

    basic_headers = _sheet_headers(basic_sheet)
    detail_headers = _sheet_headers(detail_sheet)
    extra_headers = _sheet_headers(extra_sheet)

    basic_row = 4
    detail_row = 4
    extra_row = 4

    for invoice in invoices:
        _write_row(
            basic_sheet,
            basic_row,
            basic_headers,
            _build_basic_row(invoice),
        )
        basic_row += 1

        for line in invoice.lines:
            _write_row(
                detail_sheet,
                detail_row,
                detail_headers,
                _build_detail_row(invoice.serial_no, line),
            )
            detail_row += 1

        for key, value in invoice.extra_fields.items():
            if not value.strip():
                continue
            _write_row(
                extra_sheet,
                extra_row,
                extra_headers,
                {
                    "发票流水号": invoice.serial_no,
                    "附加要素名称": key,
                    "附加要素内容": value,
                },
            )
            extra_row += 1

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output


def _build_basic_row(invoice: TemplateInvoice) -> dict[str, str]:
    row = {
        "发票流水号": invoice.serial_no,
        "发票类型": _normalize_invoice_type(invoice.invoice_type),
        "特定业务类型": invoice.special_business_type,
        "是否含税": _normalize_yes_no(invoice.price_includes_tax, default="是"),
        "受票方自然人标识": _normalize_yes_no(invoice.natural_person_flag, default="否"),
        "购买方名称": invoice.buyer.name,
        "购买方纳税人识别号": invoice.buyer.tax_id,
        "购买方证件类型": invoice.buyer_id_type,
        "购买方证件号码": invoice.buyer_id_number,
        "购买方国籍（或地区）": invoice.buyer_country_or_region,
        "购买方地址": invoice.buyer.address,
        "购买方电话": invoice.buyer.phone,
        "购买方开户银行": invoice.buyer.bank_name,
        "购买方银行账号": invoice.buyer.bank_account,
        "是否展示购买方地址电话银行账号": _contact_display_option(
            invoice.extra_fields.get("是否展示购买方地址电话银行账号", ""),
            address=invoice.buyer.address,
            phone=invoice.buyer.phone,
            bank_name=invoice.buyer.bank_name,
            bank_account=invoice.buyer.bank_account,
        ),
        "备注": invoice.note,
        "购买方邮箱": invoice.buyer.email or invoice.buyer_email,
        "收款人": invoice.payee,
        "复核人": invoice.reviewer,
    }
    for key, value in invoice.extra_fields.items():
        if key in row:
            row[key] = value
    return row


def _build_detail_row(serial_no: str, line: TemplateLine) -> dict[str, str]:
    resolved_code, resolved_short_name = _resolve_taxonomy(line)
    return {
        "发票流水号": serial_no,
        "项目名称": line.project_name,
        "商品和服务税收编码": resolved_code,
        "规格型号": line.specification,
        "单位": line.unit,
        "数量": _stringify_decimal(line.quantity),
        "单价": _stringify_decimal(line.unit_price),
        "金额": _stringify_decimal(line.amount),
        "税率": _template_tax_rate_value(line.tax_rate),
        "折扣金额": _stringify_decimal(line.discount_amount),
        "是否使用优惠政策": _normalize_yes_no(line.preferential_policy_flag, default="否"),
        "优惠政策类型": line.preferential_policy_type,
        "即征即退类型": line.immediate_refund_type,
        "煤炭种类": line.coal_type,
        # 这里保留给后续本地校验/预览使用，官方模板没有“商品和服务分类简称”列。
        "_resolved_category_short_name": resolved_short_name,
    }


def _sheet_headers(sheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in sheet[3]:
        value = (cell.value or "").strip() if isinstance(cell.value, str) else cell.value
        if not value:
            continue
        headers[str(value)] = cell.column
    return headers


def _write_row(sheet, row_index: int, headers: dict[str, int], values: dict[str, str]) -> None:
    for key, value in values.items():
        if key.startswith("_"):
            continue
        column = headers.get(key)
        if column is None:
            continue
        sheet.cell(row=row_index, column=column, value=value)


def _resolved_amount_with_tax(line: Any) -> str:
    if hasattr(line, "resolved_amount_with_tax"):
        return getattr(line, "resolved_amount_with_tax")()
    return _stringify_decimal(getattr(line, "amount_with_tax", "") or getattr(line, "amount", ""))


def _normalized_tax_rate(raw: str) -> str:
    text = (raw or "").strip().replace("％", "%")
    if not text:
        return ""
    if text in {"免税", "不征税", "免征增值税"}:
        return "免税"
    if text.endswith("%"):
        return text
    try:
        value = Decimal(text)
    except InvalidOperation:
        return text
    if value <= Decimal("1"):
        value *= Decimal("100")
    return f"{value.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP).normalize()}%".replace(".00%", "%")


def _normalize_tax_rate(raw: str) -> str:
    text = _normalized_tax_rate(raw)
    return text or ""


def _template_tax_rate_value(raw: str) -> str:
    text = _normalized_tax_rate(raw)
    if not text:
        return ""
    if text in {"免税", "不征税", "免征增值税"}:
        return "0"
    from_percent_text = text.endswith("%")
    if text.endswith("%"):
        text = text[:-1]
    try:
        value = Decimal(text)
    except InvalidOperation:
        return text
    if from_percent_text or value > Decimal("1"):
        value = value / Decimal("100")
    return format(value.quantize(Decimal("0.0000000000001"), rounding=ROUND_HALF_UP).normalize(), "f")


def _normalize_invoice_type(raw: str) -> str:
    text = (raw or "").strip()
    if any(token in text for token in ("专票", "专用", "增值税专用发票")):
        return "增值税专用发票"
    return "普通发票"


def _normalize_yes_no(raw: str | bool, *, default: str = "否") -> str:
    if isinstance(raw, bool):
        return "是" if raw else "否"
    text = (raw or "").strip()
    if text in {"是", "否"}:
        return text
    if text.lower() in {"true", "yes", "y", "1", "on"}:
        return "是"
    if text.lower() in {"false", "no", "n", "0", "off"}:
        return "否"
    return default


def _contact_display_option(
    raw: str,
    *,
    address: str = "",
    phone: str = "",
    bank_name: str = "",
    bank_account: str = "",
) -> str:
    text = (raw or "").strip()
    valid_options = {
        "展示地址、电话",
        "展示开户银行、银行账号",
        "展示地址、电话、开户银行及银行账号",
    }
    if text in valid_options:
        return text
    if text in {"否", "不展示", "无需展示", "无", "0"}:
        return ""
    has_address_phone = bool((address or "").strip() or (phone or "").strip())
    has_bank = bool((bank_name or "").strip() or (bank_account or "").strip())
    if has_address_phone and has_bank:
        return "展示地址、电话、开户银行及银行账号"
    if has_address_phone:
        return "展示地址、电话"
    if has_bank:
        return "展示开户银行、银行账号"
    return ""


def _stringify_decimal(raw: str) -> str:
    text = (raw or "").strip()
    if not text:
        return ""
    compact = text.replace(",", "").replace("，", "").replace("￥", "").replace("¥", "")
    try:
        value = Decimal(compact)
    except InvalidOperation:
        return text
    return f"{value.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP):f}"


def _resolve_taxonomy(line: TemplateLine) -> tuple[str, str]:
    if line.tax_code.strip():
        short_name = line.tax_category.strip()
        matched = _match_taxonomy_by_code(line.tax_code.strip())
        if matched:
            matched = _prefer_leaf_taxonomy_entry(matched)
            short_name = short_name or matched.category_short_name
            return matched.official_code, short_name
        return line.tax_code.strip(), short_name

    query_candidates = [
        line.tax_category.strip(),
        line.project_name.strip(),
    ]
    for query in query_candidates:
        if not query:
            continue
        matched = _match_taxonomy_by_query(query, preferred_short_name=line.tax_category.strip())
        if matched is not None:
            matched = _prefer_leaf_taxonomy_entry(matched)
            return matched.official_code, matched.category_short_name
    return "", line.tax_category.strip()


def _match_taxonomy_by_code(code: str) -> TaxonomyEntry | None:
    for entry in _load_taxonomy_entries():
        if entry.official_code == code:
            return entry
    return None


def _match_taxonomy_by_query(query: str, *, preferred_short_name: str = "") -> TaxonomyEntry | None:
    normalized_query = _normalize(query)
    normalized_preferred = _normalize(preferred_short_name)
    best: tuple[int, TaxonomyEntry] | None = None
    for entry in _load_taxonomy_entries():
        official_norm = _normalize(entry.official_name)
        short_norm = _normalize(entry.category_short_name)
        score = 0
        if normalized_preferred and short_norm == normalized_preferred:
            score = 120
        elif normalized_query and official_norm == normalized_query:
            score = 110
        elif normalized_query and short_norm == normalized_query:
            score = 105
        elif normalized_query and normalized_query in official_norm:
            score = 90
        elif normalized_query and normalized_query in short_norm:
            score = 88
        elif normalized_query and short_norm and short_norm in normalized_query:
            score = 70
        if score and (best is None or score > best[0]):
            best = (score, entry)
    return best[1] if best else None


def _prefer_leaf_taxonomy_entry(entry: TaxonomyEntry) -> TaxonomyEntry:
    children = _child_taxonomy_entries(entry.official_code)
    if not children:
        return entry

    other_children = [
        child
        for child in children
        if child.official_name.startswith("其他") or f"其他{entry.official_name}" in child.official_name
    ]
    leaf_candidates = [
        child
        for child in (other_children or children)
        if not _child_taxonomy_entries(child.official_code)
    ]
    if leaf_candidates:
        return sorted(leaf_candidates, key=lambda item: item.official_code)[0]
    return sorted(children, key=lambda item: item.official_code)[0]


def _child_taxonomy_entries(code: str) -> list[TaxonomyEntry]:
    prefix = code.rstrip("0")
    if not prefix or prefix == code:
        return []
    return [
        entry
        for entry in _load_taxonomy_entries()
        if entry.official_code != code and entry.official_code.startswith(prefix)
    ]


def _load_taxonomy_entries() -> list[TaxonomyEntry]:
    entries: list[TaxonomyEntry] = []
    with DEFAULT_TAXONOMY_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            entries.append(
                TaxonomyEntry(
                    official_code=(row.get("official_code") or "").strip(),
                    official_name=(row.get("official_name") or "").strip(),
                    category_short_name=(row.get("category_short_name") or "").strip(),
                )
            )
    return entries


def _normalize(value: str) -> str:
    return re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "", value or "").upper()
