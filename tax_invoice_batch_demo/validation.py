from __future__ import annotations

import json
import csv
from dataclasses import asdict, dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from .batch_template import DEFAULT_TAXONOMY_PATH, DEFAULT_TEMPLATE_PATH


@dataclass(frozen=True)
class ValidationIssue:
    level: str
    sheet: str
    row: int
    column: int
    header: str
    value: str
    message: str


def validate_batch_workbook(
    path: str | Path,
    *,
    template_path: str | Path = DEFAULT_TEMPLATE_PATH,
) -> list[ValidationIssue]:
    workbook = load_workbook(path, data_only=True)
    template = load_workbook(template_path, data_only=True)
    issues: list[ValidationIssue] = []
    issues.extend(_validate_basic_and_detail(workbook))
    issues.extend(_validate_inline_lists(workbook, template))
    return issues


def build_validation_report(path: str | Path) -> dict[str, Any]:
    issues = validate_batch_workbook(path)
    return {
        "source_file": str(Path(path).resolve()),
        "issue_count": len(issues),
        "error_count": sum(1 for issue in issues if issue.level == "error"),
        "warning_count": sum(1 for issue in issues if issue.level == "warning"),
        "issues": [asdict(issue) for issue in issues],
    }


def save_validation_report(path: str | Path, output_path: str | Path) -> Path:
    report = build_validation_report(path)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return output


def _validate_basic_and_detail(workbook) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    if "1-发票基本信息" not in workbook.sheetnames:
        return [
            ValidationIssue("error", "", 0, 0, "", "", "缺少 sheet：1-发票基本信息"),
        ]
    if "2-发票明细信息" not in workbook.sheetnames:
        return [
            ValidationIssue("error", "", 0, 0, "", "", "缺少 sheet：2-发票明细信息"),
        ]

    basic_sheet = workbook["1-发票基本信息"]
    detail_sheet = workbook["2-发票明细信息"]
    basic_headers = _sheet_headers(basic_sheet)
    detail_headers = _sheet_headers(detail_sheet)

    basic_serials: set[str] = set()
    for row_number in _data_rows(basic_sheet, basic_headers.get("发票流水号")):
        serial_no = _cell_text(_cell_value(basic_sheet, row_number, basic_headers, "发票流水号"))
        invoice_type = _cell_text(_cell_value(basic_sheet, row_number, basic_headers, "发票类型"))
        buyer_name = _cell_text(_cell_value(basic_sheet, row_number, basic_headers, "购买方名称"))
        buyer_tax_id = _cell_text(_cell_value(basic_sheet, row_number, basic_headers, "购买方纳税人识别号"))
        price_includes_tax = _cell_text(_cell_value(basic_sheet, row_number, basic_headers, "是否含税"))
        buyer_address = _cell_text(_cell_value(basic_sheet, row_number, basic_headers, "购买方地址"))
        buyer_phone = _cell_text(_cell_value(basic_sheet, row_number, basic_headers, "购买方电话"))
        buyer_bank_name = _cell_text(_cell_value(basic_sheet, row_number, basic_headers, "购买方开户银行"))
        buyer_bank_account = _cell_text(_cell_value(basic_sheet, row_number, basic_headers, "购买方银行账号"))
        buyer_contact_display = _cell_text(_cell_value(basic_sheet, row_number, basic_headers, "是否展示购买方地址电话银行账号"))

        issues.extend(_require_fields(basic_sheet, row_number, basic_headers, ["发票流水号", "发票类型", "是否含税", "购买方名称"]))
        if serial_no in basic_serials:
            issues.append(_issue("error", basic_sheet, row_number, basic_headers, "发票流水号", serial_no, "发票流水号重复"))
        if serial_no:
            basic_serials.add(serial_no)
        if "专用" in invoice_type and not buyer_tax_id:
            issues.append(_issue("error", basic_sheet, row_number, basic_headers, "购买方纳税人识别号", buyer_tax_id, "专票必须填写购买方纳税人识别号"))
        if buyer_name and len(buyer_name) > 100:
            issues.append(_issue("error", basic_sheet, row_number, basic_headers, "购买方名称", buyer_name, "购买方名称超过 100 字符"))
        if price_includes_tax not in {"是", "否"}:
            issues.append(_issue("error", basic_sheet, row_number, basic_headers, "是否含税", price_includes_tax, "是否含税只能填写 是 或 否"))
        if (buyer_address or buyer_phone or buyer_bank_name or buyer_bank_account) and not buyer_contact_display:
            issues.append(
                _issue(
                    "error",
                    basic_sheet,
                    row_number,
                    basic_headers,
                    "是否展示购买方地址电话银行账号",
                    buyer_contact_display,
                    "已填写购买方地址/电话/开户行/账号时，应选择官方展示项，不能留空或填写 是/否",
                )
            )

    detail_serials: set[str] = set()
    for row_number in _data_rows(detail_sheet, detail_headers.get("发票流水号")):
        serial_no = _cell_text(_cell_value(detail_sheet, row_number, detail_headers, "发票流水号"))
        tax_rate = _cell_text(_cell_value(detail_sheet, row_number, detail_headers, "税率"))
        tax_code = _cell_text(_cell_value(detail_sheet, row_number, detail_headers, "商品和服务税收编码"))
        issues.extend(_require_fields(detail_sheet, row_number, detail_headers, ["发票流水号", "项目名称", "商品和服务税收编码", "金额", "税率"]))
        if serial_no:
            detail_serials.add(serial_no)
            if serial_no not in basic_serials:
                issues.append(_issue("error", detail_sheet, row_number, detail_headers, "发票流水号", serial_no, "明细中的发票流水号未在基本信息 sheet 中出现"))
        issues.extend(_validate_tax_rate(detail_sheet, row_number, detail_headers, tax_rate))
        if tax_code and _is_summary_taxonomy_code(tax_code):
            issues.append(_issue("error", detail_sheet, row_number, detail_headers, "商品和服务税收编码", tax_code, "该商品编码是汇总商品编码，税局要求使用下级具体商编"))

    for serial_no in sorted(basic_serials - detail_serials):
        issues.append(ValidationIssue("error", "2-发票明细信息", 0, 0, "发票流水号", serial_no, "基本信息存在该发票流水号，但明细信息没有对应行"))
    return issues


def _validate_inline_lists(workbook, template) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    for sheet_name in workbook.sheetnames:
        if sheet_name not in template.sheetnames:
            continue
        sheet = workbook[sheet_name]
        template_sheet = template[sheet_name]
        headers = _sheet_headers(sheet)
        for validation in template_sheet.data_validations.dataValidation:
            options = _inline_options(validation.formula1)
            if not options:
                continue
            for range_ref in str(validation.sqref).split():
                min_col, min_row, max_col, max_row = range_boundaries(range_ref)
                header = _header_for_column(headers, min_col)
                if not header:
                    continue
                for row_number in range(max(4, min_row), min(max_row, sheet.max_row) + 1):
                    value = _cell_text(sheet.cell(row=row_number, column=min_col).value)
                    if value and value not in options:
                        issues.append(
                            ValidationIssue(
                                "error",
                                sheet_name,
                                row_number,
                                min_col,
                                header,
                                value,
                                f"字段值不在官方模板下拉选项内，可选值：{' / '.join(options)}",
                            )
                        )
    return issues


def _validate_tax_rate(sheet, row_number: int, headers: dict[str, int], tax_rate: str) -> list[ValidationIssue]:
    if not tax_rate:
        return []
    if tax_rate.endswith("%"):
        return [_issue("error", sheet, row_number, headers, "税率", tax_rate, "税率在批量导入模板中应填写小数，例如 13% 应填写 0.13")]
    try:
        value = Decimal(tax_rate)
    except InvalidOperation:
        return [_issue("error", sheet, row_number, headers, "税率", tax_rate, "税率必须是小数，例如 0.13、0.03、0.01 或 0")]
    if value < 0 or value > 1:
        return [_issue("error", sheet, row_number, headers, "税率", tax_rate, "税率应在 0 到 1 之间，例如 13% 应填写 0.13")]
    return []


def _require_fields(sheet, row_number: int, headers: dict[str, int], field_names: list[str]) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    for field_name in field_names:
        value = _cell_text(_cell_value(sheet, row_number, headers, field_name))
        if not value:
            issues.append(_issue("error", sheet, row_number, headers, field_name, value, f"{field_name} 为必填项"))
    return issues


def _data_rows(sheet, serial_column: int | None) -> list[int]:
    if serial_column is None:
        return []
    rows: list[int] = []
    for row_number in range(4, sheet.max_row + 1):
        if any(_cell_text(sheet.cell(row=row_number, column=column).value) for column in range(1, sheet.max_column + 1)):
            rows.append(row_number)
    return rows


def _sheet_headers(sheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in sheet[3]:
        header = _cell_text(cell.value).replace("\n", "").replace("\r", "").strip()
        if header:
            headers[header] = cell.column
    return headers


def _cell_value(sheet, row_number: int, headers: dict[str, int], field_name: str) -> Any:
    column = headers.get(field_name)
    if column is None:
        return ""
    return sheet.cell(row=row_number, column=column).value


def _header_for_column(headers: dict[str, int], column: int) -> str:
    for header, header_column in headers.items():
        if header_column == column:
            return header
    return ""


def _issue(level: str, sheet, row_number: int, headers: dict[str, int], field_name: str, value: str, message: str) -> ValidationIssue:
    return ValidationIssue(
        level=level,
        sheet=sheet.title,
        row=row_number,
        column=headers.get(field_name, 0),
        header=field_name,
        value=value,
        message=message,
    )


def _inline_options(formula: str | None) -> list[str]:
    text = (formula or "").strip()
    if not (text.startswith('"') and text.endswith('"')):
        return []
    return [part.strip() for part in text[1:-1].split(",") if part.strip()]


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_summary_taxonomy_code(code: str) -> bool:
    prefix = code.rstrip("0")
    if not prefix or prefix == code:
        return False
    for entry_code in _taxonomy_codes():
        if entry_code != code and entry_code.startswith(prefix):
            return True
    return False


def _taxonomy_codes() -> list[str]:
    if not DEFAULT_TAXONOMY_PATH.exists():
        return []
    with DEFAULT_TAXONOMY_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
        return [(row.get("official_code") or "").strip() for row in csv.DictReader(handle)]
