from __future__ import annotations

import csv
import json
import sys
from copy import deepcopy
from dataclasses import asdict
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from werkzeug.datastructures import FileStorage

from .batch_template import export_template_invoices, invoice_from_workbench_draft
from .failure_details import build_failure_report
from .validation import validate_batch_workbook


PACKAGE_ROOT = Path(__file__).resolve().parents[1]
PROJECT_ROOT = PACKAGE_ROOT.parent
for search_path in (str(PACKAGE_ROOT), str(PROJECT_ROOT)):
    if search_path in sys.path:
        sys.path.remove(search_path)
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(PACKAGE_ROOT))

from tax_invoice_demo.models import BuyerInfo, DraftBatch, InvoiceDraft, InvoiceLine  # noqa: E402
from tax_invoice_demo.case_events import batch_snapshot, draft_snapshot, record_case_event  # noqa: E402
from tax_invoice_demo.workbench import (  # noqa: E402
    create_draft_from_workbench,
    draft_directory,
    load_draft,
    load_draft_batch,
    save_draft,
    update_draft_from_form,
)


BATCH_OUTPUT_ROOT = PACKAGE_ROOT / "output" / "batch_import_preview"
SUCCESS_LEDGER_CSV = BATCH_OUTPUT_ROOT / "批量导入成功明细.csv"
SUCCESS_LEDGER_XLSX = BATCH_OUTPUT_ROOT / "批量导入成功明细.xlsx"

SUCCESS_LEDGER_HEADERS = [
    "recorded_at",
    "case_id",
    "draft_id",
    "company_name",
    "invoice_kind",
    "buyer_name",
    "buyer_tax_id",
    "line_no",
    "project_name",
    "tax_category",
    "tax_code",
    "specification",
    "unit",
    "quantity",
    "unit_price",
    "amount_with_tax",
    "tax_rate",
    "coding_reference",
    "note",
]

REPAIR_FIELD_TO_LINE_ATTR = {
    "line_tax_rate": "tax_rate",
    "line_tax_code": "tax_code",
    "line_project_name": "project_name",
    "line_unit": "unit",
    "line_quantity": "quantity",
    "line_unit_price": "unit_price",
    "line_amount_with_tax": "amount_with_tax",
}


def default_form() -> dict[str, str]:
    return {
        "company_name": "",
        "raw_text": "",
        "note": "",
    }


def create_lean_draft(*, company_name: str, raw_text: str, note: str, uploaded_files: list[FileStorage]) -> InvoiceDraft | DraftBatch:
    return create_draft_from_workbench(
        company_name=company_name,
        raw_text=raw_text,
        note=note,
        uploaded_files=uploaded_files,
    )


def save_lean_draft_from_form(draft_id: str, form: dict[str, Any], uploaded_files: list[FileStorage] | None = None) -> InvoiceDraft:
    return update_draft_from_form(
        draft_id,
        company_name=(form.get("company_name") or "").strip(),
        raw_text=form.get("raw_text", ""),
        note=form.get("note", ""),
        buyer=BuyerInfo(
            name=(form.get("buyer_name") or "").strip(),
            tax_id=(form.get("buyer_tax_id") or "").strip(),
            address=(form.get("buyer_address") or "").strip(),
            phone=(form.get("buyer_phone") or "").strip(),
            bank_name=(form.get("buyer_bank_name") or "").strip(),
            bank_account=(form.get("buyer_bank_account") or "").strip(),
        ),
        lines=_lines_from_form(form),
        invoice_kind=form.get("invoice_kind") or "普通发票",
        invoice_medium="电子发票",
        special_business=(form.get("special_business") or "").strip(),
        uploaded_files=uploaded_files or [],
    )


def export_draft_template(draft: InvoiceDraft) -> dict[str, Any]:
    output_path = BATCH_OUTPUT_ROOT / f"{draft.draft_id}_batch_import.xlsx"
    invoice = invoice_from_workbench_draft(draft, serial_no=draft.draft_id)
    export_template_invoices([invoice], output_path)
    issues = validate_batch_workbook(output_path)
    result = {
        "output_path": output_path,
        "validation_issues": [asdict(issue) for issue in issues],
        "error_count": sum(1 for issue in issues if issue.level == "error"),
        "warning_count": sum(1 for issue in issues if issue.level == "warning"),
    }
    record_case_event(
        case_id=draft.case_id,
        draft_id=draft.draft_id,
        event_type="template_exported",
        payload={
            **draft_snapshot(draft),
            "output_path": str(output_path),
            "error_count": result["error_count"],
            "warning_count": result["warning_count"],
        },
    )
    return result


def export_batch_template(batch_id: str) -> dict[str, Any]:
    from .workbench_bridge import export_saved_workbench_items

    output_path = BATCH_OUTPUT_ROOT / f"{batch_id}_batch_import.xlsx"
    export_saved_workbench_items([batch_id], output_path)
    issues = validate_batch_workbook(output_path)
    result = {
        "output_path": output_path,
        "validation_issues": [asdict(issue) for issue in issues],
        "error_count": sum(1 for issue in issues if issue.level == "error"),
        "warning_count": sum(1 for issue in issues if issue.level == "warning"),
    }
    batch = load_draft_batch(batch_id)
    if batch is not None:
        record_case_event(
            case_id=batch.case_id,
            batch_id=batch.batch_id,
            event_type="batch_template_exported",
            payload={
                **batch_snapshot(batch),
                "output_path": str(output_path),
                "error_count": result["error_count"],
                "warning_count": result["warning_count"],
            },
        )
    return result


def parse_failure_file(file: FileStorage, *, draft: InvoiceDraft | None = None) -> dict[str, Any]:
    BATCH_OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    safe_name = Path(file.filename or "failure.xlsx").name
    target = BATCH_OUTPUT_ROOT / f"failure_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{safe_name}"
    file.save(target)
    report = build_failure_report(target)
    if draft is not None:
        report = enrich_failure_report_for_draft(report, draft)
        failure_summary = [
            {
                "field_name": item.get("field_name", ""),
                "failure_type": item.get("failure_type", ""),
                "target_label": item.get("target_label", ""),
                "repair_focus": item.get("repair_focus", ""),
            }
            for item in report.get("records", [])
        ]
        if failure_summary:
            draft.issues = [issue for issue in draft.issues if not issue.startswith("税局退回：")]
            draft.issues.extend(
                f"税局退回：{item['target_label'] or '整张发票'}，"
                f"{item['repair_focus'] or item['field_name'] or '请人工复核'}。"
                for item in failure_summary
            )
            save_draft(draft)
        save_failure_report_for_draft(draft.draft_id, report)
    if draft is not None:
        record_case_event(
            case_id=draft.case_id,
            draft_id=draft.draft_id,
            event_type="failure_report_uploaded",
            payload={
                "file_name": safe_name,
                "stored_path": str(target),
                "report": report,
            },
        )
    return report


def enrich_failure_report_for_draft(report: dict[str, Any], draft: InvoiceDraft) -> dict[str, Any]:
    enriched = deepcopy(report)
    for record in enriched.get("records", []):
        _attach_failure_target(record, draft)
    _refresh_failure_report_summary(enriched)
    return enriched


def load_failure_report_for_draft(draft_id: str, draft: InvoiceDraft | None = None) -> dict[str, Any] | None:
    path = _failure_report_path(draft_id)
    if not path.exists():
        return None
    report = json.loads(path.read_text(encoding="utf-8"))
    return enrich_failure_report_for_draft(report, draft) if draft is not None else report


def save_failure_report_for_draft(draft_id: str, report: dict[str, Any]) -> Path:
    path = _failure_report_path(draft_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def _failure_report_path(draft_id: str) -> Path:
    return draft_directory(draft_id) / "failure_report.json"


def apply_failure_repairs_to_draft(draft: InvoiceDraft) -> dict[str, Any]:
    report = load_failure_report_for_draft(draft.draft_id)
    if not report:
        return {
            "draft": draft,
            "failure_report": None,
            "applied_count": 0,
            "applied_records": [],
        }
    report = enrich_failure_report_for_draft(report, draft)
    applied_records: list[dict[str, str]] = []
    applied_at = datetime.now().isoformat(timespec="seconds")
    for record in report.get("records", []):
        line = _repair_target_line(record, draft)
        attr = REPAIR_FIELD_TO_LINE_ATTR.get(str(record.get("repair_field") or ""))
        value = str(record.get("repair_value") or "").strip()
        if line is None or not attr or not value:
            continue
        before = getattr(line, attr)
        setattr(line, attr, value)
        record["repair_status"] = "applied"
        record["applied_at"] = applied_at
        record["applied_value"] = value
        record["previous_value"] = before
        applied_records.append(
            {
                "target_label": str(record.get("target_label") or ""),
                "field_name": str(record.get("field_name") or ""),
                "repair_field": str(record.get("repair_field") or ""),
                "previous_value": before,
                "applied_value": value,
            }
        )
    _refresh_failure_report_summary(report)
    if applied_records:
        draft.issues = [
            issue
            for issue in draft.issues
            if not issue.startswith("税局退回：") and not issue.startswith("已应用税局建议：")
        ]
        draft.issues.extend(
            f"已应用税局建议：{item['target_label'] or '明细行'}，"
            f"{item['field_name'] or '字段'}改为 {item['applied_value']}。"
            for item in applied_records
        )
        save_draft(draft)
        record_case_event(
            case_id=draft.case_id,
            draft_id=draft.draft_id,
            event_type="failure_repairs_applied",
            payload={
                "applied_at": applied_at,
                "applied_count": len(applied_records),
                "applied_records": applied_records,
            },
        )
    save_failure_report_for_draft(draft.draft_id, report)
    return {
        "draft": draft,
        "failure_report": report,
        "applied_count": len(applied_records),
        "applied_records": applied_records,
    }


def record_success_to_ledger(draft: InvoiceDraft) -> Path:
    BATCH_OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    rows = _read_success_rows()
    rows = [row for row in rows if row.get("draft_id") != draft.draft_id]
    recorded_at = datetime.now().isoformat(timespec="seconds")
    for index, line in enumerate(draft.lines, start=1):
        rows.append(
            {
                "recorded_at": recorded_at,
                "case_id": draft.case_id,
                "draft_id": draft.draft_id,
                "company_name": draft.company_name,
                "invoice_kind": draft.invoice_kind,
                "buyer_name": draft.buyer.name,
                "buyer_tax_id": draft.buyer.tax_id,
                "line_no": str(index),
                "project_name": line.project_name,
                "tax_category": line.tax_category,
                "tax_code": line.tax_code,
                "specification": line.specification,
                "unit": line.unit,
                "quantity": line.quantity,
                "unit_price": line.unit_price,
                "amount_with_tax": line.resolved_amount_with_tax(),
                "tax_rate": line.normalized_tax_rate(),
                "coding_reference": line.coding_reference,
                "note": draft.note,
            }
        )
    _write_success_rows(rows)
    record_case_event(
        case_id=draft.case_id,
        draft_id=draft.draft_id,
        event_type="success_recorded",
        payload={
            **draft_snapshot(draft),
            "recorded_at": recorded_at,
        },
    )
    return SUCCESS_LEDGER_XLSX


def draft_preview(draft: InvoiceDraft) -> dict[str, Any]:
    amount_total = Decimal("0")
    tax_total = Decimal("0")
    line_rows: list[dict[str, str]] = []
    for index, line in enumerate(draft.lines, start=1):
        amount = _decimal(line.resolved_amount_with_tax())
        rate = _rate_decimal(line.normalized_tax_rate())
        tax_amount = _money(amount - (amount / (Decimal("1") + rate))) if amount is not None and rate is not None else ""
        if amount is not None:
            amount_total += amount
        if tax_amount:
            tax_total += Decimal(tax_amount)
        line_rows.append(
            {
                "line_no": str(index),
                "project_name": line.project_name,
                "tax_category": line.tax_category,
                "tax_code": line.tax_code,
                "specification": line.specification,
                "unit": line.unit,
                "quantity": line.quantity,
                "unit_price": line.unit_price,
                "amount_with_tax": line.resolved_amount_with_tax(),
                "tax_rate": line.normalized_tax_rate(),
                "coding_reference": line.coding_reference,
                "tax_amount_preview": tax_amount,
            }
        )
    return {
        "amount_total": _money(amount_total),
        "tax_total": _money(tax_total),
        "line_rows": line_rows,
        "issues": draft.issues,
        "attachments": draft.source_images,
    }


def line_form_rows(draft: InvoiceDraft, failure_report: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    rows = draft_preview(draft)["line_rows"] or [
        {
            "project_name": "",
            "tax_category": "",
            "tax_code": "",
            "specification": "",
            "unit": "",
            "quantity": "",
            "unit_price": "",
            "amount_with_tax": "",
            "tax_rate": "3%",
            "coding_reference": "",
        }
    ]
    failures_by_line = _failures_by_line(failure_report)
    for index, row in enumerate(rows, start=1):
        row["failure_alerts"] = failures_by_line.get(index, [])
    return rows


def _attach_failure_target(record: dict[str, Any], draft: InvoiceDraft) -> None:
    field_name = str(record.get("field_name") or "")
    source_sheet = str(record.get("source_sheet") or "")
    reason = str(record.get("reason") or "")
    target_line_no = _infer_target_line_no(record, draft)
    record["target_line_no"] = str(target_line_no) if target_line_no else ""
    if target_line_no:
        project_name = draft.lines[target_line_no - 1].project_name if target_line_no <= len(draft.lines) else ""
        record["target_scope"] = "line"
        record["target_label"] = f"第 {target_line_no} 行{('：' + project_name) if project_name else ''}"
    elif source_sheet == "1-发票基本信息" or field_name in {"购买方名称", "购买方纳税人识别号", "购买方税号", "发票类型", "特定业务类型"}:
        record["target_scope"] = "draft"
        record["target_label"] = "发票抬头 / 基本信息"
    else:
        record["target_scope"] = "invoice"
        record["target_label"] = "整张发票"
    record["repair_focus"] = _repair_focus(field_name, reason)
    record["repair_field"] = _repair_field(field_name)
    record["repair_value"] = _repair_value(record)
    if record.get("repair_status") != "applied" and record["repair_field"] and record["repair_value"] and target_line_no:
        record["repair_status"] = "ready"


def _infer_target_line_no(record: dict[str, Any], draft: InvoiceDraft) -> int:
    field_name = str(record.get("field_name") or "")
    source_sheet = str(record.get("source_sheet") or "")
    reason = str(record.get("reason") or "")
    if source_sheet != "2-发票明细信息" and field_name not in {
        "商品和服务税收编码",
        "税率",
        "金额",
        "项目名称",
        "单位",
        "数量",
        "单价",
    }:
        return 0
    row_number = _extract_row_number(reason)
    if row_number >= 4:
        candidate = row_number - 3
        if 1 <= candidate <= len(draft.lines):
            return candidate
    if len(draft.lines) == 1:
        return 1
    return 0


def _extract_row_number(reason: str) -> int:
    import re

    match = re.search(r"第\s*(\d+)\s*行", reason)
    if not match:
        return 0
    try:
        return int(match.group(1))
    except ValueError:
        return 0


def _repair_focus(field_name: str, reason: str) -> str:
    if field_name in {"商品和服务税收编码", "商品和服务分类简称"}:
        return "检查税收编码，必要时改成税局允许的下级具体编码"
    if field_name == "税率":
        return "按税局允许税率调整这一行后重建模板"
    if field_name in {"购买方纳税人识别号", "购买方税号"}:
        return "核对购买方税号格式和专票必填要求"
    if field_name == "购买方名称":
        return "核对购买方名称是否为空、过长或与税号不匹配"
    if field_name in {"单位", "数量", "单价", "金额", "项目名称"}:
        return f"补齐或修正明细里的{field_name}"
    if "不属于涉税专业服务机构" in reason:
        return "这是销售方资质限制，先确认开票主体或替代税目口径"
    return "按税局失败原因复核草稿字段"


def _repair_field(field_name: str) -> str:
    if field_name == "税率":
        return "line_tax_rate"
    if field_name in {"商品和服务税收编码", "商品和服务分类简称"}:
        return "line_tax_code"
    if field_name == "项目名称":
        return "line_project_name"
    if field_name == "单位":
        return "line_unit"
    if field_name == "数量":
        return "line_quantity"
    if field_name == "单价":
        return "line_unit_price"
    if field_name == "金额":
        return "line_amount_with_tax"
    return ""


def _repair_value(record: dict[str, Any]) -> str:
    field_name = str(record.get("field_name") or "")
    if field_name == "税率":
        return str(record.get("suggested_value") or "")
    return ""


def _repair_target_line(record: dict[str, Any], draft: InvoiceDraft) -> InvoiceLine | None:
    try:
        line_no = int(record.get("target_line_no") or 0)
    except (TypeError, ValueError):
        return None
    if line_no < 1 or line_no > len(draft.lines):
        return None
    return draft.lines[line_no - 1]


def _refresh_failure_report_summary(report: dict[str, Any]) -> None:
    records = report.get("records", [])
    actionable_count = 0
    applied_count = 0
    for record in records:
        if record.get("repair_status") == "applied":
            applied_count += 1
        if (
            record.get("target_line_no")
            and record.get("repair_field")
            and record.get("repair_value")
            and record.get("repair_status") != "applied"
        ):
            actionable_count += 1
    report["actionable_count"] = actionable_count
    report["applied_count"] = applied_count


def _failures_by_line(failure_report: dict[str, Any] | None) -> dict[int, list[dict[str, str]]]:
    if not failure_report:
        return {}
    result: dict[int, list[dict[str, str]]] = {}
    for record in failure_report.get("records", []):
        try:
            line_no = int(record.get("target_line_no") or 0)
        except (TypeError, ValueError):
            line_no = 0
        if line_no <= 0:
            continue
        result.setdefault(line_no, []).append(record)
    return result


def _lines_from_form(form: dict[str, Any]) -> list[InvoiceLine]:
    project_names = _form_list(form, "line_project_name")
    lines: list[InvoiceLine] = []
    for index, project_name in enumerate(project_names):
        if not any(
            [
                project_name.strip(),
                _form_value(form, "line_amount_with_tax", index),
                _form_value(form, "line_tax_code", index),
            ]
        ):
            continue
        lines.append(
            InvoiceLine(
                project_name=project_name.strip(),
                amount_with_tax=_form_value(form, "line_amount_with_tax", index),
                tax_rate=_form_value(form, "line_tax_rate", index) or "3%",
                tax_category=_form_value(form, "line_tax_category", index),
                tax_code=_form_value(form, "line_tax_code", index),
                specification=_form_value(form, "line_specification", index),
                unit=_form_value(form, "line_unit", index),
                quantity=_form_value(form, "line_quantity", index),
                unit_price=_form_value(form, "line_unit_price", index),
                coding_reference=_form_value(form, "line_coding_reference", index),
            )
        )
    return lines


def _form_list(form: dict[str, Any], key: str) -> list[str]:
    if hasattr(form, "getlist"):
        return [str(item) for item in form.getlist(key)]
    value = form.get(key, [])
    if isinstance(value, list):
        return [str(item) for item in value]
    return [str(value)] if value else []


def _form_value(form: dict[str, Any], key: str, index: int) -> str:
    values = _form_list(form, key)
    return values[index].strip() if index < len(values) else ""


def _read_success_rows() -> list[dict[str, str]]:
    if not SUCCESS_LEDGER_CSV.exists():
        return []
    with SUCCESS_LEDGER_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _write_success_rows(rows: list[dict[str, str]]) -> None:
    with SUCCESS_LEDGER_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=SUCCESS_LEDGER_HEADERS)
        writer.writeheader()
        writer.writerows(rows)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "批量导入成功明细"
    sheet.append(SUCCESS_LEDGER_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCEFE9")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        sheet.append([row.get(header, "") for header in SUCCESS_LEDGER_HEADERS])
    sheet.freeze_panes = "A2"
    last_column = get_column_letter(len(SUCCESS_LEDGER_HEADERS))
    sheet.auto_filter.ref = f"A1:{last_column}{max(sheet.max_row, 1)}"
    for column in range(1, len(SUCCESS_LEDGER_HEADERS) + 1):
        sheet.column_dimensions[chr(64 + column)].width = 16
    workbook.save(SUCCESS_LEDGER_XLSX)


def _decimal(raw: str) -> Decimal | None:
    text = (raw or "").replace(",", "").replace("，", "").strip()
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _rate_decimal(raw: str) -> Decimal | None:
    text = (raw or "").strip().replace("％", "%")
    if text in {"免税", "不征税", "免征增值税"}:
        return Decimal("0")
    is_percent_text = text.endswith("%")
    if is_percent_text:
        text = text[:-1]
    try:
        value = Decimal(text)
    except InvalidOperation:
        return None
    if is_percent_text:
        return value / Decimal("100")
    if value > Decimal("1"):
        value = value / Decimal("100")
    return value


def _money(value: Decimal | None) -> str:
    if value is None:
        return ""
    return f"{value.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP):f}"
