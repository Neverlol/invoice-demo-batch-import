from __future__ import annotations

import json
import re
from dataclasses import asdict, dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True)
class FailureRecord:
    row_number: int
    serial_no: str
    invoice_type: str
    special_business_type: str
    buyer_name: str
    buyer_tax_id: str
    reason: str
    source_sheet: str
    field_name: str
    failure_type: str
    suggested_action: str
    allowed_values: list[str]
    suggested_value: str


HEADER_ALIASES = {
    "serial_no": ("发票流水号",),
    "invoice_type": ("发票类型",),
    "special_business_type": ("特定业务类型",),
    "buyer_name": ("购买方名称",),
    "buyer_tax_id": ("购买方税号", "购买方纳税人识别号"),
    "reason": ("导入失败原因", "失败原因", "错误原因", "导入结果"),
}

KNOWN_FAILURE_FIELDS = [
    "是否展示购买方地址电话银行账号",
    "是否展示销售方地址电话银行账号",
    "商品编码",
    "商品和服务税收编码",
    "商品和服务税收分类编码",
    "商品和服务分类简称",
    "购买方纳税人识别号",
    "购买方税号",
    "购买方名称",
    "发票流水号",
    "发票类型",
    "特定业务类型",
    "是否含税",
    "税率",
    "金额",
    "项目名称",
    "单位",
    "数量",
    "单价",
    "备注",
]


def parse_failure_workbook(path: str | Path) -> list[FailureRecord]:
    workbook = load_workbook(path, data_only=True)
    records: list[FailureRecord] = []
    for worksheet in workbook.worksheets:
        header_row, headers = _detect_header_row(worksheet)
        if not headers:
            continue
        reason_col = _find_column(headers, "reason")
        if reason_col is None:
            continue
        for row_number in range(header_row + 1, worksheet.max_row + 1):
            row_values = _row_values(worksheet, row_number)
            reason = _cell_text(row_values.get(reason_col))
            if not reason:
                continue
            field_name = _infer_field_name(reason)
            failure_type = _infer_failure_type(reason)
            allowed_values = _extract_allowed_values(reason, failure_type)
            records.append(
                FailureRecord(
                    row_number=row_number,
                    serial_no=_get_alias_value(row_values, headers, "serial_no"),
                    invoice_type=_get_alias_value(row_values, headers, "invoice_type"),
                    special_business_type=_get_alias_value(row_values, headers, "special_business_type"),
                    buyer_name=_get_alias_value(row_values, headers, "buyer_name"),
                    buyer_tax_id=_get_alias_value(row_values, headers, "buyer_tax_id"),
                    reason=reason,
                    source_sheet=_infer_source_sheet(reason, field_name),
                    field_name=field_name,
                    failure_type=failure_type,
                    suggested_action=_suggested_action(failure_type),
                    allowed_values=allowed_values,
                    suggested_value=allowed_values[0] if allowed_values else "",
                )
            )
    return records


def build_failure_report(path: str | Path) -> dict[str, Any]:
    records = parse_failure_workbook(path)
    return {
        "source_file": str(Path(path).resolve()),
        "failure_count": len(records),
        "records": [asdict(record) for record in records],
        "summary_by_field": _count_by(records, "field_name"),
        "summary_by_type": _count_by(records, "failure_type"),
        "summary_by_sheet": _count_by(records, "source_sheet"),
    }


def save_failure_report(path: str | Path, output_path: str | Path) -> Path:
    report = build_failure_report(path)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return output


def _detect_header_row(worksheet) -> tuple[int, dict[int, str]]:
    best_row = 0
    best_headers: dict[int, str] = {}
    best_score = 0
    for row_number in range(1, min(10, worksheet.max_row) + 1):
        headers = {
            cell.column: _normalize_header(cell.value)
            for cell in worksheet[row_number]
            if _normalize_header(cell.value)
        }
        score = sum(
            1
            for aliases in HEADER_ALIASES.values()
            if any(alias in headers.values() for alias in aliases)
        )
        if score > best_score:
            best_row = row_number
            best_headers = headers
            best_score = score
    return best_row, best_headers


def _row_values(worksheet, row_number: int) -> dict[int, Any]:
    return {cell.column: cell.value for cell in worksheet[row_number]}


def _get_alias_value(row_values: dict[int, Any], headers: dict[int, str], alias_key: str) -> str:
    column = _find_column(headers, alias_key)
    if column is None:
        return ""
    return _cell_text(row_values.get(column))


def _find_column(headers: dict[int, str], alias_key: str) -> int | None:
    aliases = HEADER_ALIASES[alias_key]
    for column, header in headers.items():
        if header in aliases:
            return column
    return None


def _infer_source_sheet(reason: str, field_name: str = "") -> str:
    match = re.match(r"\s*([^:：]+)[:：]", reason)
    if match:
        return match.group(1).strip()
    if field_name in {
        "商品和服务税收编码",
        "商品和服务分类简称",
        "税率",
        "金额",
        "项目名称",
        "单位",
        "数量",
        "单价",
    }:
        return "2-发票明细信息"
    return ""


def _infer_field_name(reason: str) -> str:
    if "汇总商品" in reason or "具体商编" in reason:
        return "商品和服务税收编码"
    for field_name in KNOWN_FAILURE_FIELDS:
        if field_name in reason:
            if field_name in {"商品编码", "商品和服务税收分类编码"}:
                return "商品和服务税收编码"
            return field_name
    if "不允许填写" in reason and re.search(r"\b\d{19}\b", reason):
        return "商品和服务税收编码"
    match = re.search(r"([\u4e00-\u9fffA-Za-z0-9（）()/%]+?)(?:码值不正确|不能为空|不能为(?:空|必填)|必填|格式不正确)", reason)
    if match:
        return match.group(1).strip("，,。:： ")
    return ""


def _infer_failure_type(reason: str) -> str:
    if "不属于涉税专业服务机构" in reason:
        return "seller_qualification_restriction"
    if "税率不合法" in reason and "请使用如下税率" in reason:
        return "seller_tax_rate_restriction"
    if "汇总商品" in reason or "具体商编" in reason:
        return "taxonomy_code_level_error"
    if "码值不正确" in reason or "字段值不在官方模板下拉选项" in reason:
        return "template_option_error"
    if "不能为空" in reason or "必填" in reason:
        return "missing_required_field"
    return "tax_bureau_validation_error"


def _suggested_action(failure_type: str) -> str:
    if failure_type == "seller_qualification_restriction":
        return (
            "这是当前销售方主体权限/资质限制，不是模板格式错误，也不要直接回写成赋码库错误。"
            "请改用具备对应资质的开票主体，或由会计人工确认是否存在可开具的替代税目口径。"
        )
    if failure_type == "seller_tax_rate_restriction":
        return "这是当前销售方主体/票种允许税率限制。请按税局返回的可用税率调整草稿后重建模板。"
    if failure_type == "taxonomy_code_level_error":
        return "当前编码层级过粗。请改用税局允许的下级具体商品和服务税收编码。"
    if failure_type == "template_option_error":
        return "当前字段值不符合税局模板下拉项。请按失败原因改成官方允许值。"
    if failure_type == "missing_required_field":
        return "模板存在必填字段缺失。请回到草稿补齐对应字段后重建模板。"
    return "税局返回业务校验失败。请按失败原因人工复核草稿字段和当前开票主体限制。"


def _extract_allowed_values(reason: str, failure_type: str) -> list[str]:
    if failure_type != "seller_tax_rate_restriction":
        return []
    segment = reason
    if "请使用如下税率" in reason:
        segment = reason.split("请使用如下税率", 1)[1]
    values: list[str] = []
    for raw in re.findall(r"\d+(?:\.\d+)?\s*%?|\d+(?:\.\d+)?\s*％", segment):
        normalized = _normalize_tax_rate_value(raw)
        if normalized and normalized not in values:
            values.append(normalized)
    return values


def _normalize_tax_rate_value(raw: str) -> str:
    value = raw.strip().replace("％", "%").replace(" ", "")
    if not value:
        return ""
    if value.endswith("%"):
        number = value[:-1]
        try:
            decimal_value = Decimal(number)
        except InvalidOperation:
            return value
        return _format_percent(decimal_value)
    try:
        decimal_value = Decimal(value)
    except InvalidOperation:
        return ""
    if decimal_value <= Decimal("1"):
        decimal_value *= Decimal("100")
    return _format_percent(decimal_value)


def _format_percent(value: Decimal) -> str:
    quantized = value.quantize(Decimal("0.01"))
    text = format(quantized, "f").rstrip("0").rstrip(".")
    return f"{text or '0'}%"


def _count_by(records: list[FailureRecord], field: str) -> dict[str, int]:
    result: dict[str, int] = {}
    for record in records:
        key = getattr(record, field) or "未识别"
        result[key] = result.get(key, 0) + 1
    return result


def _normalize_header(value: Any) -> str:
    return _cell_text(value).replace("\n", "").replace("\r", "").replace(" ", "")


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()
