from __future__ import annotations

import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Callable
from urllib.parse import urlparse

from openpyxl import load_workbook

from .failure_details import build_failure_report


StatusHook = Callable[[str, str], None]
DOWNLOAD_ROOT = Path(__file__).resolve().parents[1] / "output" / "batch_import_preview" / "failure_downloads"


@dataclass
class BatchRunResult:
    status: str
    current_step: str
    logs: list[str] = field(default_factory=list)
    error: str = ""
    downloaded_failure_path: str = ""
    failure_report: dict | None = None
    preview_clicked: bool = False


TAX_PORTAL_URLS = {
    "liaoning": "https://dppt.liaoning.chinatax.gov.cn:8443/",
    "jilin": "https://dppt.jilin.chinatax.gov.cn:8443/",
    "heilongjiang": "https://dppt.heilongjiang.chinatax.gov.cn:8443/",
    "beijing": "https://dppt.beijing.chinatax.gov.cn:8443/",
}


def inspect_tax_browser(cdp_endpoint: str = "http://127.0.0.1:9222") -> dict:
    try:
        from playwright.sync_api import sync_playwright
    except Exception as exc:  # noqa: BLE001
        return {"status": "error", "error": f"Playwright 不可用: {type(exc).__name__}: {exc}", "pages": []}

    try:
        with sync_playwright() as p:
            browser = p.chromium.connect_over_cdp(cdp_endpoint)
            try:
                pages = [page for context in browser.contexts for page in context.pages]
                inspected = []
                for page in pages:
                    title = _safe_title(page)
                    text = _safe_body_text(page)
                    inspected.append(
                        {
                            "url": page.url,
                            "title": title,
                            "score": _tax_page_score(page.url, title, text),
                            "subject": _extract_tax_subject(page),
                            "is_batch_import_page": _is_batch_import_page(page),
                            "is_tax_portal_home": _looks_like_tax_portal_home(page),
                        }
                    )
                inspected.sort(key=lambda item: item.get("score", 0), reverse=True)
                best = inspected[0] if inspected else {}
                return {
                    "status": "ok",
                    "cdp_endpoint": cdp_endpoint,
                    "page_count": len(inspected),
                    "best_page": best,
                    "subject": best.get("subject", ""),
                    "pages": inspected[:8],
                }
            finally:
                browser.close()
    except Exception as exc:  # noqa: BLE001
        return {"status": "error", "cdp_endpoint": cdp_endpoint, "error": f"{type(exc).__name__}: {exc}", "pages": []}


def open_tax_portal(
    cdp_endpoint: str = "http://127.0.0.1:9222",
    *,
    province: str = "liaoning",
    url: str = "",
) -> dict:
    target_url = (url or TAX_PORTAL_URLS.get(province) or TAX_PORTAL_URLS["liaoning"]).strip()
    try:
        from playwright.sync_api import sync_playwright
    except Exception as exc:  # noqa: BLE001
        return {"status": "error", "url": target_url, "error": f"Playwright 不可用: {type(exc).__name__}: {exc}"}

    try:
        with sync_playwright() as p:
            browser = p.chromium.connect_over_cdp(cdp_endpoint)
            try:
                context = browser.contexts[0] if browser.contexts else browser.new_context()
                page = context.new_page()
                page.goto(target_url, wait_until="domcontentloaded", timeout=30000)
                page.wait_for_timeout(1000)
                return {"status": "ok", "url": page.url, "title": _safe_title(page)}
            finally:
                browser.close()
    except Exception as exc:  # noqa: BLE001
        return {"status": "error", "url": target_url, "error": f"{type(exc).__name__}: {exc}"}


def focus_tax_window(cdp_endpoint: str = "http://127.0.0.1:9222") -> dict:
    """Bring the existing tax-bureau tab to the front without clicking anything.

    This is intentionally only a window/tab switch helper. The batch runner has
    already navigated the tax site to the preview stage; this function must not
    upload, preview-click, submit, or record anything.
    """

    try:
        from playwright.sync_api import sync_playwright
    except Exception as exc:  # noqa: BLE001
        return {"status": "error", "error": f"Playwright 不可用: {type(exc).__name__}: {exc}"}

    try:
        with sync_playwright() as p:
            browser = p.chromium.connect_over_cdp(cdp_endpoint)
            try:
                pages = [page for context in browser.contexts for page in context.pages]
                if not pages:
                    return {"status": "error", "error": "未找到 Edge 税局页面。请确认税局页面仍然打开。"}

                ranked = []
                for page in pages:
                    title = _safe_title(page)
                    text = _safe_body_text(page)
                    score = _tax_page_score(page.url, title, text)
                    if any(token in text for token in ("预览发票", "发票预览", "发票价税合计", "最终开具")):
                        score += 80
                    ranked.append((score, page, title))
                ranked.sort(key=lambda item: item[0], reverse=True)
                score, page, title = ranked[0]
                if score <= 0:
                    return {"status": "error", "error": "未识别到税局窗口。请确认 Edge 已打开电子税务局页面。"}
                page.bring_to_front()
                try:
                    session = page.context.new_cdp_session(page)
                    window_info = session.send("Browser.getWindowForTarget")
                    session.send(
                        "Browser.setWindowBounds",
                        {"windowId": window_info["windowId"], "bounds": {"windowState": "normal"}},
                    )
                except Exception:  # noqa: BLE001
                    pass
                try:
                    page.evaluate("() => window.focus()")
                except Exception:  # noqa: BLE001
                    pass
                os_focused = _focus_windows_browser_window(title)
                return {"status": "ok", "url": page.url, "title": title, "score": score, "os_focused": os_focused}
            finally:
                browser.close()
    except Exception as exc:  # noqa: BLE001
        return {"status": "error", "error": f"{type(exc).__name__}: {exc}"}


def _focus_windows_browser_window(title_hint: str) -> bool:
    if sys.platform != "win32":
        return False
    try:
        import ctypes
        from ctypes import wintypes
    except Exception:  # noqa: BLE001
        return False

    user32 = ctypes.windll.user32
    kernel32 = ctypes.windll.kernel32
    SW_RESTORE = 9
    ASFW_ANY = -1
    title_hint = (title_hint or "").strip()
    title_tokens = [token for token in (title_hint, "电子税务局", "发票", "税局") if token]
    candidates: list[tuple[int, int, str]] = []

    EnumWindowsProc = ctypes.WINFUNCTYPE(wintypes.BOOL, wintypes.HWND, wintypes.LPARAM)

    def callback(hwnd, _lparam):
        if not user32.IsWindowVisible(hwnd):
            return True
        length = user32.GetWindowTextLengthW(hwnd)
        if length <= 0:
            return True
        buffer = ctypes.create_unicode_buffer(length + 1)
        user32.GetWindowTextW(hwnd, buffer, length + 1)
        title = buffer.value or ""
        lowered = title.lower()
        if "edge" not in lowered and "chrome" not in lowered:
            return True
        score = 0
        for token in title_tokens:
            if token and token in title:
                score += 20 if token == title_hint else 8
        if "microsoft edge" in lowered:
            score += 5
        if score > 0:
            candidates.append((score, int(hwnd), title))
        return True

    try:
        user32.EnumWindows(EnumWindowsProc(callback), 0)
    except Exception:  # noqa: BLE001
        return False
    if not candidates:
        return False

    candidates.sort(key=lambda item: item[0], reverse=True)
    hwnd = candidates[0][1]
    try:
        user32.AllowSetForegroundWindow(ASFW_ANY)
    except Exception:  # noqa: BLE001
        pass
    try:
        foreground = user32.GetForegroundWindow()
        current_thread = kernel32.GetCurrentThreadId()
        target_thread = user32.GetWindowThreadProcessId(hwnd, None)
        foreground_thread = user32.GetWindowThreadProcessId(foreground, None) if foreground else 0
        if target_thread:
            user32.AttachThreadInput(current_thread, target_thread, True)
        if foreground_thread:
            user32.AttachThreadInput(current_thread, foreground_thread, True)
        user32.ShowWindow(hwnd, SW_RESTORE)
        user32.BringWindowToTop(hwnd)
        user32.SetForegroundWindow(hwnd)
        user32.SetActiveWindow(hwnd)
        user32.SetFocus(hwnd)
        try:
            user32.SwitchToThisWindow(hwnd, True)
        except Exception:  # noqa: BLE001
            pass
        if foreground_thread:
            user32.AttachThreadInput(current_thread, foreground_thread, False)
        if target_thread:
            user32.AttachThreadInput(current_thread, target_thread, False)
        return user32.GetForegroundWindow() == hwnd
    except Exception:  # noqa: BLE001
        return False


class BatchImportRunner:
    """Best-effort CDP runner for the tax bureau batch-import page.

    The operator must already be logged in and positioned inside the tax site.
    This runner avoids opening a new controlled browser; it attaches to the
    already opened Edge instance through CDP, then uploads the generated Excel.
    """

    def __init__(
        self,
        *,
        template_path: str | Path,
        cdp_endpoint: str = "http://127.0.0.1:9222",
        status_hook: StatusHook | None = None,
    ) -> None:
        self.template_path = Path(template_path)
        self.cdp_endpoint = cdp_endpoint
        self.status_hook = status_hook
        self.logs: list[str] = []
        self.expected_serials = _read_template_serials(self.template_path)

    def run(self) -> BatchRunResult:
        try:
            from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
            from playwright.sync_api import sync_playwright
        except Exception as exc:  # noqa: BLE001
            return self._error("init", f"Playwright 不可用: {type(exc).__name__}: {exc}")

        if not self.template_path.exists():
            return self._error("init", f"模板文件不存在: {self.template_path}")

        with sync_playwright() as p:
            browser = p.chromium.connect_over_cdp(self.cdp_endpoint)
            try:
                page = self._select_tax_page(browser)
                self._log("attach", f"已连接税局页面: {_page_label(page)}")
                subject = _extract_tax_subject(page)
                if subject:
                    self._log("subject", f"当前税局页面识别主体: {subject}")
                if self.expected_serials:
                    self._log("template", "本次模板流水号: " + "、".join(self.expected_serials))
                page = self._open_batch_import_page(page)
                self._upload_template(page, PlaywrightTimeoutError)
                import_state = self._wait_for_import_result(page)
                if import_state == "failed":
                    failure_path, failure_report = self._download_failure_details(page, PlaywrightTimeoutError)
                    if failure_path:
                        self._log("failed", "税局导入失败，已下载并解析失败明细。")
                    else:
                        self._log("failed", "税局导入失败，未能自动下载失败明细；请手动下载后回传工作台。")
                    return BatchRunResult(
                        status="failed",
                        current_step="failed",
                        logs=self.logs,
                        downloaded_failure_path=str(failure_path) if failure_path else "",
                        failure_report=failure_report,
                    )
                preview_clicked = self._try_preview_invoice(page, PlaywrightTimeoutError)
                self._log("done", "批量导入流程已执行；请在税局页面确认预览和最终开具。")
                return BatchRunResult(
                    status="done",
                    current_step="done",
                    logs=self.logs,
                    preview_clicked=preview_clicked,
                )
            except Exception as exc:  # noqa: BLE001
                self._log("error", f"{type(exc).__name__}: {exc}")
                return BatchRunResult(status="error", current_step="error", logs=self.logs, error=str(exc))
            finally:
                browser.close()

    def _select_tax_page(self, browser):
        pages = [page for context in browser.contexts for page in context.pages]
        if not pages:
            raise RuntimeError("未找到可附着的 Edge 标签页。请先用 CDP Edge 登录电子税务局。")

        scored_pages = []
        for page in pages:
            title = _safe_title(page)
            text = _safe_body_text(page)
            score = _tax_page_score(page.url, title, text)
            scored_pages.append((score, page, title))
        scored_pages.sort(key=lambda item: item[0], reverse=True)
        best_score, best_page, best_title = scored_pages[0]
        if best_score > 0:
            self._log("attach", f"已发现 {len(scored_pages)} 个浏览器页面，优先选择税局业务页: score={best_score} title={best_title}")
            return best_page
        return pages[0]

    def _open_batch_import_page(self, page):
        self._log("navigate", "准备进入批量开票页面。")
        if _is_batch_import_page(page):
            self._log("navigate", "当前页面已是批量导入页。")
            return page

        if _looks_like_tax_portal_home(page):
            business_page = self._open_invoice_business_from_portal(page)
            if business_page is not page:
                page = business_page
                self._log("navigate", f"已进入发票业务页面: {_page_label(page)}")
                subject = _extract_tax_subject(page)
                if subject:
                    self._log("subject", f"发票业务页识别主体: {subject}")
                if _is_batch_import_page(page):
                    self._log("navigate", "发票业务新窗口已在批量导入页。")
                    return page

        origin = _origin_from_url(page.url)
        direct_url = f"{origin}/blue-invoice-makeout/invoice-batch"
        self._log("navigate", f"尝试直接进入批量导入页: {direct_url}")
        page.goto(direct_url, wait_until="domcontentloaded", timeout=30000)
        page.wait_for_timeout(2500)
        if _is_batch_import_page(page):
            self._log("navigate", "已进入批量导入页。")
            return page

        # Some tax-bureau deployments block direct routing until the makeout page
        # initializes. Fall back to the visible menu entry, but do not treat the
        # blue-invoice home page itself as the batch page just because it contains
        # the words “批量开票”.
        makeout_url = f"{origin}/blue-invoice-makeout"
        self._log("navigate", "直接进入未确认成功，回到蓝字发票开具页查找批量入口。")
        page.goto(makeout_url, wait_until="domcontentloaded", timeout=30000)
        page.wait_for_timeout(2000)
        clicked = _click_first_visible_text(page, ("批量开票", "批量导入", "导入开票"), timeout=5000)
        if clicked:
            self._log("navigate", f"已点击入口: {clicked}")
            page.wait_for_timeout(2500)
            if _is_batch_import_page(page):
                self._log("navigate", "已确认进入批量导入页。")
                return page

        raise RuntimeError(
            "未进入税局批量导入页。请手动打开“批量开票/批量导入”页面，"
            "确认地址包含 /blue-invoice-makeout/invoice-batch 后再重试。"
            f" 当前地址: {page.url}"
        )

    def _open_invoice_business_from_portal(self, page):
        before_pages = set(page.context.pages)
        self._log("navigate", "当前在税局首页，尝试点击“热门服务 / 发票业务”打开发票业务新窗口。")
        try:
            with page.expect_popup(timeout=8000) as popup_info:
                clicked = _click_first_visible_text(page, ("发票业务",), timeout=5000)
                if not clicked:
                    raise RuntimeError("未找到首页“发票业务”入口。")
            popup = popup_info.value
            popup.wait_for_load_state("domcontentloaded", timeout=30000)
            popup.wait_for_timeout(2500)
            return popup
        except Exception as exc:  # noqa: BLE001
            self._log("navigate", f"未通过 popup 事件捕获发票业务窗口，改为扫描浏览器页面: {type(exc).__name__}: {exc}")
            try:
                page.wait_for_timeout(2500)
            except Exception:  # noqa: BLE001
                pass
            candidates = [candidate for candidate in page.context.pages if candidate not in before_pages]
            candidates.extend(page.context.pages)
            scored = [(_tax_page_score(candidate.url, _safe_title(candidate), _safe_body_text(candidate)), candidate) for candidate in candidates]
            scored.sort(key=lambda item: item[0], reverse=True)
            for score, candidate in scored:
                if score >= 60:
                    return candidate
            return page

    def _upload_template(self, page, timeout_error_cls) -> None:
        if not _is_batch_import_page(page):
            raise RuntimeError(
                "当前页面不是批量导入页，已停止上传以避免误点。"
                f" 请先进入 /blue-invoice-makeout/invoice-batch。当前地址: {page.url}"
            )

        self._log("upload", f"准备上传模板: {self.template_path.name}")
        file_input = page.locator("input[type=file]").first
        if file_input.count():
            file_input.set_input_files(str(self.template_path))
        else:
            with page.expect_file_chooser(timeout=8000) as chooser_info:
                clicked = _click_first_visible_text(
                    page,
                    ("选择文件", "选择模板", "上传文件", "点击上传", "重新选择", "选择"),
                    timeout=5000,
                )
                if not clicked:
                    raise RuntimeError("未找到可见的文件选择按钮。请确认已停留在税局批量导入页。")
                self._log("upload", f"已点击文件选择按钮: {clicked}")
            chooser_info.value.set_files(str(self.template_path))
        page.wait_for_timeout(1000)
        clicked_submit = _click_first_visible_text(page, ("批量导入", "开始导入", "确认导入", "导入", "上传"), timeout=5000)
        if clicked_submit:
            self._log("upload", f"已点击提交按钮: {clicked_submit}")
        else:
            self._log("upload", "已设置模板文件，但未自动找到提交按钮；请在税局页面手动点击批量导入。")
        page.wait_for_timeout(3000)

    def _wait_for_import_result(self, page) -> str:
        self._log("wait_result", "等待税局处理导入结果。")
        success_tokens = ("导入完成", "处理成功", "预览发票", "发票价税合计")
        failure_tokens = ("下载失败明细", "处理失败", "导入失败", "失败明细")
        for _ in range(90):
            page.wait_for_timeout(1000)
            text = _safe_body_text(page)
            if any(token in text for token in failure_tokens):
                self._log("wait_result", "检测到导入失败信号。")
                return "failed"
            if any(token in text for token in success_tokens) and self._current_serial_visible(text):
                self._log("wait_result", "检测到本次流水号的导入成功信号。")
                return "success"
            if any(token in text for token in success_tokens) and self.expected_serials:
                self._log("wait_result", "页面存在成功/预览字样，但未出现本次流水号，继续等待以避免点击旧记录。")
        self._log("wait_result", "未在 90 秒内识别明确成功/失败信号，继续尝试预览。")
        return "unknown"

    def _download_failure_details(self, page, timeout_error_cls) -> tuple[Path | None, dict | None]:
        DOWNLOAD_ROOT.mkdir(parents=True, exist_ok=True)
        for text in ("下载失败明细", "下载失败明细模板", "失败明细"):
            target = page.get_by_text(text, exact=False).first
            if not target.count():
                continue
            try:
                with page.expect_download(timeout=10000) as download_info:
                    target.click(timeout=5000)
                download = download_info.value
                suggested = download.suggested_filename or "tax_import_failure.xlsx"
                output_path = DOWNLOAD_ROOT / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{suggested}"
                download.save_as(str(output_path))
                self._log("download_failure", f"已下载失败明细: {output_path}")
                report = build_failure_report(output_path)
                failure_serials = {
                    record.get("serial_no", "")
                    for record in report.get("records", [])
                    if record.get("serial_no")
                }
                matched_serials = sorted(set(self.expected_serials) & failure_serials)
                if matched_serials:
                    self._log("download_failure", "失败明细确认命中本次流水号: " + "、".join(matched_serials))
                elif self.expected_serials and failure_serials:
                    self._log(
                        "download_failure",
                        "失败明细流水号与本次模板不一致，请人工确认页面是否残留旧导入状态。",
                    )
                for record in report.get("records", []):
                    field = record.get("field_name") or "未识别字段"
                    failure_type = record.get("failure_type") or "未识别类型"
                    reason = record.get("reason") or ""
                    self._log("failure_report", f"{record.get('serial_no', '')} | {field} | {failure_type} | {reason}")
                return output_path, report
            except timeout_error_cls:
                continue
        self._log("download_failure", "未能自动下载失败明细；请在税局页面手动下载后回传工作台。")
        return None, None

    def _try_preview_invoice(self, page, timeout_error_cls) -> bool:
        if self.expected_serials:
            for attempt in range(1, 9):
                for serial in self.expected_serials:
                    preview = _preview_locator_for_serial(page, serial)
                    if preview is not None:
                        try:
                            preview.click(timeout=5000)
                            page.wait_for_timeout(1500)
                            self._log("preview", f"已点击本次流水号 {serial} 的预览发票。")
                            return True
                        except timeout_error_cls:
                            pass

                    clicked = _click_preview_for_serial_by_dom(page, serial)
                    if clicked:
                        page.wait_for_timeout(1500)
                        self._log("preview", f"已通过流水号邻近定位点击预览发票: {serial}")
                        return True

                if attempt in (1, 4):
                    self._log("preview", "本次流水号已出现，但预览按钮尚未完成行内定位，继续短暂重试。")
                page.wait_for_timeout(750)
            self._log("preview", "未找到本次流水号对应的预览发票按钮；不会点击页面上的旧预览记录。")
            return False

        for text in ("预览发票", "预览"):
            target = page.get_by_text(text, exact=False).first
            if not target.count():
                continue
            try:
                target.click(timeout=5000)
                page.wait_for_timeout(1500)
                self._log("preview", "已点击预览发票。")
                return True
            except timeout_error_cls:
                continue
        self._log("preview", "未自动找到预览发票按钮；请在税局页面手动查看结果。")
        return False

    def _error(self, step: str, message: str) -> BatchRunResult:
        self._log(step, message)
        return BatchRunResult(status="error", current_step=step, logs=self.logs, error=message)

    def _log(self, step: str, message: str) -> None:
        line = f"[{datetime.now().strftime('%H:%M:%S')}] {step}: {message}"
        self.logs.append(line)
        if self.status_hook:
            self.status_hook(step, line)

    def _current_serial_visible(self, page_text: str) -> bool:
        if not self.expected_serials:
            return True
        return any(serial and serial in page_text for serial in self.expected_serials)


def _page_label(page) -> str:
    return f"{_safe_title(page) or '无标题'} | {getattr(page, 'url', '')}"


def _safe_title(page) -> str:
    try:
        return page.title()
    except Exception:  # noqa: BLE001
        return ""


def _safe_body_text(page) -> str:
    try:
        return page.locator("body").inner_text(timeout=2000)
    except Exception:  # noqa: BLE001
        return ""


def _tax_page_score(url: str, title: str = "", text: str = "") -> int:
    combined = f"{title}\n{text}"
    score = 0
    if "chinatax.gov.cn" in url:
        score += 20
    if "电子税务局" in title or "全国统一规范电子税务局" in combined:
        score += 15
    path = urlparse(url).path.rstrip("/")
    if path.endswith("/blue-invoice-makeout/invoice-batch"):
        score += 170
    elif "/blue-invoice-makeout" in path:
        score += 80
    if "发票业务" in combined:
        score += 25
    if "蓝字发票开具" in combined or "开票信息维护" in combined:
        score += 30
    if "热门服务" in combined and "我的待办" in combined:
        score += 20
    if "发票查询统计" in combined or "批量导入" in combined or "批量开票" in combined:
        score += 25
    if _is_tax_login_page_text(combined):
        score -= 40
    return score



def _looks_like_tax_portal_home(page) -> bool:
    text = _safe_body_text(page)
    title = _safe_title(page)
    return _tax_page_score(page.url, title, text) > 0 and "热门服务" in text and "发票业务" in text and "/blue-invoice-makeout" not in urlparse(page.url).path



def _is_tax_login_page_text(text: str) -> bool:
    return "登录" in text and ("密码" in text or "验证码" in text) and "热门服务" not in text



def _extract_tax_subject(page) -> str:
    text = _safe_body_text(page)
    if not text:
        return ""
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    tax_id_pattern = re.compile(r"[0-9A-Z]{15,20}")
    for index, line in enumerate(lines):
        compact = line.replace(" ", "").upper()
        if tax_id_pattern.fullmatch(compact) or tax_id_pattern.search(compact):
            tax_id = tax_id_pattern.search(compact).group(0)
            name = ""
            for candidate in reversed(lines[max(0, index - 3):index]):
                if _looks_like_taxpayer_name(candidate):
                    name = candidate
                    break
            return f"{name} / {tax_id}" if name else tax_id
    for line in lines[:80]:
        if _looks_like_taxpayer_name(line):
            return line
    return ""



def _looks_like_taxpayer_name(value: str) -> bool:
    cleaned = value.strip()
    if not cleaned or len(cleaned) < 4:
        return False
    if any(token in cleaned for token in ["全国统一", "电子税务局", "发票业务", "热门服务", "我的待办", "请输入", "操作指引"]):
        return False
    return bool(re.search(r"(有限|公司|个体工商户|商贸|科技|餐饮|饭店|工作室|学院|传媒)", cleaned))



def _origin_from_url(url: str) -> str:
    parsed = urlparse(url)
    if not parsed.scheme or not parsed.netloc:
        raise RuntimeError("当前标签页不是有效税局 URL。")
    return f"{parsed.scheme}://{parsed.netloc}"


def _is_batch_import_page(page) -> bool:
    """Return True only for the actual batch-import workspace, not its menu entry.

    The blue-invoice makeout home page can contain the text “批量开票”, which used
    to cause a false positive. Prefer the routed URL and otherwise require upload
    controls plus batch-import wording.
    """

    try:
        path = urlparse(page.url).path.rstrip("/")
        if path.endswith("/blue-invoice-makeout/invoice-batch"):
            return True
    except Exception:  # noqa: BLE001
        pass

    text = _safe_body_text(page)
    if not text or "蓝字发票开具" in _safe_title(page) and "批量导入" not in text:
        return False
    has_upload_control = False
    try:
        has_upload_control = page.locator("input[type=file]").count() > 0
    except Exception:  # noqa: BLE001
        has_upload_control = False
    upload_tokens = ("选择文件", "上传文件", "点击上传", "批量导入")
    return has_upload_control and any(token in text for token in upload_tokens)


def _click_first_visible_text(page, texts: tuple[str, ...], *, timeout: int = 5000) -> str:
    """Click the first visible interactive element matching the provided texts.

    Avoids `get_by_text("选择").first`, which can match hidden dialog titles such
    as “选择发票票种” instead of the upload button.
    """

    selectors = (
        "button",
        "a",
        "[role=button]",
        "label",
        ".t-button",
        ".el-button",
        ".ant-btn",
        ".ivu-btn",
        ".arco-btn",
        ".semi-button",
        ".t-upload__trigger",
        ".el-upload",
        ".ant-upload",
    )
    deadline = datetime.now().timestamp() + max(timeout, 0) / 1000
    last_error: Exception | None = None
    while datetime.now().timestamp() <= deadline:
        for text in texts:
            for selector in selectors:
                try:
                    locator = page.locator(selector).filter(has_text=text)
                    count = min(locator.count(), 20)
                except Exception as exc:  # noqa: BLE001
                    last_error = exc
                    continue
                for index in range(count):
                    candidate = locator.nth(index)
                    try:
                        if not candidate.is_visible(timeout=150):
                            continue
                        if hasattr(candidate, "is_enabled") and not candidate.is_enabled(timeout=150):
                            continue
                        candidate.click(timeout=1000)
                        return text
                    except Exception as exc:  # noqa: BLE001
                        last_error = exc
                        continue
        page.wait_for_timeout(200)
    if last_error:
        return ""
    return ""


def _read_template_serials(template_path: Path) -> tuple[str, ...]:
    try:
        workbook = load_workbook(template_path, read_only=True, data_only=True)
        sheet = workbook["1-发票基本信息"]
        header_row = [str(cell.value or "").strip() for cell in sheet[3]]
        serial_column = header_row.index("发票流水号") + 1
        serials: list[str] = []
        for row_index in range(4, sheet.max_row + 1):
            value = sheet.cell(row=row_index, column=serial_column).value
            serial = str(value or "").strip()
            if serial:
                serials.append(serial)
        workbook.close()
        return tuple(dict.fromkeys(serials))
    except Exception:  # noqa: BLE001
        fallback = template_path.stem
        if "_batch_import" in fallback:
            fallback = fallback.split("_batch_import", 1)[0]
        return (fallback,) if fallback else ()


def _preview_locator_for_serial(page, serial: str):
    try:
        row_selectors = (
            "tr",
            "[role='row']",
            ".ant-table-row",
            ".el-table__row",
            ".vxe-body--row",
            ".ivu-table-row",
        )
        for row_selector in row_selectors:
            rows = page.locator(row_selector).filter(has_text=serial)
            for row_index in range(rows.count()):
                row = rows.nth(row_index)
                for text in ("预览发票", "预览"):
                    preview = row.get_by_text(text, exact=False).first
                    if preview.count():
                        return preview
    except Exception:  # noqa: BLE001
        return None
    return None


def _click_preview_for_serial_by_dom(page, serial: str) -> bool:
    """Click preview only when the clickable element is inside a serial-scoped row/ancestor."""

    try:
        result = page.evaluate(
            """
            (serial) => {
              const previewPattern = /预览发票|预览/;
              const rowSelectors = [
                'tr',
                '[role="row"]',
                '.ant-table-row',
                '.el-table__row',
                '.vxe-body--row',
                '.ivu-table-row',
                '.semi-table-row',
                '.arco-table-tr'
              ];

              const norm = (value) => (value || '').replace(/\\s+/g, ' ').trim();
              const isVisible = (element) => {
                if (!element || element.closest('[aria-hidden="true"]')) {
                  return false;
                }
                const style = window.getComputedStyle(element);
                if (style.display === 'none' || style.visibility === 'hidden' || style.opacity === '0') {
                  return false;
                }
                const rect = element.getBoundingClientRect();
                return rect.width > 0 && rect.height > 0;
              };
              const isDisabled = (element) => {
                return element.disabled === true
                  || element.getAttribute('aria-disabled') === 'true'
                  || element.className.toString().includes('disabled');
              };
              const clickTarget = (root) => {
                const candidates = Array.from(root.querySelectorAll('a, button, [role="button"], span, div'))
                  .filter((element) => previewPattern.test(norm(element.innerText || element.textContent)))
                  .filter((element) => isVisible(element) && !isDisabled(element));
                const target = candidates.find((element) => norm(element.innerText || element.textContent).includes('预览发票'))
                  || candidates[0];
                if (!target) {
                  return false;
                }
                target.scrollIntoView({ block: 'center', inline: 'center' });
                target.dispatchEvent(new MouseEvent('mouseover', { bubbles: true, cancelable: true, view: window }));
                target.dispatchEvent(new MouseEvent('mousedown', { bubbles: true, cancelable: true, view: window }));
                target.dispatchEvent(new MouseEvent('mouseup', { bubbles: true, cancelable: true, view: window }));
                target.click();
                return true;
              };

              for (const selector of rowSelectors) {
                for (const row of Array.from(document.querySelectorAll(selector))) {
                  const rowText = norm(row.innerText || row.textContent);
                  if (rowText.includes(serial) && previewPattern.test(rowText) && clickTarget(row)) {
                    return true;
                  }
                }
              }

              const previewElements = Array.from(document.querySelectorAll('a, button, [role="button"], span, div'))
                .filter((element) => previewPattern.test(norm(element.innerText || element.textContent)))
                .filter((element) => isVisible(element) && !isDisabled(element));
              for (const element of previewElements) {
                let ancestor = element.parentElement;
                for (let depth = 0; ancestor && depth < 8; depth += 1, ancestor = ancestor.parentElement) {
                  const ancestorText = norm(ancestor.innerText || ancestor.textContent);
                  if (ancestorText.length <= 2000 && ancestorText.includes(serial)) {
                    element.scrollIntoView({ block: 'center', inline: 'center' });
                    element.click();
                    return true;
                  }
                }
              }
              return false;
            }
            """,
            serial,
        )
        return bool(result)
    except Exception:  # noqa: BLE001
        return False
